"""Tests for the optional debug Excel export."""

from __future__ import annotations

from datetime import timedelta
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from lecture_analyzer.core.config import PipelineConfig
from lecture_analyzer.core.models import (
    AudioSource,
    InputSource,
    LectureSession,
    MergedTranscript,
    MergedTranscriptUnit,
    Segment,
    TranscriptChunk,
)
from lecture_analyzer.core.pipeline import LectureProcessingPipeline
from lecture_analyzer.core.types import MediaType
from lecture_analyzer.output.debug_excel_exporter import (
    HEADERS,
    QA_HEADERS,
    REVIEW_CANDIDATE_HEADERS,
    SENTENCE_HEADERS,
    SENTENCE_METRIC_HEADERS,
    export_run_to_excel,
)


class DebugExcelExporterTests(unittest.TestCase):
    """Exercise the standalone Excel exporter and pipeline integration."""

    def test_export_run_to_excel_writes_readable_debug_workbook(self) -> None:
        """The exporter should create a formatted worksheet with review cues."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            json_path = temp_path / "session_structural.json"
            excel_path = temp_path / "session_structural.xlsx"

            payload = {
                "session_metadata": {
                    "schema_version": "0.6.0",
                    "session_id": "session_001",
                    "language_codes": ["it", "en"],
                    "supports_mixed_language": True,
                    "processing_status": "ready",
                    "metadata": {
                        "pipeline_execution_mode": "normal",
                        "pipeline_run_profile_label": "mixed_run",
                        "transcription_cache_enabled": True,
                        "transcription_cache_lookup_performed": True,
                        "transcription_cache_hit": True,
                        "transcription_recomputed": False,
                        "transcription_forced_recompute": False,
                    },
                },
                "input_sources": [
                    {
                        "source_id": "source_001",
                        "original_filename": "lesson.wav",
                        "original_path": str(temp_path / "lesson.wav"),
                        "duration_seconds": 120.5,
                    },
                ],
                "audio_sources": [
                    {
                        "audio_source_id": "audio_source_001",
                        "duration_seconds": 119.8,
                    },
                ],
                "transcript": {
                    "full_text": "Hello class Maybe speaker two Unknown voice",
                    "chunk_count": 3,
                    "aligned_source_count": 1,
                    "aligned_segment_count": 2,
                    "aligned_word_count": 7,
                    "diarization_segment_count": 4,
                    "speaker_count": 2,
                    "utterance_count": 3,
                    "utterance_with_speaker_count": 2,
                    "utterance_uncertain_speaker_count": 1,
                    "sentence_count": 2,
                    "sentence_with_speaker_count": 1,
                    "merged_unit_count": 3,
                    "detected_languages": ["it", "en"],
                },
                "utterances": [
                    {
                        "utterance_id": "utterance_001",
                        "audio_source_id": "audio_source_001",
                        "text": "Hello class",
                        "start_seconds": 0.0,
                        "end_seconds": 1.5,
                        "session_start_seconds": 10.0,
                        "session_end_seconds": 11.5,
                        "transcript_chunk_id": "chunk_001",
                        "source_word_ids": ["w1", "w2"],
                        "speaker_id": "speaker_1",
                        "speaker_attribution_status": "assigned",
                        "speaker_is_uncertain": False,
                        "metadata": {
                            "build_strategy": "segment_boundary_plus_gap",
                            "speaker_attribution": {
                                "status": "assigned",
                                "audio_quality": {
                                    "is_degraded": False,
                                    "degraded_reasons": [],
                                },
                            },
                        },
                    },
                    {
                        "utterance_id": "utterance_002",
                        "audio_source_id": "audio_source_001",
                        "text": "Maybe speaker two",
                        "start_seconds": 1.6,
                        "end_seconds": 3.0,
                        "session_start_seconds": 11.6,
                        "session_end_seconds": 13.0,
                        "transcript_chunk_id": "chunk_002",
                        "speaker_id": "speaker_2",
                        "speaker_attribution_status": "assigned",
                        "speaker_is_uncertain": True,
                        "metadata": {
                            "speaker_attribution": {
                                "status": "assigned",
                                "reason": "ambiguous_overlap",
                                "audio_quality": {
                                    "is_degraded": True,
                                    "degraded_reasons": [
                                        "high_zero_crossing_rate",
                                    ],
                                },
                            },
                        },
                    },
                    {
                        "utterance_id": "utterance_003",
                        "audio_source_id": "audio_source_001",
                        "text": "Unknown voice",
                        "start_seconds": 3.2,
                        "end_seconds": 4.0,
                        "transcript_chunk_id": "chunk_missing",
                        "speaker_id": None,
                        "speaker_attribution_status": "unassigned",
                        "speaker_is_uncertain": False,
                        "metadata": {
                            "speaker_attribution": {
                                "status": "unassigned",
                                "reason": "no_overlap",
                            },
                        },
                    },
                ],
                "sentences": [
                    {
                        "sentence_id": "audio_source_001_sentence_0001",
                        "audio_source_id": "audio_source_001",
                        "text": "Hello class Maybe speaker two",
                        "start_seconds": 0.0,
                        "end_seconds": 3.0,
                        "session_start_seconds": 10.0,
                        "session_end_seconds": 13.0,
                        "source_utterance_ids": ["utterance_001", "utterance_002"],
                        "source_utterance_start_index": 1,
                        "source_utterance_end_index": 2,
                        "detected_language": "it",
                        "speaker_id": None,
                        "speaker_resolution_status": "uncertain",
                        "speaker_confidence_label": "low",
                        "speaker_stability_label": "uncertain",
                        "speaker_evidence_summary": (
                            "assigned=1; uncertain=1; unassigned=0; speakers=speaker_1"
                        ),
                        "merge_safety_label": "risky",
                        "semantic_quality_label": "borderline",
                        "length_bucket": "normal",
                        "duration_bucket": "normal",
                        "review_priority": "high",
                        "sentence_review_flags": [
                            "uncertain_source",
                            "multi_utterance",
                            "merge_risky",
                        ],
                        "metadata": {
                            "source_utterance_count": 2,
                            "speaker_boundary_respected": False,
                            "has_uncertain_source": True,
                            "has_unassigned_source": False,
                            "has_speaker_change_inside": True,
                            "is_multi_utterance": True,
                            "is_semantic_fragment": False,
                            "is_semantic_run_on": False,
                            "is_merge_risky": True,
                            "semantic_quality_label": "borderline",
                        },
                    },
                    {
                        "sentence_id": "audio_source_001_sentence_0002",
                        "audio_source_id": "audio_source_001",
                        "text": "Unknown voice",
                        "start_seconds": 3.2,
                        "end_seconds": 4.0,
                        "source_utterance_ids": ["utterance_003"],
                        "source_utterance_start_index": 3,
                        "source_utterance_end_index": 3,
                        "detected_language": "en",
                        "speaker_id": "speaker_3",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "speaker_evidence_summary": (
                            "assigned=1; uncertain=0; unassigned=0; speakers=speaker_3"
                        ),
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "length_bucket": "short",
                        "duration_bucket": "short",
                        "review_priority": "low",
                        "sentence_review_flags": ["segment_missing"],
                        "metadata": {
                            "source_utterance_count": 1,
                            "speaker_boundary_respected": True,
                            "has_uncertain_source": False,
                            "has_unassigned_source": False,
                            "has_speaker_change_inside": False,
                            "is_multi_utterance": False,
                            "is_semantic_fragment": False,
                            "is_semantic_run_on": False,
                            "is_merge_risky": False,
                            "semantic_quality_label": "good",
                        },
                    },
                ],
                "segments": [
                    {
                        "segment_id": "segment_0001",
                        "transcript_chunk_ids": ["chunk_001"],
                        "sentence_ids": ["audio_source_001_sentence_0001"],
                    },
                    {
                        "segment_id": "segment_0002",
                        "transcript_chunk_ids": ["chunk_002"],
                        "sentence_ids": [],
                    },
                ],
                "qa_candidates": [
                    {
                        "qa_candidate_id": "qa_001",
                        "question_text": "What is a graph?",
                        "answer_text": "A graph is a set of nodes and edges.",
                        "context_text": (
                            "The speakers are defining what a graph is before "
                            "moving into the formal answer."
                        ),
                        "question_unit_ids": ["unit_0001"],
                        "answer_unit_ids": ["unit_0002"],
                        "question_sentence_ids": ["audio_source_001_sentence_0001"],
                        "answer_sentence_ids": ["audio_source_001_sentence_0002"],
                        "context_sentence_ids": ["audio_source_001_sentence_0001"],
                        "question_source_utterance_ids": ["utterance_001"],
                        "answer_source_utterance_ids": ["utterance_002"],
                        "context_source_utterance_ids": ["utterance_001"],
                        "question_segment_id": "segment_0001",
                        "answer_segment_id": "segment_0002",
                        "context_strategy": "local_topic_window",
                        "context_confidence": "medium",
                        "start_seconds": 10.0,
                        "end_seconds": 13.0,
                        "question_timing": {
                            "start_seconds": 10.0,
                            "end_seconds": 11.0,
                            "audio_source_id": "audio_source_001",
                            "session_start_seconds": 10.0,
                            "session_end_seconds": 11.0,
                        },
                        "answer_timing": {
                            "start_seconds": 11.6,
                            "end_seconds": 13.0,
                            "audio_source_id": "audio_source_001",
                            "session_start_seconds": 11.6,
                            "session_end_seconds": 13.0,
                        },
                        "question_speaker_role": "unknown",
                        "answer_speaker_role": "teacher",
                        "source_segment_ids": ["segment_0001", "segment_0002"],
                        "confidence": 0.68,
                        "confidence_label": "medium",
                        "confidence_score": 0.68,
                        "question_type": "direct_question",
                        "didactic_question_score": 0.79,
                        "answer_is_question": False,
                        "review_flags": ["medium_confidence"],
                        "reason_codes": [
                            "question_mark",
                            "answer_in_next_sentence",
                            "next_segment_support",
                            "speaker_turn_support",
                        ],
                        "metadata": {
                            "input_layer": "sentences",
                            "question_debug": {
                                "question_score": 0.74,
                                "question_unit_index": 0,
                                "input_layer": "sentences",
                            },
                            "answer_debug": {
                                "answer_score": 0.61,
                                "answer_distance_units": 1,
                                "gap_seconds": 0.6,
                                "answer_unit_ids": ["unit_0002"],
                                "answer_sentence_ids": [
                                    "audio_source_001_sentence_0002",
                                ],
                                "answer_source_utterance_ids": ["utterance_002"],
                                "search_stop_reason": "window_exhausted",
                                "search_signals": {
                                    "candidate_channel": "semantic_retrieval",
                                    "semantic_similarity": 0.93,
                                },
                                "ranking_debug": {
                                    "requested_ranking_strategy": "semantic_reranker",
                                    "effective_ranking_strategy": "semantic_reranker",
                                    "ranking_strategy": "semantic_reranker",
                                    "semantic_reranking_model_name": "BAAI/bge-reranker-v2-m3",
                                    "semantic_reranking_backend": "huggingface_transformers",
                                    "semantic_relevance_score": 0.97,
                                    "combined_score": 0.86,
                                    "rank_position": 1,
                                },
                            },
                            "pairing_debug": {
                                "segment_relation": "next_segment",
                                "question_segment_ids": ["segment_0001"],
                                "answer_segment_ids": ["segment_0002"],
                                "requested_search_strategy": "semantic_retrieval",
                                "search_strategy": "semantic_retrieval",
                                "effective_search_strategy": "semantic_retrieval",
                                "requested_ranking_strategy": "semantic_reranker",
                                "ranking_strategy": "semantic_reranker",
                                "effective_ranking_strategy": "semantic_reranker",
                                "search_stop_reason": "window_exhausted",
                                "search_fallback_reason": "",
                                "search_backend_error": "",
                                "ranking_fallback_reason": "",
                                "semantic_backend_status": "available",
                                "semantic_model_name": "intfloat/multilingual-e5-base",
                                "semantic_backend": "sentence_transformers",
                                "semantic_reranking_model_name": "BAAI/bge-reranker-v2-m3",
                                "semantic_reranking_backend": "huggingface_transformers",
                                "semantic_relevance_score": 0.97,
                                "speaker_influence": "boost",
                                "deferred_answer_search_used": True,
                            },
                            "grounding_debug": {
                                "question_sentence_ids": [
                                    "audio_source_001_sentence_0001",
                                ],
                                "answer_sentence_ids": [
                                    "audio_source_001_sentence_0002",
                                ],
                                "question_source_utterance_ids": ["utterance_001"],
                                "answer_source_utterance_ids": ["utterance_002"],
                                "question_timing_source": "utterances",
                                "answer_timing_source": "utterances",
                                "question_grounded_utterance_ids": ["utterance_001"],
                                "answer_grounded_utterance_ids": ["utterance_002"],
                            },
                            "context_debug": {
                                "context_text": (
                                    "The speakers are defining what a graph is "
                                    "before moving into the formal answer."
                                ),
                                "context_strategy": "local_topic_window",
                                "context_confidence": "medium",
                                "context_sentence_ids": [
                                    "audio_source_001_sentence_0001",
                                ],
                                "context_source_utterance_ids": ["utterance_001"],
                            },
                            "confidence_debug": {
                                "question_score": 0.74,
                                "didactic_question_score": 0.79,
                                "answer_score": 0.61,
                                "question_weight": 0.55,
                                "answer_weight": 0.45,
                                "base_confidence": 0.6815,
                                "competing_question_penalty": 0.0,
                                "fallback_penalty": 0.0,
                                "answer_is_question_penalty": 0.0,
                                "final_confidence": 0.68,
                            },
                        },
                    },
                ],
                "speaker_role_estimates": [{"speaker_label": "speaker_1"}],
                "pipeline_timing": {
                    "summary": {
                        "pipeline_execution_mode": "normal",
                        "run_profile_label": "mixed_run",
                        "total_duration_seconds": 12.34,
                        "stage_count": 5,
                        "executed_stage_count": 4,
                        "completed_stage_count": 4,
                        "skipped_stage_count": 1,
                        "failed_stage_count": 0,
                        "reused_cache_stage_count": 1,
                        "reused_artifact_stage_count": 0,
                        "forced_recompute_stage_count": 0,
                        "any_cache_hit": True,
                        "any_artifact_reuse": False,
                        "full_recompute_requested": False,
                        "most_expensive_stage_name": "transcription",
                        "most_expensive_stage_duration_seconds": 5.67,
                    },
                    "stages": [
                        {
                            "stage_name": "transcription",
                            "status": "reused_from_cache",
                            "duration_seconds": 5.67,
                            "used_cache": True,
                            "used_existing_artifact": False,
                            "forced_recompute": False,
                            "note": None,
                            "metadata": {"backend": "fake"},
                        },
                        {
                            "stage_name": "qa_extraction",
                            "status": "executed",
                            "duration_seconds": 0.31,
                            "note": None,
                            "metadata": {"qa_candidate_count": 1},
                        },
                        {
                            "stage_name": "json_export",
                            "status": "executed",
                            "duration_seconds": 0.02,
                            "note": None,
                            "metadata": {"segmentation_mode": "structural"},
                        },
                    ],
                },
            }
            json_path.write_text(
                json.dumps(payload, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            export_run_to_excel(json_path, excel_path)

            workbook = load_workbook(excel_path)
            worksheet = workbook["Utterances"]
            sentences_worksheet = workbook["Sentences"]
            qa_worksheet = workbook["QACandidates"]
            review_candidates_worksheet = workbook["ReviewCandidates"]
            sentence_metrics_worksheet = workbook["SentenceMetrics"]
            summary_worksheet = workbook["run_summary"]

            self.assertEqual(tuple(cell.value for cell in worksheet[1]), HEADERS)
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet["A2"].value, "utterance_001")
            self.assertEqual(worksheet["B2"].value, timedelta(seconds=10))
            self.assertEqual(worksheet["B2"].number_format, "[h]:mm:ss.000")
            self.assertEqual(worksheet["D2"].value, timedelta(seconds=1.5))
            self.assertEqual(worksheet["M2"].value, "segment_0001")
            self.assertEqual(worksheet["N2"].value, "audio_source_001_sentence_0001")
            self.assertEqual(worksheet["O2"].value, 1)
            self.assertEqual(worksheet["R2"].value, 2)
            self.assertEqual(worksheet["F3"].value, "uncertain")
            self.assertEqual(worksheet["G3"].value, True)
            self.assertEqual(worksheet["K3"].value, True)
            self.assertEqual(worksheet["L3"].value, "medium")
            self.assertIn("uncertain_speaker", str(worksheet["S3"].value))
            self.assertIn("audio_quality_degraded", str(worksheet["S3"].value))
            self.assertIn("speaker_reason=ambiguous_overlap", str(worksheet["T3"].value))
            self.assertEqual(worksheet["J4"].value, True)
            self.assertEqual(worksheet["L4"].value, "high")
            self.assertIn("speaker_unassigned", str(worksheet["S4"].value))
            self.assertIn("segment_missing", str(worksheet["S4"].value))
            self.assertTrue(bool(worksheet["Q2"].alignment.wrap_text))
            self.assertEqual(worksheet["Q2"].fill.fill_type, "solid")
            self.assertEqual(worksheet["E2"].fill.fill_type, "solid")
            self.assertIsNone(worksheet["Q4"].fill.fill_type)
            self.assertEqual(worksheet["Q3"].font.underline, "single")
            self.assertTrue(bool(worksheet["Q3"].font.italic))
            self.assertTrue(bool(worksheet["Q4"].font.bold))
            self.assertEqual(worksheet["F3"].fill.fill_type, "solid")
            self.assertEqual(worksheet["S3"].fill.fill_type, "solid")
            self.assertEqual(
                tuple(cell.value for cell in sentences_worksheet[1]),
                SENTENCE_HEADERS,
            )
            self.assertEqual(sentences_worksheet.freeze_panes, "A2")
            sentence_header_index = {
                cell.value: index
                for index, cell in enumerate(sentences_worksheet[1], start=1)
            }
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["sentence_id"],
                ).value,
                "audio_source_001_sentence_0001",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["time_range"],
                ).value,
                "00:00:10.000 - 00:00:13.000",
            )
            self.assertIn(
                "QUESTION | qa_001",
                str(
                    sentences_worksheet.cell(
                        2,
                        sentence_header_index["qa_debug_summary"],
                    ).value,
                ),
            )
            self.assertIn(
                "answer: 0002",
                str(
                    sentences_worksheet.cell(
                        2,
                        sentence_header_index["qa_debug_summary"],
                    ).value,
                ),
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["human_comment"],
                ).value,
                None,
            )
            self.assertEqual(
                sentences_worksheet.cell(2, sentence_header_index["start"]).value,
                timedelta(seconds=10),
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["source_utterance_count"],
                ).value,
                2,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["source_utterance_span"],
                ).value,
                "1-2",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["is_multi_utterance"],
                ).value,
                True,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["has_uncertain_source"],
                ).value,
                True,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["has_speaker_change_inside"],
                ).value,
                True,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["is_merge_risky"],
                ).value,
                True,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["semantic_quality_label"],
                ).value,
                "borderline",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["speaker_stability_label"],
                ).value,
                "uncertain",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["merge_safety_label"],
                ).value,
                "risky",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["review_priority"],
                ).value,
                "high",
            )
            self.assertEqual(
                sentences_worksheet.cell(2, sentence_header_index["segment_id"]).value,
                "segment_0001",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["distinct_source_speaker_count"],
                ).value,
                2,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["source_speaker_ids"],
                ).value,
                "speaker_1, speaker_2",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["assigned_source_utterance_count"],
                ).value,
                1,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["uncertain_source_utterance_count"],
                ).value,
                1,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["unassigned_source_utterance_count"],
                ).value,
                0,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["speaker_resolution_status"],
                ).value,
                "uncertain",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["speaker_confidence_label"],
                ).value,
                "low",
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["has_provenance_overlap"],
                ).value,
                False,
            )
            self.assertIn(
                "source_utterance_count=2",
                str(
                    sentences_worksheet.cell(
                        2,
                        sentence_header_index["notes"],
                    ).value,
                ),
            )
            text_row2 = sentences_worksheet.cell(2, sentence_header_index["text"])
            text_row3 = sentences_worksheet.cell(3, sentence_header_index["text"])
            human_comment_row2 = sentences_worksheet.cell(
                2,
                sentence_header_index["human_comment"],
            )
            self.assertIsNone(text_row2.fill.fill_type)
            self.assertEqual(text_row2.font.underline, "single")
            self.assertTrue(bool(text_row2.font.bold))
            self.assertEqual(text_row2.border.top.style, "medium")
            self.assertEqual(text_row3.fill.fill_type, "solid")
            self.assertIsNone(text_row3.font.underline)
            self.assertTrue(bool(text_row2.alignment.wrap_text))
            self.assertTrue(bool(human_comment_row2.alignment.wrap_text))
            self.assertEqual(human_comment_row2.fill.fill_type, "solid")
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["review_priority"],
                ).fill.fill_type,
                "solid",
            )
            self.assertEqual(
                tuple(cell.value for cell in qa_worksheet[1]),
                QA_HEADERS,
            )
            self.assertEqual(qa_worksheet.freeze_panes, "I2")
            qa_header_index = {
                cell.value: index
                for index, cell in enumerate(qa_worksheet[1], start=1)
            }
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["qa_id"]).value,
                "qa_001",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["question_sentence_id"],
                ).value,
                "0001",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["answer_sentence_id"],
                ).value,
                "0002",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["context_sentence_id"],
                ).value,
                "0001",
            )
            self.assertIn(
                "formal answer",
                str(
                    qa_worksheet.cell(
                        2,
                        qa_header_index["context_text"],
                    ).value,
                ),
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["context_strategy"],
                ).value,
                "local_topic_window",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["context_confidence"],
                ).value,
                "medium",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["context_sentence_ids"],
                ).value,
                "audio_source_001_sentence_0001",
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["confidence"]).value,
                "medium (0.68)",
            )
            self.assertIn(
                "medium_confidence",
                str(
                    qa_worksheet.cell(
                        2,
                        qa_header_index["review_flag_summary"],
                    ).value,
                ),
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["id"]).value,
                "qa_001",
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["start"]).value,
                timedelta(seconds=10),
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["duration_s"]).value,
                timedelta(seconds=3),
            )
            self.assertAlmostEqual(
                float(qa_worksheet.cell(2, qa_header_index["confidence_score"]).value),
                0.68,
            )
            self.assertAlmostEqual(
                float(
                    qa_worksheet.cell(
                        2,
                        qa_header_index["didactic_question_score"],
                    ).value,
                ),
                0.79,
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["review_priority"]).value,
                "medium",
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["answer_is_question"]).value,
                False,
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["deferred_answer_search_used"],
                ).value,
                True,
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["input_layer"]).value,
                "sentences",
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["speaker_influence"]).value,
                "boost",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["requested_search_strategy"],
                ).value,
                "semantic_retrieval",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["effective_ranking_strategy"],
                ).value,
                "semantic_reranker",
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["search_model_name"]).value,
                "intfloat/multilingual-e5-base",
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["search_backend_error"]).value,
                None,
            )
            self.assertAlmostEqual(
                float(
                    qa_worksheet.cell(2, qa_header_index["semantic_similarity"]).value,
                ),
                0.93,
            )
            self.assertEqual(
                qa_worksheet.cell(2, qa_header_index["reranking_model_name"]).value,
                "BAAI/bge-reranker-v2-m3",
            )
            self.assertAlmostEqual(
                float(
                    qa_worksheet.cell(
                        2,
                        qa_header_index["semantic_relevance_score"],
                    ).value,
                ),
                0.97,
            )
            self.assertAlmostEqual(
                float(
                    qa_worksheet.cell(2, qa_header_index["combined_rank_score"]).value,
                ),
                0.86,
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["question_timing_source"],
                ).value,
                "utterances",
            )
            self.assertIn(
                "speaker_turn_support",
                str(qa_worksheet.cell(2, qa_header_index["reason_codes"]).value),
            )
            self.assertIn(
                "medium_confidence",
                str(qa_worksheet.cell(2, qa_header_index["review_flags"]).value),
            )
            self.assertIn(
                "base_confidence=0.68",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertIn(
                "effective_search=semantic_retrieval",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertIn(
                "semantic_relevance=0.97",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertIn(
                "didactic_question_score=0.79",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertIn(
                "context_strategy=local_topic_window",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertIn(
                "context_sentence_ids=audio_source_001_sentence_0001",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertIn(
                "deferred_answer_search=true",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertNotIn(
                "search_backend_error=",
                str(qa_worksheet.cell(2, qa_header_index["notes"]).value),
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["confidence_score"],
                ).number_format,
                "0.00",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["semantic_relevance_score"],
                ).number_format,
                "0.00",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["review_priority"],
                ).fill.fill_type,
                "solid",
            )
            self.assertTrue(
                bool(
                    qa_worksheet.cell(
                        2,
                        qa_header_index["question_text"],
                    ).alignment.wrap_text,
                ),
            )
            self.assertEqual(review_candidates_worksheet.freeze_panes, "A2")
            self.assertEqual(
                tuple(cell.value for cell in review_candidates_worksheet[1]),
                REVIEW_CANDIDATE_HEADERS,
            )
            review_header_index = {
                cell.value: index
                for index, cell in enumerate(review_candidates_worksheet[1], start=1)
            }
            self.assertEqual(
                review_candidates_worksheet.cell(
                    2,
                    review_header_index["kind"],
                ).value,
                "sentence",
            )
            self.assertEqual(
                review_candidates_worksheet.cell(
                    2,
                    review_header_index["review_priority"],
                ).value,
                "high",
            )
            speaker_review_row = next(
                row_index
                for row_index in range(2, review_candidates_worksheet.max_row + 1)
                if review_candidates_worksheet.cell(
                    row_index,
                    review_header_index["speaker"],
                ).value
            )
            self.assertEqual(
                review_candidates_worksheet.cell(
                    speaker_review_row,
                    review_header_index["text"],
                ).fill.fill_type,
                "solid",
            )
            if (
                review_candidates_worksheet.cell(
                    speaker_review_row,
                    review_header_index["review_priority"],
                ).value
                == "medium"
            ):
                self.assertTrue(
                    bool(
                        review_candidates_worksheet.cell(
                            speaker_review_row,
                            review_header_index["text"],
                        ).font.italic,
                    ),
                )
            if (
                review_candidates_worksheet.cell(
                    speaker_review_row,
                    review_header_index["review_priority"],
                ).value
                == "high"
            ):
                self.assertTrue(
                    bool(
                        review_candidates_worksheet.cell(
                            speaker_review_row,
                            review_header_index["text"],
                        ).font.bold,
                    ),
                )
            self.assertEqual(
                tuple(cell.value for cell in sentence_metrics_worksheet[1]),
                SENTENCE_METRIC_HEADERS,
            )
            self.assertEqual(sentence_metrics_worksheet.freeze_panes, "A2")
            sentence_metric_rows = {
                row[0]: row
                for row in sentence_metrics_worksheet.iter_rows(
                    min_row=2,
                    values_only=True,
                )
            }
            sentence_metric_row_indexes = {
                sentence_metrics_worksheet[f"A{row_index}"].value: row_index
                for row_index in range(2, sentence_metrics_worksheet.max_row + 1)
            }
            self.assertEqual(
                sentence_metric_rows["sentence_count"],
                ("sentence_count", 2, "count", "volume", None),
            )
            self.assertEqual(
                sentence_metric_rows["sentence_with_speaker_pct"],
                ("sentence_with_speaker_pct", 0.5, "pct", "volume", None),
            )
            self.assertEqual(
                sentence_metric_rows["sentence_word_count_le_3_count"],
                ("sentence_word_count_le_3_count", 1, "count", "word_buckets", None),
            )
            self.assertEqual(
                sentence_metric_rows["sentence_duration_2_5s_count"],
                ("sentence_duration_2_5s_count", 1, "count", "duration_buckets", None),
            )
            self.assertEqual(
                sentence_metric_rows["sentences_with_speaker_change_inside_count"],
                (
                    "sentences_with_speaker_change_inside_count",
                    1,
                    "count",
                    "speaker",
                    None,
                ),
            )
            self.assertEqual(
                sentence_metric_rows["sentences_with_provenance_overlap_count"],
                (
                    "sentences_with_provenance_overlap_count",
                    0,
                    "count",
                    "speaker",
                    None,
                ),
            )
            self.assertEqual(
                sentence_metric_rows["semantic_quality_borderline_count"],
                (
                    "semantic_quality_borderline_count",
                    1,
                    "count",
                    "semantic",
                    None,
                ),
            )
            self.assertEqual(
                sentence_metric_rows["sentence_length_profile"],
                ("sentence_length_profile", "short_heavy", "label", "profiles", None),
            )
            self.assertEqual(
                sentence_metric_rows["speaker_stability_profile"],
                ("speaker_stability_profile", "fragile", "label", "profiles", None),
            )
            self.assertEqual(
                sentence_metric_rows["sentence_debug_overview"],
                ("sentence_debug_overview", "review_needed", "label", "profiles", None),
            )
            self.assertEqual(
                sentence_metric_rows["sentence_total_duration_seconds"],
                (
                    "sentence_total_duration_seconds",
                    timedelta(seconds=3.8),
                    "seconds",
                    "duration_stats",
                    "Displayed as Excel duration with millisecond precision.",
                ),
            )
            self.assertAlmostEqual(
                float(sentence_metric_rows["sentence_avg_word_count"][1]),
                3.5,
            )
            self.assertAlmostEqual(
                float(sentence_metric_rows["sentence_p25_word_count"][1]),
                2.75,
            )
            self.assertAlmostEqual(
                float(sentence_metric_rows["speaker_instability_pressure_pct"][1]),
                0.5,
            )
            self.assertAlmostEqual(
                float(sentence_metric_rows["semantic_risk_pressure_pct"][1]),
                0.5,
            )
            sentence_with_speaker_pct_row = sentence_metric_row_indexes["sentence_with_speaker_pct"]
            sentence_total_duration_row = sentence_metric_row_indexes["sentence_total_duration_seconds"]
            self.assertEqual(
                sentence_metrics_worksheet[f"B{sentence_with_speaker_pct_row}"].number_format,
                "0.00%",
            )
            self.assertEqual(
                sentence_metrics_worksheet[f"B{sentence_total_duration_row}"].number_format,
                "[h]:mm:ss.000",
            )
            self.assertTrue(bool(sentence_metrics_worksheet["E2"].alignment.wrap_text))
            self.assertEqual(
                tuple(cell.value for cell in summary_worksheet[1]),
                ("section", "metric", "value", "value2", "notes"),
            )
            self.assertEqual(summary_worksheet.freeze_panes, "A2")
            summary_rows = {
                row[1]: row
                for row in summary_worksheet.iter_rows(min_row=2, values_only=True)
            }
            summary_row_indexes = {
                summary_worksheet[f"B{row_index}"].value: row_index
                for row_index in range(2, summary_worksheet.max_row + 1)
            }
            self.assertEqual(
                summary_rows["input_filenames"],
                ("input", "input_filenames", "lesson.wav", None, None),
            )
            self.assertEqual(
                summary_rows["total_input_duration"],
                (
                    "input",
                    "total_input_duration",
                    timedelta(seconds=120.5),
                    None,
                    "Sum of available original input durations.",
                ),
            )
            self.assertEqual(
                summary_rows["utterance_word_count_total"],
                ("debug", "utterance_word_count_total", 7, None, None),
            )
            self.assertEqual(
                summary_rows["sentence_count"],
                ("debug", "sentence_count", 2, None, None),
            )
            self.assertEqual(
                summary_rows["qa_search_strategies"],
                ("debug", "qa_search_strategies", "semantic_retrieval", None, None),
            )
            self.assertEqual(
                summary_rows["qa_ranking_strategies"],
                ("debug", "qa_ranking_strategies", "semantic_reranker", None, None),
            )
            self.assertEqual(
                summary_rows["qa_search_models"],
                (
                    "debug",
                    "qa_search_models",
                    "intfloat/multilingual-e5-base",
                    None,
                    None,
                ),
            )
            self.assertEqual(
                summary_rows["qa_reranking_models"],
                (
                    "debug",
                    "qa_reranking_models",
                    "BAAI/bge-reranker-v2-m3",
                    None,
                    None,
                ),
            )
            self.assertEqual(
                summary_rows["sentence_word_count_total"],
                ("debug", "sentence_word_count_total", 7, None, None),
            )
            self.assertEqual(
                summary_rows["pipeline_execution_mode"],
                ("timing_summary", "pipeline_execution_mode", "normal", None, None),
            )
            self.assertEqual(
                summary_rows["run_profile_label"],
                ("timing_summary", "run_profile_label", "mixed_run", None, None),
            )
            self.assertEqual(
                summary_rows["reused_cache_stage_count"],
                ("timing_summary", "reused_cache_stage_count", 1, None, None),
            )
            self.assertEqual(
                summary_rows["segment_count"],
                ("debug", "segment_count", 2, None, None),
            )
            self.assertEqual(
                summary_rows["sentences_with_speaker"],
                (
                    "debug",
                    "sentences_with_speaker",
                    1,
                    0.5,
                    "Speaker coverage across sentences.",
                ),
            )
            self.assertEqual(
                summary_rows["segment_missing_utterances"],
                (
                    "debug",
                    "segment_missing_utterances",
                    1,
                    1 / 3,
                    "Utterances with no resolved segment mapping.",
                ),
            )
            self.assertEqual(
                summary_rows["high_priority_sentence_reviews"],
                (
                    "debug",
                    "high_priority_sentence_reviews",
                    2,
                    1,
                    "Sentences flagged as high review priority.",
                ),
            )
            self.assertEqual(
                summary_rows["utterances_assigned_to_multiple_sentences"],
                (
                    "debug",
                    "utterances_assigned_to_multiple_sentences",
                    0,
                    0,
                    "Utterances linked to more than one final sentence.",
                ),
            )
            self.assertEqual(
                summary_rows["sentences_with_shared_source_utterances"],
                (
                    "debug",
                    "sentences_with_shared_source_utterances",
                    0,
                    0,
                    "Sentences that share at least one source utterance with another sentence.",
                ),
            )
            self.assertEqual(
                summary_rows["total_duration"],
                (
                    "timing_summary",
                    "total_duration",
                    timedelta(seconds=12.34),
                    summary_rows["total_duration"][3],
                    "Pipeline real-time factor over the original input duration.",
                ),
            )
            self.assertAlmostEqual(
                float(summary_rows["total_duration"][3]),
                12.34 / 120.5,
            )
            self.assertEqual(
                summary_rows["transcription"],
                (
                    "timing_stage",
                    "transcription",
                    timedelta(seconds=5.67),
                    summary_rows["transcription"][3],
                    "status=reused_from_cache | used_cache=true | rtf=0.05x | metadata=backend=fake",
                ),
            )
            self.assertAlmostEqual(
                float(summary_rows["transcription"][3]),
                5.67 / 12.34,
            )
            self.assertEqual(
                summary_rows["qa_extraction_duration"],
                (
                    "timing_summary",
                    "qa_extraction_duration",
                    timedelta(seconds=0.31),
                    summary_rows["qa_extraction_duration"][3],
                    "Duration of the qa_extraction pipeline stage when available.",
                ),
            )
            total_input_row = summary_row_indexes["total_input_duration"]
            utterances_with_speaker_row = summary_row_indexes["utterances_with_speaker"]
            sentences_with_speaker_row = summary_row_indexes["sentences_with_speaker"]
            total_duration_row = summary_row_indexes["total_duration"]
            self.assertEqual(
                summary_worksheet[f"C{total_input_row}"].number_format,
                "[h]:mm:ss.000",
            )
            self.assertEqual(
                summary_worksheet[f"D{utterances_with_speaker_row}"].number_format,
                "0.00%",
            )
            self.assertEqual(
                summary_worksheet[f"D{sentences_with_speaker_row}"].number_format,
                "0.00%",
            )
            self.assertEqual(
                summary_worksheet[f"D{total_duration_row}"].number_format,
                '0.00"x"',
            )
            self.assertTrue(bool(summary_worksheet["E2"].alignment.wrap_text))

    def test_export_run_to_excel_marks_context_only_sentences_in_debug_summary(
        self,
    ) -> None:
        """Context-only sentences should appear in the sentence QA debug summary."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            json_path = temp_path / "session_contextual.json"
            excel_path = temp_path / "session_contextual.xlsx"

            payload = {
                "session_metadata": {
                    "schema_version": "0.6.0",
                    "session_id": "session_contextual",
                    "language_codes": ["en"],
                    "processing_status": "ready",
                    "metadata": {},
                },
                "utterances": [
                    {
                        "utterance_id": "utterance_001",
                        "audio_source_id": "audio_source_001",
                        "text": "We are comparing inches and centimeters on a ruler.",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "speaker_id": "speaker_1",
                    },
                    {
                        "utterance_id": "utterance_002",
                        "audio_source_id": "audio_source_001",
                        "text": (
                            "Where is there a point where an integer number of "
                            "centimeters lines up exactly with an integer number "
                            "of inches?"
                        ),
                        "start_seconds": 1.1,
                        "end_seconds": 2.5,
                        "speaker_id": "speaker_1",
                    },
                    {
                        "utterance_id": "utterance_003",
                        "audio_source_id": "audio_source_001",
                        "text": "Yeah, it happens at 50 inches.",
                        "start_seconds": 2.6,
                        "end_seconds": 3.6,
                        "speaker_id": "speaker_2",
                    },
                ],
                "sentences": [
                    {
                        "sentence_id": "sentence_001",
                        "audio_source_id": "audio_source_001",
                        "text": "We are comparing inches and centimeters on a ruler.",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "source_utterance_ids": ["utterance_001"],
                        "speaker_id": "speaker_1",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "low",
                        "sentence_review_flags": [],
                        "metadata": {"source_utterance_count": 1},
                    },
                    {
                        "sentence_id": "sentence_002",
                        "audio_source_id": "audio_source_001",
                        "text": (
                            "Where is there a point where an integer number of "
                            "centimeters lines up exactly with an integer number "
                            "of inches?"
                        ),
                        "start_seconds": 1.1,
                        "end_seconds": 2.5,
                        "source_utterance_ids": ["utterance_002"],
                        "speaker_id": "speaker_1",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "low",
                        "sentence_review_flags": [],
                        "metadata": {"source_utterance_count": 1},
                    },
                    {
                        "sentence_id": "sentence_003",
                        "audio_source_id": "audio_source_001",
                        "text": "Yeah, it happens at 50 inches.",
                        "start_seconds": 2.6,
                        "end_seconds": 3.6,
                        "source_utterance_ids": ["utterance_003"],
                        "speaker_id": "speaker_2",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "low",
                        "sentence_review_flags": [],
                        "metadata": {"source_utterance_count": 1},
                    },
                ],
                "qa_candidates": [
                    {
                        "qa_candidate_id": "qa_100",
                        "question_text": (
                            "Where is there a point where an integer number of "
                            "centimeters lines up exactly with an integer number "
                            "of inches?"
                        ),
                        "answer_text": "Yeah, it happens at 50 inches.",
                        "context_text": (
                            "The speakers are comparing inches and centimeters "
                            "on a ruler and looking for a matching point."
                        ),
                        "question_sentence_ids": ["sentence_002"],
                        "answer_sentence_ids": ["sentence_003"],
                        "context_sentence_ids": ["sentence_001"],
                        "question_source_utterance_ids": ["utterance_002"],
                        "answer_source_utterance_ids": ["utterance_003"],
                        "context_source_utterance_ids": ["utterance_001"],
                        "question_speaker_role": "student",
                        "answer_speaker_role": "teacher",
                        "confidence_score": 0.84,
                        "confidence_label": "high",
                        "question_type": "direct_question",
                        "context_strategy": "local_topic_window",
                        "context_confidence": "high",
                        "reason_codes": ["deferred_answer_search"],
                        "review_flags": [],
                        "metadata": {
                            "input_layer": "sentences",
                            "question_debug": {"input_layer": "sentences"},
                            "answer_debug": {},
                            "pairing_debug": {
                                "deferred_answer_search_used": True,
                            },
                            "grounding_debug": {},
                            "context_debug": {
                                "context_strategy": "local_topic_window",
                                "context_confidence": "high",
                                "context_sentence_ids": ["sentence_001"],
                            },
                            "confidence_debug": {},
                        },
                    },
                ],
                "segments": [
                    {
                        "segment_id": "segment_001",
                        "sentence_ids": ["sentence_001", "sentence_002", "sentence_003"],
                    },
                ],
            }
            json_path.write_text(
                json.dumps(payload, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            export_run_to_excel(json_path, excel_path)

            workbook = load_workbook(excel_path)
            worksheet = workbook["Sentences"]
            header_index = {
                cell.value: index
                for index, cell in enumerate(worksheet[1], start=1)
            }

            self.assertIn(
                "CONTEXT | for qa_100",
                str(worksheet.cell(2, header_index["qa_debug_summary"]).value),
            )
            self.assertIn(
                "QUESTION | qa_100 | deferred answer | context: 001",
                str(worksheet.cell(3, header_index["qa_debug_summary"]).value),
            )
            self.assertIn(
                "ANSWER | for qa_100",
                str(worksheet.cell(4, header_index["qa_debug_summary"]).value),
            )

    def test_sentence_debug_export_includes_speaker_decision_fields_and_metrics(
        self,
    ) -> None:
        """Sentence rows and metrics should expose the new speaker decision signals."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            json_path = temp_path / "session_structural.json"
            excel_path = temp_path / "session_structural.xlsx"

            payload = {
                "session_metadata": {
                    "schema_version": "0.6.0",
                    "session_id": "session_001",
                    "processing_status": "ready",
                    "metadata": {},
                },
                "input_sources": [
                    {
                        "source_id": "source_001",
                        "original_filename": "lesson.wav",
                        "original_path": str(temp_path / "lesson.wav"),
                        "duration_seconds": 30.0,
                    },
                ],
                "audio_sources": [
                    {
                        "audio_source_id": "audio_source_001",
                        "duration_seconds": 30.0,
                    },
                ],
                "transcript": {
                    "full_text": "Teacher explains briefly",
                    "utterance_count": 2,
                    "sentence_count": 1,
                    "sentence_with_speaker_count": 1,
                },
                "utterances": [
                    {
                        "utterance_id": "utterance_001",
                        "audio_source_id": "audio_source_001",
                        "text": "Teacher explains",
                        "start_seconds": 0.0,
                        "end_seconds": 1.4,
                        "speaker_id": "speaker_1",
                        "speaker_is_uncertain": False,
                    },
                    {
                        "utterance_id": "utterance_002",
                        "audio_source_id": "audio_source_001",
                        "text": "briefly",
                        "start_seconds": 1.45,
                        "end_seconds": 1.7,
                        "speaker_id": "speaker_2",
                        "speaker_is_uncertain": False,
                    },
                ],
                "sentences": [
                    {
                        "sentence_id": "audio_source_001_sentence_0001",
                        "audio_source_id": "audio_source_001",
                        "text": "Teacher explains briefly",
                        "start_seconds": 0.0,
                        "end_seconds": 1.7,
                        "source_utterance_ids": ["utterance_001", "utterance_002"],
                        "source_utterance_start_index": 1,
                        "source_utterance_end_index": 2,
                        "speaker_id": "speaker_1",
                        "speaker_resolution_status": "mostly_stable",
                        "speaker_confidence_label": "medium",
                        "speaker_stability_label": "mostly_stable",
                        "speaker_assignment_method": "direct_weighted_majority",
                        "speaker_evidence_summary": (
                            "method=direct_weighted_majority; dominant=speaker_1:2.40; "
                            "second=speaker_2:0.60; share=0.80; margin=1.80; "
                            "assigned=2; uncertain=0; unassigned=0; short=1"
                        ),
                        "merge_safety_label": "borderline",
                        "semantic_quality_label": "good",
                        "length_bucket": "short",
                        "duration_bucket": "short",
                        "review_priority": "medium",
                        "sentence_review_flags": ["multi_utterance"],
                        "metadata": {
                            "source_utterance_count": 2,
                            "has_uncertain_source": False,
                            "has_unassigned_source": False,
                            "has_speaker_change_inside": True,
                            "is_multi_utterance": True,
                            "speaker_assignment_method": "direct_weighted_majority",
                            "dominant_speaker_weight": 2.4,
                            "second_speaker_weight": 0.6,
                            "dominance_margin": 1.8,
                            "dominant_speaker_share": 0.8,
                            "assigned_source_utterance_count": 2,
                            "uncertain_source_utterance_count": 0,
                            "unassigned_source_utterance_count": 0,
                            "short_fragment_source_utterance_count": 1,
                            "speaker_evidence": {
                                "dominant_weight": 2.4,
                                "second_weight": 0.6,
                                "dominance_margin": 1.8,
                                "dominant_share": 0.8,
                            },
                        },
                    },
                ],
                "segments": [
                    {
                        "segment_id": "segment_0001",
                        "sentence_ids": ["audio_source_001_sentence_0001"],
                    },
                ],
            }
            json_path.write_text(
                json.dumps(payload, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            export_run_to_excel(json_path, excel_path)

            workbook = load_workbook(excel_path)
            sentences_worksheet = workbook["Sentences"]
            sentence_metrics_worksheet = workbook["SentenceMetrics"]
            header_index = {
                cell.value: index
                for index, cell in enumerate(sentences_worksheet[1], start=1)
            }

            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    header_index["speaker_assignment_method"],
                ).value,
                "direct_weighted_majority",
            )
            self.assertEqual(
                float(
                    sentences_worksheet.cell(
                        2,
                        header_index["dominant_speaker_weight"],
                    ).value,
                ),
                2.4,
            )
            self.assertEqual(
                float(
                    sentences_worksheet.cell(
                        2,
                        header_index["second_speaker_weight"],
                    ).value,
                ),
                0.6,
            )
            self.assertEqual(
                float(
                    sentences_worksheet.cell(
                        2,
                        header_index["dominance_margin"],
                    ).value,
                ),
                1.8,
            )
            self.assertEqual(
                int(
                    sentences_worksheet.cell(
                        2,
                        header_index["short_fragment_source_utterance_count"],
                    ).value,
                ),
                1,
            )

            sentence_metric_rows = {
                row[0]: row
                for row in sentence_metrics_worksheet.iter_rows(
                    min_row=2,
                    values_only=True,
                )
            }
            self.assertEqual(
                sentence_metric_rows[
                    "speaker_assignment_method_direct_weighted_majority_count"
                ][1],
                1,
            )
            self.assertEqual(
                sentence_metric_rows[
                    "speaker_resolution_status_mostly_stable_count"
                ][1],
                1,
            )
            self.assertEqual(
                sentence_metric_rows[
                    "speaker_stability_label_mostly_stable_count"
                ][1],
                1,
            )

    def test_export_run_to_excel_highlights_provenance_conflicts(self) -> None:
        """The exporter should surface utterance-to-sentence mapping conflicts."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            json_path = temp_path / "session_conflict.json"
            excel_path = temp_path / "session_conflict.xlsx"

            payload = {
                "session_metadata": {
                    "schema_version": "0.6.0",
                    "session_id": "session_conflict",
                    "processing_status": "ready",
                },
                "utterances": [
                    {
                        "utterance_id": "utterance_001",
                        "audio_source_id": "audio_source_001",
                        "text": "Shared utterance",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "speaker_id": "speaker_1",
                        "speaker_attribution_status": "assigned",
                        "speaker_is_uncertain": False,
                    },
                ],
                "sentences": [
                    {
                        "sentence_id": "sentence_001",
                        "audio_source_id": "audio_source_001",
                        "text": "Shared utterance",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "source_utterance_ids": ["utterance_001"],
                        "speaker_id": "speaker_1",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "low",
                        "metadata": {
                            "source_utterance_count": 1,
                            "speaker_evidence": {
                                "assigned_speakers": ["speaker_1"],
                            },
                        },
                    },
                    {
                        "sentence_id": "sentence_002",
                        "audio_source_id": "audio_source_001",
                        "text": "Shared utterance duplicate",
                        "start_seconds": 1.1,
                        "end_seconds": 2.0,
                        "source_utterance_ids": ["utterance_001"],
                        "speaker_id": "speaker_1",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "low",
                        "metadata": {
                            "source_utterance_count": 1,
                            "speaker_evidence": {
                                "assigned_speakers": ["speaker_1"],
                            },
                        },
                    },
                ],
                "segments": [],
            }
            json_path.write_text(
                json.dumps(payload, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            export_run_to_excel(json_path, excel_path)

            workbook = load_workbook(excel_path)
            utterances_worksheet = workbook["Utterances"]
            sentences_worksheet = workbook["Sentences"]
            summary_worksheet = workbook["run_summary"]
            sentence_header_index = {
                cell.value: index
                for index, cell in enumerate(sentences_worksheet[1], start=1)
            }

            self.assertEqual(utterances_worksheet["O2"].value, 2)
            self.assertEqual(utterances_worksheet["P2"].value, True)
            self.assertIn(
                "sentence_mapping_conflict",
                str(utterances_worksheet["S2"].value),
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["has_provenance_overlap"],
                ).value,
                True,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    3,
                    sentence_header_index["has_provenance_overlap"],
                ).value,
                True,
            )
            self.assertIn(
                "sentence_002",
                str(
                    sentences_worksheet.cell(
                        2,
                        sentence_header_index["provenance_overlap_sentence_ids"],
                    ).value,
                ),
            )
            self.assertIn(
                "sentence_001",
                str(
                    sentences_worksheet.cell(
                        3,
                        sentence_header_index["provenance_overlap_sentence_ids"],
                    ).value,
                ),
            )
            self.assertNotIn(
                "sentence_001",
                str(
                    sentences_worksheet.cell(
                        2,
                        sentence_header_index["provenance_overlap_sentence_ids"],
                    ).value,
                ),
            )
            self.assertNotIn(
                "sentence_002",
                str(
                    sentences_worksheet.cell(
                        3,
                        sentence_header_index["provenance_overlap_sentence_ids"],
                    ).value,
                ),
            )
            self.assertIn(
                "provenance_overlap",
                str(
                    sentences_worksheet.cell(
                        2,
                        sentence_header_index["review_flags"],
                    ).value,
                ),
            )

            summary_rows = {
                row[1]: row
                for row in summary_worksheet.iter_rows(min_row=2, values_only=True)
            }
            self.assertEqual(
                summary_rows["utterances_assigned_to_multiple_sentences"][2],
                1,
            )
            self.assertEqual(
                summary_rows["sentence_assignment_total"][2],
                2,
            )
            self.assertEqual(
                summary_rows["sentences_with_shared_source_utterances"][2],
                2,
            )
            self.assertEqual(
                summary_rows["max_sentence_reuse_per_utterance"][2],
                2,
            )
            self.assertEqual(
                summary_rows["all_sentences_have_provenance_overlap"][2],
                True,
            )

    def test_export_run_to_excel_preserves_existing_human_comments(self) -> None:
        """Manual comments should survive when the workbook is regenerated."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            json_path = temp_path / "session_comments.json"
            excel_path = temp_path / "session_comments.xlsx"

            payload = {
                "session_metadata": {
                    "schema_version": "0.6.0",
                    "session_id": "session_comments",
                    "processing_status": "ready",
                },
                "utterances": [
                    {
                        "utterance_id": "utterance_001",
                        "audio_source_id": "audio_source_001",
                        "text": "What is a graph?",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "speaker_id": "speaker_1",
                        "speaker_attribution_status": "assigned",
                        "speaker_is_uncertain": False,
                    },
                ],
                "sentences": [
                    {
                        "sentence_id": "sentence_001",
                        "audio_source_id": "audio_source_001",
                        "text": "What is a graph?",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "source_utterance_ids": ["utterance_001"],
                        "speaker_id": "speaker_1",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "high",
                        "sentence_review_flags": ["uncertain_source"],
                        "metadata": {
                            "source_utterance_count": 1,
                            "has_uncertain_source": True,
                            "speaker_evidence": {
                                "assigned_speakers": ["speaker_1"],
                            },
                        },
                    },
                ],
                "qa_candidates": [
                    {
                        "qa_candidate_id": "qa_010",
                        "question_text": "What is a graph?",
                        "answer_text": "A graph is a structure of nodes and edges.",
                        "question_sentence_ids": ["sentence_001"],
                        "answer_sentence_ids": ["sentence_001"],
                        "question_source_utterance_ids": ["utterance_001"],
                        "answer_source_utterance_ids": ["utterance_001"],
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "question_speaker_role": "student",
                        "answer_speaker_role": "teacher",
                        "confidence_score": 0.82,
                        "confidence_label": "high",
                        "question_type": "direct_question",
                        "reason_codes": ["question_mark"],
                        "metadata": {
                            "input_layer": "sentences",
                            "question_debug": {"input_layer": "sentences"},
                            "answer_debug": {},
                            "pairing_debug": {},
                            "grounding_debug": {},
                            "confidence_debug": {},
                        },
                    },
                ],
                "segments": [
                    {
                        "segment_id": "segment_001",
                        "sentence_ids": ["sentence_001"],
                    },
                ],
            }
            json_path.write_text(
                json.dumps(payload, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            export_run_to_excel(json_path, excel_path)

            workbook = load_workbook(excel_path)
            sentences_worksheet = workbook["Sentences"]
            qa_worksheet = workbook["QACandidates"]
            review_candidates_worksheet = workbook["ReviewCandidates"]

            sentence_header_index = {
                cell.value: index
                for index, cell in enumerate(sentences_worksheet[1], start=1)
            }
            qa_header_index = {
                cell.value: index
                for index, cell in enumerate(qa_worksheet[1], start=1)
            }
            review_header_index = {
                cell.value: index
                for index, cell in enumerate(review_candidates_worksheet[1], start=1)
            }

            sentences_worksheet.cell(
                2,
                sentence_header_index["human_comment"],
            ).value = "Check speaker assignment"
            qa_worksheet.cell(
                2,
                qa_header_index["human_comment"],
            ).value = "Looks good"
            review_candidates_worksheet.cell(
                2,
                review_header_index["human_comment"],
            ).value = "Review during playback"
            workbook.save(excel_path)

            export_run_to_excel(json_path, excel_path)

            workbook = load_workbook(excel_path)
            sentences_worksheet = workbook["Sentences"]
            qa_worksheet = workbook["QACandidates"]
            review_candidates_worksheet = workbook["ReviewCandidates"]

            sentence_header_index = {
                cell.value: index
                for index, cell in enumerate(sentences_worksheet[1], start=1)
            }
            qa_header_index = {
                cell.value: index
                for index, cell in enumerate(qa_worksheet[1], start=1)
            }
            review_header_index = {
                cell.value: index
                for index, cell in enumerate(review_candidates_worksheet[1], start=1)
            }

            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["human_comment"],
                ).value,
                "Check speaker assignment",
            )
            self.assertEqual(
                qa_worksheet.cell(
                    2,
                    qa_header_index["human_comment"],
                ).value,
                "Looks good",
            )
            self.assertEqual(
                review_candidates_worksheet.cell(
                    2,
                    review_header_index["human_comment"],
                ).value,
                "Review during playback",
            )
            self.assertTrue(
                bool(
                    sentences_worksheet.column_dimensions["G"].hidden,
                ),
            )

    def test_export_run_to_excel_scopes_utterance_ids_by_audio_source(self) -> None:
        """Identical local utterance ids across sources should not create false overlap."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            json_path = temp_path / "session_scoped.json"
            excel_path = temp_path / "session_scoped.xlsx"

            payload = {
                "session_metadata": {
                    "schema_version": "0.6.0",
                    "session_id": "session_scoped",
                    "processing_status": "ready",
                },
                "utterances": [
                    {
                        "utterance_id": "utterance_shared",
                        "audio_source_id": "audio_source_001",
                        "text": "Alpha statement",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "speaker_id": "speaker_1",
                        "speaker_attribution_status": "assigned",
                        "speaker_is_uncertain": False,
                    },
                    {
                        "utterance_id": "utterance_shared",
                        "audio_source_id": "audio_source_002",
                        "text": "Beta statement",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "speaker_id": "speaker_2",
                        "speaker_attribution_status": "assigned",
                        "speaker_is_uncertain": False,
                    },
                ],
                "sentences": [
                    {
                        "sentence_id": "audio_source_001_sentence_0001",
                        "audio_source_id": "audio_source_001",
                        "text": "Alpha statement",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "source_utterance_ids": ["utterance_shared"],
                        "speaker_id": "speaker_1",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "low",
                        "metadata": {
                            "source_utterance_count": 1,
                            "speaker_evidence": {
                                "assigned_speakers": ["speaker_1"],
                            },
                        },
                    },
                    {
                        "sentence_id": "audio_source_002_sentence_0001",
                        "audio_source_id": "audio_source_002",
                        "text": "Beta statement",
                        "start_seconds": 0.0,
                        "end_seconds": 1.0,
                        "source_utterance_ids": ["utterance_shared"],
                        "speaker_id": "speaker_2",
                        "speaker_resolution_status": "stable",
                        "speaker_confidence_label": "high",
                        "speaker_stability_label": "stable",
                        "merge_safety_label": "safe",
                        "semantic_quality_label": "good",
                        "review_priority": "low",
                        "metadata": {
                            "source_utterance_count": 1,
                            "speaker_evidence": {
                                "assigned_speakers": ["speaker_2"],
                            },
                        },
                    },
                ],
                "segments": [],
            }
            json_path.write_text(
                json.dumps(payload, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            export_run_to_excel(json_path, excel_path)

            workbook = load_workbook(excel_path)
            utterances_worksheet = workbook["Utterances"]
            sentences_worksheet = workbook["Sentences"]
            summary_worksheet = workbook["run_summary"]
            sentence_header_index = {
                cell.value: index
                for index, cell in enumerate(sentences_worksheet[1], start=1)
            }

            self.assertEqual(utterances_worksheet["O2"].value, 1)
            self.assertEqual(utterances_worksheet["P2"].value, False)
            self.assertEqual(utterances_worksheet["O3"].value, 1)
            self.assertEqual(utterances_worksheet["P3"].value, False)
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["has_provenance_overlap"],
                ).value,
                False,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    3,
                    sentence_header_index["has_provenance_overlap"],
                ).value,
                False,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    2,
                    sentence_header_index["provenance_overlap_sentence_ids"],
                ).value,
                None,
            )
            self.assertEqual(
                sentences_worksheet.cell(
                    3,
                    sentence_header_index["provenance_overlap_sentence_ids"],
                ).value,
                None,
            )

            summary_rows = {
                row[1]: row
                for row in summary_worksheet.iter_rows(min_row=2, values_only=True)
            }
            self.assertEqual(
                summary_rows["utterances_assigned_to_multiple_sentences"][2],
                0,
            )
            self.assertEqual(
                summary_rows["sentence_assignment_total"][2],
                2,
            )
            self.assertEqual(
                summary_rows["sentences_with_shared_source_utterances"][2],
                0,
            )
            self.assertEqual(
                summary_rows["max_sentence_reuse_per_utterance"][2],
                1,
            )
            self.assertEqual(
                summary_rows["all_sentences_have_provenance_overlap"][2],
                False,
            )

    def test_pipeline_invokes_debug_export_only_when_enabled(self) -> None:
        """The pipeline should call the Excel exporter only behind the flag."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            media_path = temp_path / "lesson.wav"
            media_path.write_bytes(b"placeholder")

            enabled_config = PipelineConfig(
                working_directory=temp_path / "artifacts_enabled",
                segmentation_mode="both",
                transcript_alignment_enabled=False,
                diarization_enabled=False,
                enable_qa_extraction=False,
                export_debug_excel=True,
                debug_excel_path=Path("debug_exports"),
            )
            enabled_pipeline = self._build_fake_pipeline(enabled_config, media_path)

            with patch(
                "lecture_analyzer.core._processing_pipeline_impl.export_run_to_excel",
            ) as export_mock:
                enabled_pipeline.process(
                    media_path,
                    output_path=temp_path / "exports_enabled",
                )

            self.assertEqual(export_mock.call_count, 3)
            exported_excel_names = {
                Path(call.args[1]).name
                for call in export_mock.call_args_list
            }
            self.assertEqual(
                exported_excel_names,
                {
                    "lesson_structural.xlsx",
                    "lesson_windowed.xlsx",
                    "lesson_adaptive.xlsx",
                },
            )

            disabled_config = PipelineConfig(
                working_directory=temp_path / "artifacts_disabled",
                segmentation_mode="structural",
                transcript_alignment_enabled=False,
                diarization_enabled=False,
                enable_qa_extraction=False,
                export_debug_excel=False,
            )
            disabled_pipeline = self._build_fake_pipeline(disabled_config, media_path)

            with patch(
                "lecture_analyzer.core._processing_pipeline_impl.export_run_to_excel",
            ) as export_mock:
                disabled_pipeline.process(
                    media_path,
                    output_path=temp_path / "exports_disabled",
                )

            export_mock.assert_not_called()

    def test_pipeline_uses_input_name_when_debug_path_is_file(self) -> None:
        """An explicit Excel filename should still include the exported input stem."""

        with tempfile.TemporaryDirectory() as temp_directory:
            temp_path = Path(temp_directory)
            media_path = temp_path / "lesson.wav"
            media_path.write_bytes(b"placeholder")

            config = PipelineConfig(
                working_directory=temp_path / "artifacts_named",
                segmentation_mode="structural",
                transcript_alignment_enabled=False,
                diarization_enabled=False,
                enable_qa_extraction=False,
                export_debug_excel=True,
                debug_excel_path=Path("manual_debug.xlsx"),
            )
            pipeline = self._build_fake_pipeline(config, media_path)

            with patch(
                "lecture_analyzer.core._processing_pipeline_impl.export_run_to_excel",
            ) as export_mock:
                pipeline.process(
                    media_path,
                    output_path=temp_path / "exports_named",
                )

            export_mock.assert_called_once()
            self.assertEqual(
                Path(export_mock.call_args.args[1]).name,
                "manual_debug_lesson_structural.xlsx",
            )

    @staticmethod
    def _build_fake_pipeline(
        config: PipelineConfig,
        media_path: Path,
    ) -> LectureProcessingPipeline:
        """Build a pipeline wired with lightweight in-memory collaborators."""

        pipeline = LectureProcessingPipeline(config)
        pipeline.session_loader = _FakeSessionLoader(media_path)
        pipeline.audio_normalizer = _FakeAudioNormalizer(media_path)
        pipeline.transcriber = _FakeTranscriber()
        pipeline.whisperx_aligner = _FakeAligner()
        pipeline.utterance_builder = _FakeUtteranceBuilder()
        pipeline.pyannote_diarizer = _FakeDiarizer()
        pipeline.speaker_attributor = _FakeSpeakerAttributor()
        pipeline.transcript_merger = _FakeTranscriptMerger()
        pipeline.transcript_normalizer = _IdentityTranscriptNormalizer()
        pipeline.segmenter = _FakeSegmenter()
        pipeline.qa_extractor = _FakeQAExtractor()
        return pipeline


class _FakeSessionLoader:
    """Return a deterministic in-memory session for exporter tests."""

    def __init__(self, media_path: Path) -> None:
        self.media_path = media_path

    def load_session(
        self,
        input_paths: str | Path | list[str | Path],
        session_id: str | None = None,
    ) -> LectureSession:
        return LectureSession(
            session_id=session_id or "session_001",
            input_sources=[
                InputSource(
                    source_id="source_001",
                    original_path=self.media_path,
                    media_type=MediaType.AUDIO,
                    order_index=1,
                    original_filename=self.media_path.name,
                ),
            ],
            language_codes=["it"],
        )


class _FakeAudioNormalizer:
    """Produce one deterministic normalized audio source."""

    def __init__(self, media_path: Path) -> None:
        self.media_path = media_path

    def normalize_sources(self, input_sources: list[InputSource]) -> list[AudioSource]:
        return [
            AudioSource(
                audio_source_id="audio_source_001",
                input_source_id=input_sources[0].source_id,
                audio_path=self.media_path,
                audio_format="wav",
                order_index=1,
                duration_seconds=12.0,
            ),
        ]


class _FakeTranscriber:
    """Populate one deterministic transcript chunk."""

    def transcribe_session(self, session: LectureSession) -> list[TranscriptChunk]:
        session.transcript_chunks = [
            TranscriptChunk(
                chunk_id="audio_source_001_chunk_0001",
                audio_source_id="audio_source_001",
                start_seconds=0.0,
                end_seconds=4.0,
                text="What is a matrix?",
                detected_language="it",
            ),
        ]
        session.transcript_text = "What is a matrix?"
        session.metadata["transcription_backend"] = "fake"
        return session.transcript_chunks


class _FakeAligner:
    """Expose the disabled-alignment branch without external dependencies."""

    def align_session(self, session: LectureSession) -> list[object]:
        session.aligned_transcripts = []
        session.metadata["transcript_alignment_enabled"] = False
        session.metadata["transcript_alignment_status"] = "disabled"
        session.metadata["transcript_alignment_word_count"] = 0
        session.metadata["transcript_alignment_failed_sources"] = []
        for audio_source in session.audio_sources:
            audio_source.metadata["alignment"] = {
                "status": "disabled",
                "reason": "alignment_disabled",
            }
        return []


class _FakeUtteranceBuilder:
    """Expose the skipped utterance-building path."""

    def build_session(self, session: LectureSession) -> list[object]:
        session.utterances = []
        session.metadata["utterance_build_status"] = "skipped"
        session.metadata["utterance_build_reason"] = "aligned_transcripts_unavailable"
        session.metadata["utterance_failed_sources"] = []
        return []


class _FakeDiarizer:
    """Expose the disabled diarization path."""

    def diarize_session(self, session: LectureSession) -> list[object]:
        session.diarization_segments = []
        session.metadata["diarization_enabled"] = False
        session.metadata["diarization_status"] = "disabled"
        session.metadata["diarization_failed_sources"] = []
        session.metadata["diarization_speaker_count"] = 0
        for audio_source in session.audio_sources:
            audio_source.metadata["diarization"] = {
                "status": "disabled",
                "reason": "diarization_disabled",
            }
        return []


class _FakeSpeakerAttributor:
    """Expose skipped speaker attribution when diarization is unavailable."""

    def attribute_session(self, session: LectureSession) -> list[object]:
        session.metadata["speaker_attribution_status"] = "skipped"
        session.metadata["speaker_attribution_reason"] = "diarization_unavailable"
        session.metadata["speaker_attribution_assigned_count"] = 0
        session.metadata["speaker_attribution_unassigned_count"] = 0
        return session.utterances


class _FakeTranscriptMerger:
    """Build a minimal merged transcript for segmentation and export."""

    def merge_session(self, session: LectureSession) -> MergedTranscript:
        return MergedTranscript(
            session_id=session.session_id,
            units=[
                MergedTranscriptUnit(
                    unit_id="unit_0001",
                    chunk_id="audio_source_001_chunk_0001",
                    chunk_occurrence=1,
                    audio_source_id="audio_source_001",
                    source_order_index=1,
                    input_source_id="source_001",
                    start_seconds=0.0,
                    end_seconds=4.0,
                    session_start_seconds=0.0,
                    session_end_seconds=4.0,
                    text="What is a matrix?",
                    detected_language="it",
                ),
            ],
            full_text="What is a matrix?",
            detected_languages=["it"],
        )


class _IdentityTranscriptNormalizer:
    """Leave the merged transcript unchanged."""

    def normalize(self, merged_transcript: MergedTranscript) -> MergedTranscript:
        return merged_transcript


class _FakeSegmenter:
    """Return one deterministic segment per requested mode."""

    def resolved_mode(self, mode: str | None = None) -> str:
        if mode in {"structural", "windowed", "adaptive"}:
            return str(mode)
        return "structural"

    def segment_session(
        self,
        session: LectureSession,
        mode: str | None = None,
    ) -> list[Segment]:
        resolved_mode = self.resolved_mode(mode)
        return [
            Segment(
                segment_id=f"{resolved_mode}_segment_0001",
                start_seconds=0.0,
                end_seconds=4.0,
                text=session.transcript_text,
                transcript_chunk_ids=["audio_source_001_chunk_0001"],
                merged_transcript_unit_ids=["unit_0001"],
                audio_source_ids=["audio_source_001"],
                metadata={"segmentation_mode": resolved_mode},
            ),
        ]


class _FakeQAExtractor:
    """Return no QA candidates so the test stays focused on export hooks."""

    def extract(self, session: LectureSession) -> list[object]:
        return []


if __name__ == "__main__":
    unittest.main()
