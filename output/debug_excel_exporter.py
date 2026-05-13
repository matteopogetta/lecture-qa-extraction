"""Excel export helpers for manual utterance-level debugging.

This module reads the session JSON artifact produced by the pipeline and writes
an `.xlsx` workbook tailored for human review. The exporter intentionally
depends only on the serialized JSON structure so it stays decoupled from the
pipeline orchestration and the in-memory data models.
"""

from __future__ import annotations

from datetime import timedelta
import hashlib
import json
from pathlib import Path
import statistics
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from output.sentence_provenance_validation import (
    SentenceProvenanceValidation,
    scoped_utterance_key,
    validate_sentence_provenance,
)


COMMENT_HEADER_ALIASES = (
    "human_comment",
    "manual_comment",
    "comment",
    "commento",
    "human_note",
)

HEADERS = (
    "id",
    "start",
    "end",
    "duration_s",
    "speaker",
    "speaker_status",
    "is_uncertain_speaker",
    "is_unassigned_speaker",
    "is_speaker_changed",
    "is_segment_missing",
    "is_audio_quality_degraded",
    "review_priority",
    "segment_id",
    "final_sentence_id",
    "sentence_assignment_count",
    "has_sentence_mapping_conflict",
    "text",
    "word_count",
    "review_flags",
    "notes",
)

SENTENCE_TECHNICAL_HEADERS = (
    "id",
    "start",
    "end",
    "duration_s",
    "speaker",
    "detected_language",
    "source_utterance_count",
    "source_utterance_span",
    "is_multi_utterance",
    "has_uncertain_source",
    "has_unassigned_source",
    "has_speaker_change_inside",
    "is_semantic_fragment",
    "is_semantic_run_on",
    "is_merge_risky",
    "semantic_quality_label",
    "length_bucket",
    "duration_bucket",
    "speaker_stability_label",
    "merge_safety_label",
    "review_priority",
    "segment_id",
    "text",
    "word_count",
    "source_utterance_ids",
    "distinct_source_speaker_count",
    "source_speaker_ids",
    "assigned_source_utterance_count",
    "uncertain_source_utterance_count",
    "unassigned_source_utterance_count",
    "speaker_resolution_status",
    "speaker_confidence_label",
    "has_provenance_overlap",
    "provenance_overlap_sentence_ids",
    "review_flags",
    "notes",
    "speaker_assignment_method",
    "dominant_speaker_weight",
    "second_speaker_weight",
    "dominance_margin",
    "dominant_speaker_share",
    "short_fragment_source_utterance_count",
)

SENTENCE_CORE_HEADERS = (
    "sentence_id",
    "time_range",
    "speaker",
    "text",
    "qa_debug_summary",
    "human_comment",
)

SENTENCE_HEADERS = SENTENCE_CORE_HEADERS + tuple(
    header
    for header in SENTENCE_TECHNICAL_HEADERS
    if header not in {"speaker", "text"}
)

QA_TECHNICAL_HEADERS = (
    "id",
    "start",
    "end",
    "duration_s",
    "question_type",
    "didactic_question_score",
    "confidence_score",
    "confidence_label",
    "question_score",
    "answer_score",
    "base_confidence",
    "has_answer",
    "answer_is_question",
    "deferred_answer_search_used",
    "review_priority",
    "input_layer",
    "question_speaker_role",
    "answer_speaker_role",
    "question_segment_id",
    "answer_segment_id",
    "segment_relation",
    "requested_search_strategy",
    "effective_search_strategy",
    "requested_ranking_strategy",
    "effective_ranking_strategy",
    "search_stop_reason",
    "search_fallback_reason",
    "search_backend_error",
    "ranking_fallback_reason",
    "search_backend_status",
    "search_model_name",
    "search_backend",
    "semantic_similarity",
    "candidate_channel",
    "reranking_model_name",
    "reranking_backend",
    "semantic_relevance_score",
    "combined_rank_score",
    "rank_position",
    "speaker_influence",
    "question_timing_source",
    "answer_timing_source",
    "question_text",
    "answer_text",
    "context_text",
    "question_sentence_ids",
    "answer_sentence_ids",
    "context_sentence_ids",
    "question_source_utterance_ids",
    "answer_source_utterance_ids",
    "context_source_utterance_ids",
    "question_unit_ids",
    "answer_unit_ids",
    "context_strategy",
    "context_confidence",
    "source_segment_ids",
    "reason_codes",
    "review_flags",
    "notes",
)

QA_CORE_HEADERS = (
    "qa_id",
    "question_sentence_id",
    "answer_sentence_id",
    "context_sentence_id",
    "question_text",
    "answer_text",
    "context_text",
    "question_sentence_ids",
    "answer_sentence_ids",
    "context_sentence_ids",
    "context_strategy",
    "context_confidence",
    "confidence",
    "review_flag_summary",
    "review_flags",
    "human_comment",
)

QA_HEADERS = QA_CORE_HEADERS + tuple(
    header
    for header in QA_TECHNICAL_HEADERS
    if header not in set(QA_CORE_HEADERS)
)

REVIEW_CANDIDATE_TECHNICAL_HEADERS = (
    "kind",
    "id",
    "start",
    "review_priority",
    "speaker",
    "segment_id",
    "quality",
    "review_flags",
    "text",
    "notes",
)

REVIEW_CANDIDATE_CORE_HEADERS = (
    "candidate_id",
    "kind",
    "review_reason",
    "text",
    "suggested_action",
    "human_comment",
)

REVIEW_CANDIDATE_HEADERS = REVIEW_CANDIDATE_CORE_HEADERS + tuple(
    header
    for header in REVIEW_CANDIDATE_TECHNICAL_HEADERS
    if header not in {"kind", "text"}
)

SUMMARY_HEADERS = ("section", "metric", "value", "value2", "notes")
SENTENCE_METRIC_HEADERS = (
    "metric_name",
    "metric_value",
    "metric_unit",
    "metric_group",
    "notes",
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
CORE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="CFE2F3")
EDITABLE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
EDITABLE_CELL_FILL = PatternFill(fill_type="solid", fgColor="F3FBF1")
UNCERTAIN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
HEADER_FONT = Font(bold=True)
SPEAKER_FILL_COLORS = (
    "E8F4EA",
    "E7F0FA",
    "FCE8E6",
    "FFF4D6",
    "F1E8FB",
    "E6F7F5",
    "FDEBF3",
    "EDF5D6",
    "E9ECF8",
    "F8EBDD",
)
SPEAKER_CHANGE_BORDER = Border(
    top=Side(style="medium", color="A6A6A6"),
)
TOP_ALIGNMENT = Alignment(vertical="top")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

COLUMN_WIDTHS = {
    "A": 22,
    "B": 13,
    "C": 13,
    "D": 13,
    "E": 16,
    "F": 18,
    "G": 16,
    "H": 16,
    "I": 16,
    "J": 16,
    "K": 18,
    "L": 16,
    "M": 18,
    "N": 24,
    "O": 12,
    "P": 18,
    "Q": 72,
    "R": 12,
    "S": 32,
    "T": 42,
}

SUMMARY_COLUMN_WIDTHS = {
    "A": 18,
    "B": 32,
    "C": 18,
    "D": 14,
    "E": 72,
}

SENTENCE_METRIC_COLUMN_WIDTHS = {
    "A": 42,
    "B": 18,
    "C": 16,
    "D": 22,
    "E": 72,
}

WORD_COUNT_BUCKETS = (
    ("le_2", lambda value: value <= 2),
    ("le_3", lambda value: value <= 3),
    ("4_6", lambda value: 4 <= value <= 6),
    ("7_12", lambda value: 7 <= value <= 12),
    ("13_24", lambda value: 13 <= value <= 24),
    ("ge_25", lambda value: value >= 25),
    ("ge_35", lambda value: value >= 35),
    ("ge_50", lambda value: value >= 50),
)

DURATION_BUCKETS = (
    ("lt_1s", lambda value: value < 1.0),
    ("1_2s", lambda value: 1.0 <= value < 2.0),
    ("2_5s", lambda value: 2.0 <= value < 5.0),
    ("5_10s", lambda value: 5.0 <= value < 10.0),
    ("ge_10s", lambda value: value >= 10.0),
    ("ge_15s", lambda value: value >= 15.0),
    ("ge_20s", lambda value: value >= 20.0),
)

SENTENCE_COLUMN_WIDTHS = {
    "sentence_id": 26,
    "time_range": 28,
    "speaker": 16,
    "text": 108,
    "qa_debug_summary": 44,
    "human_comment": 64,
    "id": 24,
    "start": 13,
    "end": 13,
    "duration_s": 13,
    "detected_language": 14,
    "source_utterance_count": 18,
    "source_utterance_span": 18,
    "is_multi_utterance": 16,
    "has_uncertain_source": 18,
    "has_unassigned_source": 18,
    "has_speaker_change_inside": 18,
    "is_semantic_fragment": 18,
    "is_semantic_run_on": 18,
    "is_merge_risky": 16,
    "semantic_quality_label": 18,
    "length_bucket": 12,
    "duration_bucket": 12,
    "speaker_stability_label": 18,
    "merge_safety_label": 18,
    "review_priority": 18,
    "segment_id": 18,
    "word_count": 12,
    "source_utterance_ids": 28,
    "distinct_source_speaker_count": 14,
    "source_speaker_ids": 28,
    "assigned_source_utterance_count": 14,
    "uncertain_source_utterance_count": 14,
    "unassigned_source_utterance_count": 14,
    "speaker_resolution_status": 20,
    "speaker_confidence_label": 16,
    "has_provenance_overlap": 14,
    "provenance_overlap_sentence_ids": 34,
    "review_flags": 36,
    "notes": 42,
    "speaker_assignment_method": 28,
    "dominant_speaker_weight": 16,
    "second_speaker_weight": 16,
    "dominance_margin": 16,
    "dominant_speaker_share": 16,
    "short_fragment_source_utterance_count": 18,
}

QA_COLUMN_WIDTHS = {
    "qa_id": 18,
    "question_sentence_id": 20,
    "answer_sentence_id": 20,
    "context_sentence_id": 20,
    "question_text": 80,
    "answer_text": 80,
    "context_text": 72,
    "question_sentence_ids": 24,
    "answer_sentence_ids": 24,
    "context_sentence_ids": 24,
    "context_strategy": 24,
    "context_confidence": 18,
    "confidence": 20,
    "review_flag_summary": 34,
    "review_flags": 36,
    "human_comment": 56,
    "id": 18,
    "start": 13,
    "end": 13,
    "duration_s": 13,
    "question_type": 18,
    "didactic_question_score": 16,
    "confidence_score": 14,
    "confidence_label": 16,
    "question_score": 14,
    "answer_score": 14,
    "base_confidence": 14,
    "has_answer": 12,
    "answer_is_question": 14,
    "deferred_answer_search_used": 18,
    "review_priority": 16,
    "input_layer": 20,
    "question_speaker_role": 16,
    "answer_speaker_role": 16,
    "question_segment_id": 18,
    "answer_segment_id": 18,
    "segment_relation": 18,
    "requested_search_strategy": 18,
    "effective_search_strategy": 18,
    "requested_ranking_strategy": 18,
    "effective_ranking_strategy": 20,
    "search_stop_reason": 20,
    "search_fallback_reason": 20,
    "search_backend_error": 22,
    "ranking_fallback_reason": 22,
    "search_backend_status": 18,
    "search_model_name": 24,
    "search_backend": 22,
    "semantic_similarity": 16,
    "candidate_channel": 18,
    "reranking_model_name": 24,
    "reranking_backend": 24,
    "semantic_relevance_score": 18,
    "combined_rank_score": 18,
    "rank_position": 14,
    "speaker_influence": 18,
    "question_timing_source": 18,
    "answer_timing_source": 18,
    "question_source_utterance_ids": 26,
    "answer_source_utterance_ids": 26,
    "context_source_utterance_ids": 26,
    "question_unit_ids": 22,
    "answer_unit_ids": 22,
    "source_segment_ids": 24,
    "reason_codes": 36,
    "notes": 72,
}

REVIEW_CANDIDATE_COLUMN_WIDTHS = {
    "candidate_id": 24,
    "kind": 12,
    "review_reason": 28,
    "text": 90,
    "suggested_action": 22,
    "human_comment": 56,
    "id": 24,
    "start": 13,
    "review_priority": 14,
    "speaker": 16,
    "segment_id": 18,
    "quality": 18,
    "review_flags": 36,
    "notes": 42,
}


def export_run_to_excel(json_path: str | Path, excel_path: str | Path) -> None:
    """Export serialized utterances from one pipeline JSON artifact to Excel.

    Args:
        json_path: Path to an existing pipeline JSON output file.
        excel_path: Destination path for the generated `.xlsx` workbook.

    Returns:
        None. The function writes the Excel file to disk.
    """

    resolved_json_path = Path(json_path).expanduser().resolve()
    resolved_excel_path = Path(excel_path).expanduser().resolve()

    existing_comments = _load_existing_human_comments(resolved_excel_path)
    payload = json.loads(resolved_json_path.read_text(encoding="utf-8"))
    provenance_validation = validate_sentence_provenance(
        utterances=_list_of_dicts(payload.get("utterances")),
        sentences=_list_of_dicts(payload.get("sentences")),
    )
    qa_rows = _build_qa_rows(
        payload,
        existing_comments=existing_comments.get("QACandidates", {}),
    )
    sentence_qa_debug_lookup = _build_sentence_qa_debug_lookup(qa_rows)
    utterance_rows = _build_utterance_rows(
        payload,
        provenance_validation=provenance_validation,
    )
    sentence_rows = _build_sentence_rows(
        payload,
        provenance_validation=provenance_validation,
        qa_debug_lookup=sentence_qa_debug_lookup,
        existing_comments=existing_comments.get("Sentences", {}),
    )
    review_candidate_rows = _build_review_candidate_rows(
        utterance_rows,
        sentence_rows,
        existing_comments=existing_comments.get("ReviewCandidates", {}),
    )
    sentence_metric_rows = _build_sentence_metric_rows(
        sentence_rows,
        provenance_validation=provenance_validation,
    )
    summary_rows = _build_summary_rows(
        payload,
        utterance_rows=utterance_rows,
        sentence_rows=sentence_rows,
        review_candidate_rows=review_candidate_rows,
        provenance_validation=provenance_validation,
    )

    workbook = Workbook()
    utterances_worksheet = workbook.active
    utterances_worksheet.title = "Utterances"
    sentences_worksheet = workbook.create_sheet(title="Sentences")
    qa_worksheet = workbook.create_sheet(title="QACandidates")
    review_candidates_worksheet = workbook.create_sheet(title="ReviewCandidates")
    sentence_metrics_worksheet = workbook.create_sheet(title="SentenceMetrics")
    summary_worksheet = workbook.create_sheet(title="run_summary")

    _populate_utterances_sheet(utterances_worksheet, utterance_rows)
    _populate_sentences_sheet(sentences_worksheet, sentence_rows)
    _populate_qa_sheet(qa_worksheet, qa_rows)
    _populate_review_candidates_sheet(
        review_candidates_worksheet,
        review_candidate_rows,
    )
    _populate_sentence_metrics_sheet(
        sentence_metrics_worksheet,
        sentence_metric_rows,
    )
    _populate_summary_sheet(summary_worksheet, summary_rows)

    resolved_excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(resolved_excel_path)


def _populate_utterances_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Populate the utterances sheet and apply formatting."""

    worksheet.append(list(HEADERS))
    _style_header_row(worksheet)

    for row in rows:
        worksheet.append([row[header] for header in HEADERS])

    _format_utterances_sheet(worksheet)


def _populate_summary_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Populate the run summary sheet and apply formatting."""

    worksheet.append(list(SUMMARY_HEADERS))
    _style_header_row(worksheet)

    for row in rows:
        worksheet.append(
            [
                row["section"],
                row["metric"],
                row["value"],
                row["value2"],
                row["notes"],
            ],
        )

    _format_summary_sheet(worksheet, rows)


def _populate_sentences_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Populate the sentences sheet and apply formatting."""

    worksheet.append(list(SENTENCE_HEADERS))
    _style_header_row(worksheet)

    for row in rows:
        worksheet.append([row[header] for header in SENTENCE_HEADERS])

    _format_sentences_sheet(worksheet)


def _populate_qa_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Populate the QA candidate sheet and apply formatting."""

    worksheet.append(list(QA_HEADERS))
    _style_header_row(worksheet)

    for row in rows:
        worksheet.append([row[header] for header in QA_HEADERS])

    _format_qa_sheet(worksheet)


def _populate_review_candidates_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Populate the review-candidate sheet and apply formatting."""

    worksheet.append(list(REVIEW_CANDIDATE_HEADERS))
    _style_header_row(worksheet)

    for row in rows:
        worksheet.append([row[header] for header in REVIEW_CANDIDATE_HEADERS])

    _format_review_candidates_sheet(worksheet)


def _populate_sentence_metrics_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Populate the aggregate sentence-metrics worksheet."""

    worksheet.append(list(SENTENCE_METRIC_HEADERS))
    _style_header_row(worksheet)

    for row in rows:
        worksheet.append(
            [
                row["metric_name"],
                row["metric_value"],
                row["metric_unit"],
                row["metric_group"],
                row["notes"],
            ],
        )

    _format_sentence_metrics_sheet(worksheet, rows)


def _load_existing_human_comments(excel_path: Path) -> dict[str, dict[str, str]]:
    """Return existing manual comments keyed by sheet and row identifier."""

    if not excel_path.exists():
        return {}

    try:
        workbook = load_workbook(excel_path, data_only=True)
    except Exception:
        return {}

    return {
        "Sentences": _extract_sheet_human_comments(
            workbook,
            "Sentences",
            key_headers=("sentence_id", "id"),
        ),
        "QACandidates": _extract_sheet_human_comments(
            workbook,
            "QACandidates",
            key_headers=("qa_id", "id"),
        ),
        "ReviewCandidates": _extract_sheet_human_comments(
            workbook,
            "ReviewCandidates",
            key_headers=("candidate_id", "id"),
        ),
    }


def _extract_sheet_human_comments(
    workbook,
    sheet_name: str,
    *,
    key_headers: Sequence[str],
) -> dict[str, str]:
    """Return preserved manual comments from one existing worksheet."""

    if sheet_name not in workbook.sheetnames:
        return {}

    worksheet = workbook[sheet_name]
    if worksheet.max_row < 2:
        return {}

    header_index = {
        str(cell.value).strip().lower(): index
        for index, cell in enumerate(worksheet[1], start=1)
        if str(cell.value or "").strip()
    }
    comment_index = _find_first_header_index(header_index, COMMENT_HEADER_ALIASES)
    key_index = _find_first_header_index(header_index, key_headers)
    if comment_index is None or key_index is None:
        return {}

    comments: dict[str, str] = {}
    for row_index in range(2, worksheet.max_row + 1):
        key_value = str(worksheet.cell(row_index, key_index).value or "").strip()
        comment_value = str(worksheet.cell(row_index, comment_index).value or "").strip()
        if key_value and comment_value:
            comments[key_value] = comment_value
    return comments


def _find_first_header_index(
    header_index: dict[str, int],
    header_candidates: Sequence[str],
) -> int | None:
    """Return the first available header index from a list of aliases."""

    for header in header_candidates:
        normalized_header = str(header or "").strip().lower()
        if normalized_header in header_index:
            return header_index[normalized_header]
    return None


def _format_time_range(
    start_seconds: float | None,
    end_seconds: float | None,
) -> str:
    """Return a compact human-readable time range."""

    if start_seconds is None and end_seconds is None:
        return ""
    if end_seconds is None:
        return _format_clock_value(start_seconds)
    return (
        f"{_format_clock_value(start_seconds)} - "
        f"{_format_clock_value(end_seconds)}"
    )


def _format_clock_value(value: float | None) -> str:
    """Return an Excel-like clock string with millisecond precision."""

    if value is None:
        return ""

    total_milliseconds = max(0, int(round(value * 1000)))
    hours, remainder = divmod(total_milliseconds, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    seconds, milliseconds = divmod(remainder, 1000)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{milliseconds:03d}"


def _human_sentence_reference(sentence_ids: Sequence[str]) -> str:
    """Return a compact sentence-id summary for human review columns."""

    normalized_ids = [str(sentence_id).strip() for sentence_id in sentence_ids if str(sentence_id).strip()]
    if not normalized_ids:
        return ""
    return ", ".join(_compact_identifier(sentence_id) for sentence_id in normalized_ids)


def _compact_identifier(identifier: str) -> str:
    """Return a shortened identifier suitable for compact worksheet columns."""

    normalized_identifier = str(identifier or "").strip()
    if not normalized_identifier:
        return ""
    if "_sentence_" in normalized_identifier:
        return normalized_identifier.split("_sentence_", maxsplit=1)[-1]
    if "_utterance_" in normalized_identifier:
        return normalized_identifier.split("_utterance_", maxsplit=1)[-1]
    if "_" in normalized_identifier:
        return normalized_identifier.rsplit("_", maxsplit=1)[-1]
    return normalized_identifier


def _format_qa_confidence(
    *,
    confidence_label: str,
    confidence_score: float | None,
) -> str:
    """Return a compact confidence summary for QA review."""

    normalized_label = str(confidence_label or "").strip()
    if normalized_label and confidence_score is not None:
        return f"{normalized_label} ({confidence_score:.2f})"
    if normalized_label:
        return normalized_label
    if confidence_score is not None:
        return f"{confidence_score:.2f}"
    return ""


def _summarize_review_flags(review_flags: Sequence[str] | str) -> str:
    """Return a short review-flag summary suitable for manual scanning."""

    flags = _split_delimited_values(review_flags)
    if not flags:
        return ""
    return " | ".join(flags[:3])


def _build_review_reason(*, quality: str, review_flags: str) -> str:
    """Return a compact human-facing reason string for review rows."""

    reason_parts: list[str] = []
    normalized_quality = str(quality or "").strip()
    if normalized_quality:
        reason_parts.append(normalized_quality.upper())
    flags = _split_delimited_values(review_flags)
    if flags:
        reason_parts.append(", ".join(flags[:2]))
    return " | ".join(reason_parts)


def _suggest_review_action(
    *,
    kind: str,
    review_priority: str,
    review_flags: str,
) -> str:
    """Return a short suggested action for one review candidate."""

    flags = set(_split_delimited_values(review_flags))
    normalized_kind = str(kind or "").strip()
    normalized_priority = str(review_priority or "").strip().lower()

    if "merge_risky" in flags or "semantic_run_on" in flags or "semantic_fragment" in flags:
        return "check_split"
    if "uncertain_source" in flags or "uncertain_speaker" in flags:
        return "verify_speaker"
    if "segment_missing" in flags or "provenance_overlap" in flags:
        return "check_alignment"
    if normalized_kind == "utterance" and "speaker_change" in flags:
        return "check_turn_boundary"
    if normalized_priority == "high":
        return "review_now"
    if normalized_priority == "medium":
        return "review_if_time"
    return "monitor"


def _build_sentence_qa_debug_lookup(
    qa_rows: Sequence[dict[str, Any]],
) -> dict[str, str]:
    """Return one compact QA debug summary per sentence id."""

    lookup: dict[str, list[str]] = {}
    for row in qa_rows:
        qa_id = str(row.get("id") or "").strip()
        if not qa_id:
            continue

        question_ids = _split_delimited_values(row.get("question_sentence_ids"))
        answer_ids = _split_delimited_values(row.get("answer_sentence_ids"))
        context_ids = _split_delimited_values(row.get("context_sentence_ids"))
        question_set = set(question_ids)
        answer_set = set(answer_ids)
        shared_ids = question_set & answer_set

        for sentence_id in sorted(shared_ids):
            _append_sentence_qa_summary(
                lookup,
                sentence_id,
                _build_sentence_qa_summary(
                    mode="question+answer",
                    qa_id=qa_id,
                    counterpart_ids=[],
                    context_ids=context_ids,
                    context_strategy=str(row.get("context_strategy") or ""),
                    deferred_answer_search_used=bool(
                        row.get("deferred_answer_search_used"),
                    ),
                    confidence_label=str(row.get("confidence_label") or ""),
                    review_flags=str(row.get("review_flags") or ""),
                ),
            )
        for sentence_id in question_ids:
            if sentence_id in shared_ids:
                continue
            _append_sentence_qa_summary(
                lookup,
                sentence_id,
                _build_sentence_qa_summary(
                    mode="question",
                    qa_id=qa_id,
                    counterpart_ids=answer_ids,
                    context_ids=context_ids,
                    context_strategy=str(row.get("context_strategy") or ""),
                    deferred_answer_search_used=bool(
                        row.get("deferred_answer_search_used"),
                    ),
                    confidence_label=str(row.get("confidence_label") or ""),
                    review_flags=str(row.get("review_flags") or ""),
                ),
            )
        for sentence_id in answer_ids:
            if sentence_id in shared_ids:
                continue
            _append_sentence_qa_summary(
                lookup,
                sentence_id,
                _build_sentence_qa_summary(
                    mode="answer",
                    qa_id=qa_id,
                    counterpart_ids=question_ids,
                    context_ids=context_ids,
                    context_strategy=str(row.get("context_strategy") or ""),
                    deferred_answer_search_used=bool(
                        row.get("deferred_answer_search_used"),
                    ),
                    confidence_label=str(row.get("confidence_label") or ""),
                    review_flags=str(row.get("review_flags") or ""),
                ),
            )
        for sentence_id in context_ids:
            if sentence_id in question_set or sentence_id in answer_set:
                continue
            _append_sentence_qa_summary(
                lookup,
                sentence_id,
                _build_sentence_qa_summary(
                    mode="context",
                    qa_id=qa_id,
                    counterpart_ids=question_ids,
                    context_ids=context_ids,
                    context_strategy=str(row.get("context_strategy") or ""),
                    deferred_answer_search_used=bool(
                        row.get("deferred_answer_search_used"),
                    ),
                    confidence_label=str(row.get("confidence_label") or ""),
                    review_flags=str(row.get("review_flags") or ""),
                ),
            )

    return {
        sentence_id: " || ".join(summary_parts[:2])
        for sentence_id, summary_parts in lookup.items()
    }


def _append_sentence_qa_summary(
    lookup: dict[str, list[str]],
    sentence_id: str,
    summary: str,
) -> None:
    """Append a QA summary to one sentence without duplicates."""

    if not sentence_id or not summary:
        return
    bucket = lookup.setdefault(sentence_id, [])
    if summary not in bucket:
        bucket.append(summary)


def _build_sentence_qa_summary(
    *,
    mode: str,
    qa_id: str,
    counterpart_ids: Sequence[str],
    context_ids: Sequence[str],
    context_strategy: str,
    deferred_answer_search_used: bool,
    confidence_label: str,
    review_flags: str,
) -> str:
    """Return a compact QA summary string for one sentence."""

    parts: list[str] = []
    warnings = set(_split_delimited_values(review_flags))

    if mode == "question+answer":
        parts.extend(["QUESTION+ANSWER", qa_id])
        if context_strategy == "intra_sentence_context":
            parts.append("intra-sentence")
    elif mode == "question":
        parts.extend(["QUESTION", qa_id])
        if deferred_answer_search_used:
            parts.append("deferred answer")
        context_label = _context_reference_label(
            context_ids=context_ids,
            context_strategy=context_strategy,
        )
        if context_label:
            parts.append(f"context: {context_label}")
        counterpart_reference = _human_sentence_reference(counterpart_ids)
        if counterpart_reference:
            parts.append(f"answer: {counterpart_reference}")
    elif mode == "answer":
        parts.extend(["ANSWER", f"for {qa_id}"])
        counterpart_reference = _human_sentence_reference(counterpart_ids)
        if counterpart_reference:
            parts.append(f"from question: {counterpart_reference}")
    elif mode == "context":
        parts.extend(["CONTEXT", f"for {qa_id}"])
        context_label = _context_reference_label(
            context_ids=context_ids,
            context_strategy=context_strategy,
        )
        if context_label:
            parts.append(context_label)

    normalized_confidence = str(confidence_label or "").strip()
    if normalized_confidence:
        parts.append(f"confidence: {normalized_confidence}")

    compact_warning_labels = []
    warning_map = {
        "missing_answer": "missing_answer",
        "low_confidence": "low_confidence",
        "medium_confidence": "medium_confidence",
        "speaker_conflict": "speaker_penalty",
        "answer_grounding_missing": "answer_grounding",
        "question_grounding_missing": "question_grounding",
        "fallback_input_layer": "fallback_layer",
    }
    for source_flag, label in warning_map.items():
        if source_flag in warnings:
            compact_warning_labels.append(label)
    if compact_warning_labels:
        parts.append(", ".join(compact_warning_labels[:2]))

    return " | ".join(parts)


def _context_reference_label(
    *,
    context_ids: Sequence[str],
    context_strategy: str,
) -> str:
    """Return one short human-facing context label for QA debug summaries."""

    normalized_strategy = str(context_strategy or "").strip()
    if normalized_strategy == "previous_sentence_context":
        return "previous sentence"
    if normalized_strategy == "intra_sentence_context":
        return "intra-sentence"

    context_reference = _human_sentence_reference(context_ids)
    if context_reference:
        return context_reference
    return normalized_strategy.replace("_", " ").strip()


def _build_utterance_rows(
    payload: dict[str, Any],
    *,
    provenance_validation: SentenceProvenanceValidation,
) -> list[dict[str, Any]]:
    """Return Excel-ready row dictionaries built from a session JSON payload."""

    utterances = payload.get("utterances", [])
    sentences = _list_of_dicts(payload.get("sentences"))
    segments = payload.get("segments", [])
    if not isinstance(utterances, list):
        utterances = []
    if not isinstance(segments, list):
        segments = []

    segment_lookup = _build_segment_lookup(segments)
    sentence_segment_lookup = _build_segment_lookup_for_field(
        segments,
        lookup_field="sentence_ids",
    )
    ordered_utterances = sorted(utterances, key=_utterance_sort_key)

    rows: list[dict[str, Any]] = []
    previous_speaker = ""
    for utterance in ordered_utterances:
        if not isinstance(utterance, dict):
            continue
        utterance_id = str(utterance.get("utterance_id") or "").strip()
        utterance_key = scoped_utterance_key(
            utterance.get("audio_source_id"),
            utterance_id,
        )
        linked_sentence_ids = provenance_validation.utterance_to_sentence_ids.get(
            utterance_key,
            [],
        )
        sentence_assignment_count = len(linked_sentence_ids)
        has_sentence_mapping_conflict = sentence_assignment_count > 1
        segment_ids = _resolve_segment_ids(
            utterance,
            segment_lookup,
            sentence_segment_lookup=sentence_segment_lookup,
            utterance_sentence_lookup=provenance_validation.utterance_to_sentence_ids,
        )
        segment_is_missing = not segment_ids
        speaker = str(utterance.get("speaker_id") or "").strip()
        speaker_status = _resolve_speaker_status(utterance)
        speaker_change = bool(
            previous_speaker and speaker and speaker != previous_speaker
        )
        review_flags = _build_review_flags(
            speaker=speaker,
            speaker_status=speaker_status,
            speaker_change=speaker_change,
            segment_is_missing=segment_is_missing,
            has_sentence_mapping_conflict=has_sentence_mapping_conflict,
            utterance=utterance,
        )
        review_priority = _utterance_review_priority(
            review_flags=review_flags,
        )
        notes = _build_notes(
            utterance,
            segment_ids=segment_ids,
            utterance_sentence_lookup=provenance_validation.utterance_to_sentence_ids,
        )
        display_segment_ids = segment_ids or ["UNASSIGNED_SEGMENT"]
        is_audio_quality_degraded = _is_audio_quality_degraded(utterance)

        rows.append(
            {
                "id": str(utterance.get("utterance_id") or ""),
                "start": _excel_duration(_utterance_time(utterance, "start")),
                "end": _excel_duration(_utterance_time(utterance, "end")),
                "duration_s": _excel_duration(_utterance_duration(utterance)),
                "speaker": speaker,
                "speaker_status": speaker_status,
                "is_uncertain_speaker": speaker_status == "uncertain",
                "is_unassigned_speaker": not speaker,
                "is_speaker_changed": speaker_change,
                "is_segment_missing": segment_is_missing,
                "is_audio_quality_degraded": is_audio_quality_degraded,
                "review_priority": review_priority,
                "segment_id": ", ".join(display_segment_ids),
                "final_sentence_id": ", ".join(linked_sentence_ids),
                "sentence_assignment_count": sentence_assignment_count,
                "has_sentence_mapping_conflict": has_sentence_mapping_conflict,
                "text": str(utterance.get("text") or ""),
                "word_count": _word_count(utterance),
                "review_flags": ", ".join(review_flags),
                "notes": notes,
            },
        )

        if speaker:
            previous_speaker = speaker

    return rows


def _build_summary_rows(
    payload: dict[str, Any],
    *,
    utterance_rows: list[dict[str, Any]],
    sentence_rows: list[dict[str, Any]],
    review_candidate_rows: list[dict[str, Any]],
    provenance_validation: SentenceProvenanceValidation,
) -> list[dict[str, Any]]:
    """Return structured summary rows with derived values and formats."""

    session_metadata = _dict_value(payload.get("session_metadata"))
    session_run_metadata = _dict_value(session_metadata.get("metadata"))
    transcript = _dict_value(payload.get("transcript"))
    input_sources = _list_of_dicts(payload.get("input_sources"))
    audio_sources = _list_of_dicts(payload.get("audio_sources"))
    utterances = _list_of_dicts(payload.get("utterances"))
    sentences = _list_of_dicts(payload.get("sentences"))
    segments = _list_of_dicts(payload.get("segments"))
    qa_candidates = _list_of_dicts(payload.get("qa_candidates"))
    speaker_role_estimates = _list_of_dicts(payload.get("speaker_role_estimates"))
    pipeline_timing = _dict_value(payload.get("pipeline_timing"))
    timing_summary = _dict_value(pipeline_timing.get("summary"))
    timing_stages = _list_of_dicts(pipeline_timing.get("stages"))

    transcript_word_count = _text_word_count(transcript.get("full_text"))
    utterance_word_count = sum(_word_count(utterance) for utterance in utterances)
    sentence_word_count = sum(_word_count(sentence) for sentence in sentences)
    utterance_speaker_count = len(
        {
            str(utterance.get("speaker_id") or "").strip()
            for utterance in utterances
            if str(utterance.get("speaker_id") or "").strip()
        },
    )
    utterance_durations = [
        duration
        for utterance in utterances
        if (duration := _utterance_duration(utterance)) is not None
    ]
    sentence_speaker_count = len(
        {
            str(sentence.get("speaker_id") or "").strip()
            for sentence in sentences
            if str(sentence.get("speaker_id") or "").strip()
        },
    )
    sentence_durations = [
        duration
        for sentence in sentences
        if (duration := _sentence_duration(sentence)) is not None
    ]
    utterance_segment_missing_count = sum(
        1 for row in utterance_rows if bool(row.get("is_segment_missing"))
    )
    sentence_segment_missing_count = sum(
        1
        for row in sentence_rows
        if "UNASSIGNED_SEGMENT" in str(row.get("segment_id") or "")
    )
    high_priority_utterance_count = sum(
        1 for row in utterance_rows if row.get("review_priority") == "high"
    )
    high_priority_sentence_count = sum(
        1 for row in sentence_rows if row.get("review_priority") == "high"
    )
    medium_priority_sentence_count = sum(
        1 for row in sentence_rows if row.get("review_priority") == "medium"
    )
    fragment_sentence_count = sum(
        1
        for row in sentence_rows
        if row.get("semantic_quality_label") == "fragment"
    )
    run_on_sentence_count = sum(
        1
        for row in sentence_rows
        if row.get("semantic_quality_label") == "run_on"
    )
    utterances_assigned_to_multiple_sentences_count = (
        provenance_validation.utterances_assigned_to_multiple_sentences
    )
    sentences_with_shared_source_utterances_count = (
        provenance_validation.sentences_with_provenance_overlap_count
    )
    max_sentence_reuse_per_utterance = (
        provenance_validation.max_sentence_reuse_per_utterance
    )
    duplicate_source_utterance_id_count = (
        provenance_validation.duplicate_source_utterance_id_count
    )
    utterance_without_sentence_count = provenance_validation.utterance_without_sentence_count
    provenance_anomaly_count = provenance_validation.provenance_anomaly_count
    sentence_assignment_total = provenance_validation.sentence_assignment_total
    all_sentences_have_provenance_overlap = (
        provenance_validation.all_sentences_have_provenance_overlap
    )

    input_filenames = [
        str(
            source.get("original_filename")
            or Path(str(source.get("original_path") or "")).name,
        )
        for source in input_sources
    ]

    total_input_duration_seconds = _sum_duration(input_sources)
    total_normalized_audio_duration_seconds = _sum_duration(audio_sources)
    total_pipeline_duration_seconds = _safe_float(
        timing_summary.get("total_duration_seconds"),
    )
    most_expensive_stage_duration_seconds = _safe_float(
        timing_summary.get("most_expensive_stage_duration_seconds"),
    )
    segment_count = len(segments)
    utterance_count = len(utterances)
    sentence_count = len(sentences)
    qa_candidate_count = len(qa_candidates)
    qa_pairing_debug_rows = [
        _dict_value(_dict_value(candidate.get("metadata")).get("pairing_debug"))
        for candidate in qa_candidates
    ]
    qa_answer_debug_rows = [
        _dict_value(_dict_value(candidate.get("metadata")).get("answer_debug"))
        for candidate in qa_candidates
    ]
    qa_ranking_debug_rows = [
        _dict_value(answer_debug.get("ranking_debug"))
        for answer_debug in qa_answer_debug_rows
    ]
    qa_search_strategies = _unique_non_empty_strings(
        [
            str(
                pairing_debug.get("effective_search_strategy")
                or pairing_debug.get("search_strategy")
                or "",
            ).strip()
            for pairing_debug in qa_pairing_debug_rows
        ],
    )
    qa_ranking_strategies = _unique_non_empty_strings(
        [
            str(
                pairing_debug.get("effective_ranking_strategy")
                or pairing_debug.get("ranking_strategy")
                or "",
            ).strip()
            for pairing_debug in qa_pairing_debug_rows
        ],
    )
    qa_search_models = _unique_non_empty_strings(
        [
            str(pairing_debug.get("semantic_model_name") or "").strip()
            for pairing_debug in qa_pairing_debug_rows
        ],
    )
    qa_search_fallback_reasons = _unique_non_empty_strings(
        [
            str(pairing_debug.get("search_fallback_reason") or "").strip()
            for pairing_debug in qa_pairing_debug_rows
        ],
    )
    qa_search_backend_errors = _unique_non_empty_strings(
        [
            str(pairing_debug.get("search_backend_error") or "").strip()
            for pairing_debug in qa_pairing_debug_rows
        ],
    )
    qa_ranking_fallback_reasons = _unique_non_empty_strings(
        [
            str(pairing_debug.get("ranking_fallback_reason") or "").strip()
            for pairing_debug in qa_pairing_debug_rows
        ],
    )
    qa_reranking_models = _unique_non_empty_strings(
        [
            str(
                pairing_debug.get("semantic_reranking_model_name")
                or ranking_debug.get("semantic_reranking_model_name")
                or "",
            ).strip()
            for pairing_debug, ranking_debug in zip(
                qa_pairing_debug_rows,
                qa_ranking_debug_rows,
            )
        ],
    )
    qa_extraction_stage = _find_timing_stage(timing_stages, "qa_extraction")
    qa_extraction_duration_seconds = _safe_float(
        qa_extraction_stage.get("duration_seconds"),
    )
    utterances_with_speaker = _safe_float(
        transcript.get("utterance_with_speaker_count"),
        fallback=0.0,
    )
    uncertain_utterances = _safe_float(
        transcript.get("utterance_uncertain_speaker_count"),
        fallback=0.0,
    )
    total_utterance_duration_seconds = sum(utterance_durations) if utterance_durations else None
    sentences_with_speaker = _safe_float(
        transcript.get("sentence_with_speaker_count"),
        fallback=0.0,
    )
    total_sentence_duration_seconds = sum(sentence_durations) if sentence_durations else None

    rows = [
        _summary_row("session", "session_id", session_metadata.get("session_id")),
        _summary_row(
            "session",
            "schema_version",
            session_metadata.get("schema_version"),
        ),
        _summary_row(
            "session",
            "processing_status",
            session_metadata.get("processing_status"),
        ),
        _summary_row(
            "session",
            "pipeline_execution_mode",
            session_run_metadata.get("pipeline_execution_mode"),
        ),
        _summary_row(
            "session",
            "pipeline_run_profile_label",
            session_run_metadata.get("pipeline_run_profile_label"),
        ),
        _summary_row(
            "session",
            "language_codes",
            _join_values(session_metadata.get("language_codes", [])),
        ),
        _summary_row("input", "input_file_count", len(input_sources), value_kind="int"),
        _summary_row("input", "input_filenames", _join_values(input_filenames)),
        _summary_row(
            "input",
            "total_input_duration",
            _excel_duration(total_input_duration_seconds),
            value_kind="duration",
            notes="Sum of available original input durations.",
        ),
        _summary_row(
            "input",
            "total_normalized_audio_duration",
            _excel_duration(total_normalized_audio_duration_seconds),
            value_kind="duration",
            value2=_safe_ratio(
                total_normalized_audio_duration_seconds,
                total_input_duration_seconds,
            ),
            value2_kind="percent",
            notes="Normalized audio duration / original input duration.",
        ),
        _summary_row(
            "transcript",
            "transcript_word_count",
            transcript_word_count,
            value_kind="int",
        ),
        _summary_row(
            "transcript",
            "chunk_count",
            transcript.get("chunk_count"),
            value_kind="int",
        ),
        _summary_row(
            "transcript",
            "aligned_source_count",
            transcript.get("aligned_source_count"),
            value_kind="int",
        ),
        _summary_row(
            "transcript",
            "aligned_segment_count",
            transcript.get("aligned_segment_count"),
            value_kind="int",
        ),
        _summary_row(
            "transcript",
            "aligned_word_count",
            transcript.get("aligned_word_count"),
            value_kind="int",
        ),
        _summary_row(
            "transcript",
            "merged_unit_count",
            transcript.get("merged_unit_count"),
            value_kind="int",
        ),
        _summary_row(
            "transcript",
            "detected_languages",
            _join_values(transcript.get("detected_languages", [])),
        ),
        _summary_row(
            "transcript",
            "transcription_cache_enabled",
            session_run_metadata.get("transcription_cache_enabled"),
        ),
        _summary_row(
            "transcript",
            "transcription_cache_lookup_performed",
            session_run_metadata.get("transcription_cache_lookup_performed"),
        ),
        _summary_row(
            "transcript",
            "transcription_cache_hit",
            session_run_metadata.get("transcription_cache_hit"),
        ),
        _summary_row(
            "transcript",
            "transcription_recomputed",
            session_run_metadata.get("transcription_recomputed"),
        ),
        _summary_row(
            "transcript",
            "transcription_forced_recompute",
            session_run_metadata.get("transcription_forced_recompute"),
        ),
        _summary_row("debug", "segment_count", segment_count, value_kind="int"),
        _summary_row("debug", "utterance_count", utterance_count, value_kind="int"),
        _summary_row("debug", "sentence_count", sentence_count, value_kind="int"),
        _summary_row(
            "debug",
            "utterance_word_count_total",
            utterance_word_count,
            value_kind="int",
        ),
        _summary_row(
            "debug",
            "sentence_word_count_total",
            sentence_word_count,
            value_kind="int",
        ),
        _summary_row(
            "debug",
            "utterance_speaker_count",
            utterance_speaker_count,
            value_kind="int",
            notes="Distinct speaker ids found on utterances.",
        ),
        _summary_row(
            "debug",
            "diarization_speaker_count",
            transcript.get("speaker_count"),
            value_kind="int",
            notes="Speaker count derived from diarization segments.",
        ),
        _summary_row(
            "debug",
            "sentence_speaker_count",
            sentence_speaker_count,
            value_kind="int",
            notes="Distinct speaker ids found on sentences.",
        ),
        _summary_row(
            "debug",
            "utterances_with_speaker",
            transcript.get("utterance_with_speaker_count"),
            value_kind="int",
            value2=_safe_ratio(utterances_with_speaker, utterance_count),
            value2_kind="percent",
            notes="Speaker coverage across utterances.",
        ),
        _summary_row(
            "debug",
            "sentences_with_speaker",
            transcript.get("sentence_with_speaker_count"),
            value_kind="int",
            value2=_safe_ratio(sentences_with_speaker, sentence_count),
            value2_kind="percent",
            notes="Speaker coverage across sentences.",
        ),
        _summary_row(
            "debug",
            "uncertain_utterances",
            transcript.get("utterance_uncertain_speaker_count"),
            value_kind="int",
            value2=_safe_ratio(uncertain_utterances, utterance_count),
            value2_kind="percent",
            notes="Share of utterances marked as uncertain.",
        ),
        _summary_row(
            "debug",
            "segment_missing_utterances",
            utterance_segment_missing_count,
            value_kind="int",
            value2=_safe_ratio(utterance_segment_missing_count, utterance_count),
            value2_kind="percent",
            notes="Utterances with no resolved segment mapping.",
        ),
        _summary_row(
            "debug",
            "segment_missing_sentences",
            sentence_segment_missing_count,
            value_kind="int",
            value2=_safe_ratio(sentence_segment_missing_count, sentence_count),
            value2_kind="percent",
            notes="Sentences with no resolved segment mapping.",
        ),
        _summary_row(
            "debug",
            "high_priority_utterance_reviews",
            high_priority_utterance_count,
            value_kind="int",
            value2=_safe_ratio(high_priority_utterance_count, utterance_count),
            value2_kind="percent",
            notes="Utterances flagged as high review priority.",
        ),
        _summary_row(
            "debug",
            "high_priority_sentence_reviews",
            high_priority_sentence_count,
            value_kind="int",
            value2=_safe_ratio(high_priority_sentence_count, sentence_count),
            value2_kind="percent",
            notes="Sentences flagged as high review priority.",
        ),
        _summary_row(
            "debug",
            "medium_priority_sentence_reviews",
            medium_priority_sentence_count,
            value_kind="int",
            value2=_safe_ratio(medium_priority_sentence_count, sentence_count),
            value2_kind="percent",
            notes="Sentences flagged as medium review priority.",
        ),
        _summary_row(
            "debug",
            "fragment_sentence_count",
            fragment_sentence_count,
            value_kind="int",
            value2=_safe_ratio(fragment_sentence_count, sentence_count),
            value2_kind="percent",
            notes="Sentences classified as semantic fragments.",
        ),
        _summary_row(
            "debug",
            "run_on_sentence_count",
            run_on_sentence_count,
            value_kind="int",
            value2=_safe_ratio(run_on_sentence_count, sentence_count),
            value2_kind="percent",
            notes="Sentences classified as semantic run-ons.",
        ),
        _summary_row(
            "debug",
            "review_candidate_count",
            len(review_candidate_rows),
            value_kind="int",
            notes="Rows emitted in the ReviewCandidates sheet.",
        ),
        _summary_row(
            "debug",
            "utterances_assigned_to_multiple_sentences",
            utterances_assigned_to_multiple_sentences_count,
            value_kind="int",
            value2=_safe_ratio(
                utterances_assigned_to_multiple_sentences_count,
                utterance_count,
            ),
            value2_kind="percent",
            notes=_join_notes(
                "Utterances linked to more than one final sentence.",
                provenance_validation.mapping_conflict_examples,
            ),
        ),
        _summary_row(
            "debug",
            "sentences_with_shared_source_utterances",
            sentences_with_shared_source_utterances_count,
            value_kind="int",
            value2=_safe_ratio(
                sentences_with_shared_source_utterances_count,
                sentence_count,
            ),
            value2_kind="percent",
            notes=_join_notes(
                "Sentences that share at least one source utterance with another sentence.",
                provenance_validation.overlap_examples,
            ),
        ),
        _summary_row(
            "debug",
            "max_sentence_reuse_per_utterance",
            max_sentence_reuse_per_utterance,
            value_kind="int",
            notes="Maximum number of final sentences linked to one utterance.",
        ),
        _summary_row(
            "debug",
            "duplicate_source_utterance_id_count",
            duplicate_source_utterance_id_count,
            value_kind="int",
            notes=_join_notes(
                "Duplicate source utterance ids found inside sentence provenance.",
                provenance_validation.duplicate_source_examples,
            ),
        ),
        _summary_row(
            "debug",
            "utterances_without_final_sentence",
            utterance_without_sentence_count,
            value_kind="int",
            value2=_safe_ratio(utterance_without_sentence_count, utterance_count),
            value2_kind="percent",
            notes="Final utterances not linked to any sentence.",
        ),
        _summary_row(
            "debug",
            "sentence_assignment_total",
            sentence_assignment_total,
            value_kind="int",
            notes="Total source utterance assignments across final sentences.",
        ),
        _summary_row(
            "debug",
            "all_sentences_have_provenance_overlap",
            all_sentences_have_provenance_overlap,
            notes=(
                "True only when every final sentence overlaps with another sentence"
            ),
        ),
        _summary_row(
            "debug",
            "provenance_anomaly_count",
            provenance_anomaly_count,
            value_kind="int",
            notes=_join_notes(
                "Combined count of mapping conflicts, duplicate source ids, empty provenance and unknown references.",
                provenance_validation.unknown_reference_examples,
            ),
        ),
        _summary_row(
            "debug",
            "qa_candidate_count",
            qa_candidate_count,
            value_kind="int",
        ),
        _summary_row(
            "debug",
            "qa_search_strategies",
            _join_values(qa_search_strategies),
        ),
        _summary_row(
            "debug",
            "qa_ranking_strategies",
            _join_values(qa_ranking_strategies),
        ),
        _summary_row(
            "debug",
            "qa_search_models",
            _join_values(qa_search_models),
        ),
        _summary_row(
            "debug",
            "qa_search_fallback_reasons",
            _join_values(qa_search_fallback_reasons),
            notes="Unique QA search fallback reasons observed in final pairs.",
        ),
        _summary_row(
            "debug",
            "qa_search_backend_errors",
            _join_values(qa_search_backend_errors),
            notes="Unique semantic search backend errors propagated to QA debug.",
        ),
        _summary_row(
            "debug",
            "qa_reranking_models",
            _join_values(qa_reranking_models),
        ),
        _summary_row(
            "debug",
            "qa_ranking_fallback_reasons",
            _join_values(qa_ranking_fallback_reasons),
            notes="Unique QA ranking fallback reasons observed in final pairs.",
        ),
        _summary_row(
            "timing_summary",
            "qa_extraction_duration",
            _excel_duration(qa_extraction_duration_seconds),
            value_kind="duration",
            value2=_safe_ratio(
                qa_extraction_duration_seconds,
                total_pipeline_duration_seconds,
            ),
            value2_kind="percent",
            notes="Duration of the qa_extraction pipeline stage when available.",
        ),
        _summary_row(
            "debug",
            "speaker_role_estimate_count",
            len(speaker_role_estimates),
            value_kind="int",
        ),
        _summary_row(
            "derived",
            "avg_words_per_utterance",
            _safe_ratio(utterance_word_count, utterance_count),
            value_kind="float",
        ),
        _summary_row(
            "derived",
            "avg_words_per_sentence",
            _safe_ratio(sentence_word_count, sentence_count),
            value_kind="float",
        ),
        _summary_row(
            "derived",
            "avg_utterance_duration",
            _excel_duration(_safe_average(utterance_durations)),
            value_kind="duration",
            value2=_safe_ratio(
                total_utterance_duration_seconds,
                total_input_duration_seconds,
            ),
            value2_kind="percent",
            notes="Average utterance duration and utterance coverage over input time.",
        ),
        _summary_row(
            "derived",
            "avg_sentence_duration",
            _excel_duration(_safe_average(sentence_durations)),
            value_kind="duration",
            value2=_safe_ratio(
                total_sentence_duration_seconds,
                total_input_duration_seconds,
            ),
            value2_kind="percent",
            notes="Average sentence duration and sentence coverage over input time.",
        ),
        _summary_row(
            "derived",
            "avg_segment_duration",
            _excel_duration(_safe_ratio(total_input_duration_seconds, segment_count)),
            value_kind="duration",
            value2=_safe_ratio(utterance_count, segment_count),
            value2_kind="float",
            notes="Average segment duration and utterances per segment.",
        ),
        _summary_row(
            "derived",
            "avg_utterances_per_sentence",
            _safe_average(
                [
                    _safe_float(
                        _dict_value(sentence.get("metadata")).get("source_utterance_count"),
                    )
                    for sentence in sentences
                    if _safe_float(
                        _dict_value(sentence.get("metadata")).get("source_utterance_count"),
                    )
                    is not None
                ],
            ),
            value_kind="float",
        ),
        _summary_row(
            "derived",
            "sentences_per_segment",
            _safe_ratio(sentence_count, segment_count),
            value_kind="float",
        ),
        _summary_row(
            "derived",
            "words_per_minute",
            _per_minute(transcript_word_count, total_input_duration_seconds),
            value_kind="float",
            notes="Transcript throughput over the original input duration.",
        ),
        _summary_row(
            "derived",
            "utterances_per_minute",
            _per_minute(utterance_count, total_input_duration_seconds),
            value_kind="float",
        ),
        _summary_row(
            "derived",
            "sentences_per_minute",
            _per_minute(sentence_count, total_input_duration_seconds),
            value_kind="float",
        ),
        _summary_row(
            "derived",
            "segments_per_hour",
            _per_hour(segment_count, total_input_duration_seconds),
            value_kind="float",
        ),
        _summary_row(
            "derived",
            "qa_candidates_per_segment",
            _safe_ratio(qa_candidate_count, segment_count),
            value_kind="float",
        ),
        _summary_row(
            "timing_summary",
            "total_duration",
            _excel_duration(total_pipeline_duration_seconds),
            value_kind="duration",
            value2=_safe_ratio(
                total_pipeline_duration_seconds,
                total_input_duration_seconds,
            ),
            value2_kind="ratio",
            notes="Pipeline real-time factor over the original input duration.",
        ),
        _summary_row(
            "timing_summary",
            "stage_count",
            timing_summary.get("stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "pipeline_execution_mode",
            timing_summary.get("pipeline_execution_mode"),
        ),
        _summary_row(
            "timing_summary",
            "run_profile_label",
            timing_summary.get("run_profile_label"),
        ),
        _summary_row(
            "timing_summary",
            "executed_stage_count",
            timing_summary.get("executed_stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "reused_cache_stage_count",
            timing_summary.get("reused_cache_stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "reused_artifact_stage_count",
            timing_summary.get("reused_artifact_stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "forced_recompute_stage_count",
            timing_summary.get("forced_recompute_stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "completed_stage_count",
            timing_summary.get("completed_stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "skipped_stage_count",
            timing_summary.get("skipped_stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "failed_stage_count",
            timing_summary.get("failed_stage_count"),
            value_kind="int",
        ),
        _summary_row(
            "timing_summary",
            "any_cache_hit",
            timing_summary.get("any_cache_hit"),
        ),
        _summary_row(
            "timing_summary",
            "any_artifact_reuse",
            timing_summary.get("any_artifact_reuse"),
        ),
        _summary_row(
            "timing_summary",
            "full_recompute_requested",
            timing_summary.get("full_recompute_requested"),
        ),
        _summary_row(
            "timing_summary",
            "most_expensive_stage_name",
            timing_summary.get("most_expensive_stage_name"),
        ),
        _summary_row(
            "timing_summary",
            "most_expensive_stage_duration",
            _excel_duration(most_expensive_stage_duration_seconds),
            value_kind="duration",
            value2=_safe_ratio(
                most_expensive_stage_duration_seconds,
                total_pipeline_duration_seconds,
            ),
            value2_kind="percent",
        ),
    ]

    rows.extend(_build_input_detail_rows(input_sources))
    rows.extend(
        _build_timing_stage_rows(
            timing_stages,
            total_pipeline_duration_seconds=total_pipeline_duration_seconds,
            total_input_duration_seconds=total_input_duration_seconds,
        ),
    )
    return rows


def _build_sentence_rows(
    payload: dict[str, Any],
    *,
    provenance_validation: SentenceProvenanceValidation,
    qa_debug_lookup: dict[str, str],
    existing_comments: dict[str, str],
) -> list[dict[str, Any]]:
    """Return Excel-ready row dictionaries for reconstructed sentences."""

    sentences = _list_of_dicts(payload.get("sentences"))
    utterances = _list_of_dicts(payload.get("utterances"))
    segments = _list_of_dicts(payload.get("segments"))
    utterance_lookup = _build_utterance_lookup(utterances)
    sentence_segment_lookup = _build_segment_lookup_for_field(
        segments,
        lookup_field="sentence_ids",
    )
    ordered_sentences = sorted(sentences, key=_sentence_sort_key)

    rows: list[dict[str, Any]] = []
    for sentence in ordered_sentences:
        segment_ids = _resolve_unit_segment_ids(
            unit_id=str(sentence.get("sentence_id") or "").strip(),
            lookup=sentence_segment_lookup,
        )
        segment_is_missing = not segment_ids
        speaker = str(sentence.get("speaker_id") or "").strip()
        source_utterance_ids = _coerce_string_list(sentence.get("source_utterance_ids"))
        sentence_id = str(sentence.get("sentence_id") or "").strip()
        scoped_source_utterance_keys = provenance_validation.sentence_to_utterance_keys.get(
            sentence_id,
            [
                scoped_utterance_key(sentence.get("audio_source_id"), utterance_id)
                for utterance_id in source_utterance_ids
            ],
        )
        source_diagnostics = _source_utterance_diagnostics(
            source_utterance_keys=scoped_source_utterance_keys,
            utterance_lookup=utterance_lookup,
            sentence=sentence,
        )
        metadata = _dict_value(sentence.get("metadata"))
        speaker_evidence = _dict_value(metadata.get("speaker_evidence"))
        source_utterance_count = _safe_int(
            metadata.get("source_utterance_count"),
            fallback=len(source_utterance_ids),
        )
        semantic_quality_label = str(
            sentence.get("semantic_quality_label")
            or metadata.get("semantic_quality_label")
            or "good",
        ).strip()
        merge_safety_label = str(
            sentence.get("merge_safety_label")
            or metadata.get("merge_safety_label")
            or "safe",
        ).strip()
        review_flags = _build_sentence_review_flags(
            sentence=sentence,
            segment_is_missing=segment_is_missing,
            has_provenance_overlap=bool(
                provenance_validation.sentence_to_overlap_sentence_ids.get(sentence_id),
            ),
        )
        review_priority = _sentence_review_priority(
            sentence=sentence,
            review_flags=review_flags,
            segment_is_missing=segment_is_missing,
        )
        display_segment_ids = segment_ids or ["UNASSIGNED_SEGMENT"]
        rows.append(
            {
                "sentence_id": sentence_id,
                "time_range": _format_time_range(
                    _sentence_time(sentence, "start"),
                    _sentence_time(sentence, "end"),
                ),
                "qa_debug_summary": qa_debug_lookup.get(sentence_id, ""),
                "human_comment": existing_comments.get(sentence_id, ""),
                "id": str(sentence.get("sentence_id") or ""),
                "start": _excel_duration(_sentence_time(sentence, "start")),
                "end": _excel_duration(_sentence_time(sentence, "end")),
                "duration_s": _excel_duration(_sentence_duration(sentence)),
                "speaker": speaker,
                "detected_language": str(sentence.get("detected_language") or ""),
                "source_utterance_count": source_utterance_count,
                "source_utterance_span": _format_sentence_span(sentence),
                "is_multi_utterance": bool(metadata.get("is_multi_utterance"))
                or source_utterance_count > 1,
                "has_uncertain_source": bool(metadata.get("has_uncertain_source")),
                "has_unassigned_source": bool(metadata.get("has_unassigned_source")),
                "has_speaker_change_inside": bool(
                    metadata.get("has_speaker_change_inside"),
                ),
                "is_semantic_fragment": semantic_quality_label == "fragment",
                "is_semantic_run_on": semantic_quality_label == "run_on",
                "is_merge_risky": merge_safety_label == "risky",
                "semantic_quality_label": semantic_quality_label,
                "length_bucket": str(sentence.get("length_bucket") or ""),
                "duration_bucket": str(sentence.get("duration_bucket") or ""),
                "speaker_stability_label": str(
                    sentence.get("speaker_stability_label") or "",
                ),
                "merge_safety_label": merge_safety_label,
                "review_priority": review_priority,
                "segment_id": ", ".join(display_segment_ids),
                "text": str(sentence.get("text") or ""),
                "word_count": _word_count(sentence),
                "source_utterance_ids": ", ".join(str(item) for item in source_utterance_ids),
                "distinct_source_speaker_count": source_diagnostics[
                    "distinct_source_speaker_count"
                ],
                "source_speaker_ids": ", ".join(source_diagnostics["source_speaker_ids"]),
                "assigned_source_utterance_count": source_diagnostics[
                    "assigned_source_utterance_count"
                ],
                "uncertain_source_utterance_count": source_diagnostics[
                    "uncertain_source_utterance_count"
                ],
                "unassigned_source_utterance_count": source_diagnostics[
                    "unassigned_source_utterance_count"
                ],
                "speaker_resolution_status": str(
                    sentence.get("speaker_resolution_status") or "",
                ),
                "speaker_confidence_label": str(
                    sentence.get("speaker_confidence_label") or "",
                ),
                "has_provenance_overlap": bool(
                    provenance_validation.sentence_to_overlap_sentence_ids.get(sentence_id),
                ),
                "provenance_overlap_sentence_ids": ", ".join(
                    provenance_validation.sentence_to_overlap_sentence_ids.get(
                        sentence_id,
                        [],
                    ),
                ),
                "review_flags": ", ".join(review_flags),
                "notes": _build_sentence_notes(
                    sentence,
                    display_segment_ids,
                    source_diagnostics=source_diagnostics,
                    provenance_overlap_sentence_ids=provenance_validation.sentence_to_overlap_sentence_ids.get(
                        sentence_id,
                        [],
                    ),
                ),
                "speaker_assignment_method": str(
                    sentence.get("speaker_assignment_method")
                    or metadata.get("speaker_assignment_method")
                    or "",
                ),
                "dominant_speaker_weight": _safe_float(
                    metadata.get("dominant_speaker_weight"),
                    fallback=_safe_float(speaker_evidence.get("dominant_weight")),
                ),
                "second_speaker_weight": _safe_float(
                    metadata.get("second_speaker_weight"),
                    fallback=_safe_float(speaker_evidence.get("second_weight")),
                ),
                "dominance_margin": _safe_float(
                    metadata.get("dominance_margin"),
                    fallback=_safe_float(speaker_evidence.get("dominance_margin")),
                ),
                "dominant_speaker_share": _safe_float(
                    metadata.get("dominant_speaker_share"),
                    fallback=_safe_float(speaker_evidence.get("dominant_share")),
                ),
                "short_fragment_source_utterance_count": source_diagnostics[
                    "short_fragment_source_utterance_count"
                ],
            },
        )

    return rows


def _build_qa_rows(
    payload: dict[str, Any],
    *,
    existing_comments: dict[str, str],
) -> list[dict[str, Any]]:
    """Return Excel-ready row dictionaries for extracted QA candidates."""

    qa_candidates = sorted(
        _list_of_dicts(payload.get("qa_candidates")),
        key=_qa_sort_key,
    )
    rows: list[dict[str, Any]] = []

    for candidate in qa_candidates:
        metadata = _dict_value(candidate.get("metadata"))
        question_debug = _dict_value(metadata.get("question_debug"))
        answer_debug = _dict_value(metadata.get("answer_debug"))
        pairing_debug = _dict_value(metadata.get("pairing_debug"))
        grounding_debug = _dict_value(metadata.get("grounding_debug"))
        confidence_debug = _dict_value(metadata.get("confidence_debug"))
        context_debug = _dict_value(metadata.get("context_debug"))
        search_signals = _dict_value(answer_debug.get("search_signals"))
        ranking_debug = _dict_value(answer_debug.get("ranking_debug"))

        confidence_score = _safe_float(
            candidate.get("confidence_score"),
            fallback=_safe_float(candidate.get("confidence")),
        )
        review_flags = _build_qa_review_flags(
            candidate=candidate,
            metadata=metadata,
            confidence_score=confidence_score,
        )
        review_priority = _qa_review_priority(
            candidate=candidate,
            review_flags=review_flags,
            confidence_score=confidence_score,
        )
        qa_id = str(candidate.get("qa_candidate_id") or "")
        question_sentence_ids = _coerce_string_list(candidate.get("question_sentence_ids"))
        answer_sentence_ids = _coerce_string_list(candidate.get("answer_sentence_ids"))
        context_sentence_ids = _coerce_string_list(
            candidate.get("context_sentence_ids"),
            fallback=_coerce_string_list(context_debug.get("context_sentence_ids")),
        )
        context_source_utterance_ids = _coerce_string_list(
            candidate.get("context_source_utterance_ids"),
            fallback=_coerce_string_list(
                context_debug.get("context_source_utterance_ids"),
            ),
        )
        context_text = str(
            candidate.get("context_text")
            or context_debug.get("context_text")
            or context_debug.get("context_raw_text")
            or "",
        )
        context_strategy = str(
            candidate.get("context_strategy")
            or context_debug.get("context_strategy")
            or "",
        )
        context_confidence = str(
            candidate.get("context_confidence")
            or context_debug.get("context_confidence")
            or "",
        )

        rows.append(
            {
                "qa_id": qa_id,
                "question_sentence_id": _human_sentence_reference(question_sentence_ids),
                "answer_sentence_id": _human_sentence_reference(answer_sentence_ids),
                "context_sentence_id": _human_sentence_reference(context_sentence_ids),
                "question_text": str(candidate.get("question_text") or ""),
                "answer_text": str(candidate.get("answer_text") or ""),
                "context_text": context_text,
                "question_sentence_ids": _join_values(question_sentence_ids),
                "answer_sentence_ids": _join_values(answer_sentence_ids),
                "context_sentence_ids": _join_values(context_sentence_ids),
                "context_strategy": context_strategy,
                "context_confidence": context_confidence,
                "confidence": _format_qa_confidence(
                    confidence_label=str(candidate.get("confidence_label") or ""),
                    confidence_score=confidence_score,
                ),
                "review_flag_summary": _summarize_review_flags(review_flags),
                "review_flags": ", ".join(review_flags),
                "human_comment": existing_comments.get(qa_id, ""),
                "id": qa_id,
                "start": _excel_duration(_qa_time(candidate, "start")),
                "end": _excel_duration(_qa_time(candidate, "end")),
                "duration_s": _excel_duration(_qa_duration(candidate)),
                "question_type": str(candidate.get("question_type") or ""),
                "didactic_question_score": _safe_float(
                    candidate.get("didactic_question_score"),
                    fallback=_safe_float(
                        confidence_debug.get("didactic_question_score"),
                    ),
                ),
                "confidence_score": confidence_score,
                "confidence_label": str(candidate.get("confidence_label") or ""),
                "question_score": _safe_float(
                    confidence_debug.get("question_score"),
                    fallback=_safe_float(question_debug.get("question_score")),
                ),
                "answer_score": _safe_float(
                    confidence_debug.get("answer_score"),
                    fallback=_safe_float(answer_debug.get("answer_score")),
                ),
                "base_confidence": _safe_float(confidence_debug.get("base_confidence")),
                "has_answer": bool(str(candidate.get("answer_text") or "").strip()),
                "answer_is_question": bool(candidate.get("answer_is_question")),
                "deferred_answer_search_used": bool(
                    pairing_debug.get("deferred_answer_search_used"),
                ),
                "review_priority": review_priority,
                "input_layer": str(
                    metadata.get("input_layer")
                    or question_debug.get("input_layer")
                    or "",
                ),
                "question_speaker_role": str(
                    candidate.get("question_speaker_role") or "",
                ),
                "answer_speaker_role": str(
                    candidate.get("answer_speaker_role") or "",
                ),
                "question_segment_id": str(
                    candidate.get("question_segment_id") or "",
                ),
                "answer_segment_id": str(
                    candidate.get("answer_segment_id") or "",
                ),
                "segment_relation": str(pairing_debug.get("segment_relation") or ""),
                "requested_search_strategy": str(
                    pairing_debug.get("requested_search_strategy") or "",
                ),
                "effective_search_strategy": str(
                    pairing_debug.get("effective_search_strategy")
                    or pairing_debug.get("search_strategy")
                    or "",
                ),
                "requested_ranking_strategy": str(
                    pairing_debug.get("requested_ranking_strategy") or "",
                ),
                "effective_ranking_strategy": str(
                    pairing_debug.get("effective_ranking_strategy")
                    or pairing_debug.get("ranking_strategy")
                    or "",
                ),
                "search_stop_reason": str(
                    pairing_debug.get("search_stop_reason") or "",
                ),
                "search_fallback_reason": str(
                    pairing_debug.get("search_fallback_reason") or "",
                ),
                "search_backend_error": str(
                    pairing_debug.get("search_backend_error") or "",
                ),
                "ranking_fallback_reason": str(
                    pairing_debug.get("ranking_fallback_reason") or "",
                ),
                "search_backend_status": str(
                    pairing_debug.get("semantic_backend_status") or "",
                ),
                "search_model_name": str(
                    pairing_debug.get("semantic_model_name") or "",
                ),
                "search_backend": str(pairing_debug.get("semantic_backend") or ""),
                "semantic_similarity": _safe_float(
                    search_signals.get("semantic_similarity"),
                ),
                "candidate_channel": str(
                    search_signals.get("candidate_channel") or "",
                ),
                "reranking_model_name": str(
                    pairing_debug.get("semantic_reranking_model_name")
                    or ranking_debug.get("semantic_reranking_model_name")
                    or "",
                ),
                "reranking_backend": str(
                    pairing_debug.get("semantic_reranking_backend")
                    or ranking_debug.get("semantic_reranking_backend")
                    or "",
                ),
                "semantic_relevance_score": _safe_float(
                    pairing_debug.get("semantic_relevance_score"),
                    fallback=_safe_float(
                        ranking_debug.get("semantic_relevance_score"),
                    ),
                ),
                "combined_rank_score": _safe_float(
                    ranking_debug.get("combined_score"),
                    fallback=_safe_float(
                        answer_debug.get("partial_scores", {}).get(
                            "combined_rank_score",
                        ),
                    ),
                ),
                "rank_position": _safe_float(ranking_debug.get("rank_position")),
                "speaker_influence": str(
                    pairing_debug.get("speaker_influence") or "",
                ),
                "question_timing_source": str(
                    grounding_debug.get("question_timing_source") or "",
                ),
                "answer_timing_source": str(
                    grounding_debug.get("answer_timing_source") or "",
                ),
                "question_source_utterance_ids": _join_values(
                    candidate.get("question_source_utterance_ids"),
                ),
                "answer_source_utterance_ids": _join_values(
                    candidate.get("answer_source_utterance_ids"),
                ),
                "context_source_utterance_ids": _join_values(
                    context_source_utterance_ids,
                ),
                "question_unit_ids": _join_values(candidate.get("question_unit_ids")),
                "answer_unit_ids": _join_values(candidate.get("answer_unit_ids")),
                "source_segment_ids": _join_values(candidate.get("source_segment_ids")),
                "reason_codes": _join_values(candidate.get("reason_codes")),
                "notes": _build_qa_notes(
                    candidate=candidate,
                    metadata=metadata,
                    question_debug=question_debug,
                    answer_debug=answer_debug,
                    pairing_debug=pairing_debug,
                    grounding_debug=grounding_debug,
                    confidence_debug=confidence_debug,
                    context_debug=context_debug,
                ),
            },
        )

    return rows


def _build_sentence_metric_rows(
    sentence_rows: list[dict[str, Any]],
    *,
    provenance_validation: SentenceProvenanceValidation,
) -> list[dict[str, Any]]:
    """Return aggregate sentence metrics optimized for run-to-run comparison."""

    sentence_count = len(sentence_rows)
    word_counts = [
        int(row["word_count"])
        for row in sentence_rows
        if row.get("word_count") is not None
    ]
    duration_values = [
        _timedelta_seconds(row.get("duration_s"))
        for row in sentence_rows
        if _timedelta_seconds(row.get("duration_s")) is not None
    ]

    sentence_with_speaker_count = sum(
        1 for row in sentence_rows if str(row.get("speaker") or "").strip()
    )
    sentence_without_speaker_count = sentence_count - sentence_with_speaker_count
    multi_utterance_sentence_count = sum(
        1 for row in sentence_rows if bool(row.get("is_multi_utterance"))
    )
    single_utterance_sentence_count = sentence_count - multi_utterance_sentence_count
    sentences_with_uncertain_source_count = sum(
        1 for row in sentence_rows if bool(row.get("has_uncertain_source"))
    )
    sentences_with_unassigned_source_count = sum(
        1 for row in sentence_rows if bool(row.get("has_unassigned_source"))
    )
    sentences_with_speaker_change_inside_count = sum(
        1 for row in sentence_rows if bool(row.get("has_speaker_change_inside"))
    )
    sentences_with_provenance_overlap_count = (
        provenance_validation.sentences_with_provenance_overlap_count
    )
    semantic_quality_counts = _count_by_label(
        [
            str(row.get("semantic_quality_label") or "").strip()
            for row in sentence_rows
            if str(row.get("semantic_quality_label") or "").strip()
        ],
    )
    speaker_resolution_counts = _count_by_label(
        [
            str(row.get("speaker_resolution_status") or "").strip()
            for row in sentence_rows
            if str(row.get("speaker_resolution_status") or "").strip()
        ],
    )
    speaker_stability_counts = _count_by_label(
        [
            str(row.get("speaker_stability_label") or "").strip()
            for row in sentence_rows
            if str(row.get("speaker_stability_label") or "").strip()
        ],
    )
    speaker_assignment_method_counts = _count_by_label(
        [
            str(row.get("speaker_assignment_method") or "").strip()
            for row in sentence_rows
            if str(row.get("speaker_assignment_method") or "").strip()
        ],
    )
    speaker_instability_pressure = sum(
        1
        for row in sentence_rows
        if (
            bool(row.get("has_speaker_change_inside"))
            or bool(row.get("has_uncertain_source"))
            or bool(row.get("has_unassigned_source"))
        )
    )
    semantic_risk_pressure = sum(
        1
        for row in sentence_rows
        if str(row.get("semantic_quality_label") or "").strip()
        in {"fragment", "run_on", "borderline"}
    )
    avg_utterances_per_sentence = _safe_average(
        [
            _safe_float(row.get("source_utterance_count"))
            for row in sentence_rows
            if _safe_float(row.get("source_utterance_count")) is not None
        ],
    )

    rows: list[dict[str, Any]] = []

    def add_metric(
        metric_name: str,
        metric_value: Any,
        metric_unit: str,
        metric_group: str,
        *,
        notes: str = "",
        value_kind: str = "float",
    ) -> None:
        rows.append(
            {
                "metric_name": metric_name,
                "metric_value": metric_value,
                "metric_unit": metric_unit,
                "metric_group": metric_group,
                "notes": notes,
                "value_kind": value_kind,
            },
        )

    add_metric("sentence_count", sentence_count, "count", "volume", value_kind="int")
    add_metric(
        "sentence_with_speaker_count",
        sentence_with_speaker_count,
        "count",
        "volume",
        value_kind="int",
    )
    add_metric(
        "sentence_without_speaker_count",
        sentence_without_speaker_count,
        "count",
        "volume",
        value_kind="int",
    )
    add_metric(
        "multi_utterance_sentence_count",
        multi_utterance_sentence_count,
        "count",
        "volume",
        value_kind="int",
    )
    add_metric(
        "single_utterance_sentence_count",
        single_utterance_sentence_count,
        "count",
        "volume",
        value_kind="int",
    )
    add_metric(
        "sentence_with_speaker_pct",
        _safe_ratio(sentence_with_speaker_count, sentence_count),
        "pct",
        "volume",
        value_kind="percent",
    )
    add_metric(
        "sentence_without_speaker_pct",
        _safe_ratio(sentence_without_speaker_count, sentence_count),
        "pct",
        "volume",
        value_kind="percent",
    )
    add_metric(
        "multi_utterance_sentence_pct",
        _safe_ratio(multi_utterance_sentence_count, sentence_count),
        "pct",
        "volume",
        value_kind="percent",
    )
    add_metric(
        "single_utterance_sentence_pct",
        _safe_ratio(single_utterance_sentence_count, sentence_count),
        "pct",
        "volume",
        value_kind="percent",
    )

    add_metric(
        "sentence_total_word_count",
        sum(word_counts),
        "words",
        "word_stats",
        value_kind="int",
    )
    add_metric(
        "sentence_avg_word_count",
        _safe_average(word_counts),
        "words",
        "word_stats",
    )
    add_metric(
        "sentence_median_word_count",
        _median(word_counts),
        "words",
        "word_stats",
    )
    add_metric(
        "sentence_min_word_count",
        min(word_counts) if word_counts else None,
        "words",
        "word_stats",
        value_kind="int",
    )
    add_metric(
        "sentence_max_word_count",
        max(word_counts) if word_counts else None,
        "words",
        "word_stats",
        value_kind="int",
    )
    add_metric(
        "sentence_p25_word_count",
        _percentile(word_counts, 25),
        "words",
        "word_stats",
    )
    add_metric(
        "sentence_p75_word_count",
        _percentile(word_counts, 75),
        "words",
        "word_stats",
    )
    add_metric(
        "sentence_p90_word_count",
        _percentile(word_counts, 90),
        "words",
        "word_stats",
    )

    add_metric(
        "sentence_total_duration_seconds",
        _excel_duration(sum(duration_values) if duration_values else None),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )
    add_metric(
        "sentence_avg_duration_seconds",
        _excel_duration(_safe_average(duration_values)),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )
    add_metric(
        "sentence_median_duration_seconds",
        _excel_duration(_median(duration_values)),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )
    add_metric(
        "sentence_min_duration_seconds",
        _excel_duration(min(duration_values) if duration_values else None),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )
    add_metric(
        "sentence_max_duration_seconds",
        _excel_duration(max(duration_values) if duration_values else None),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )
    add_metric(
        "sentence_p25_duration_seconds",
        _excel_duration(_percentile(duration_values, 25)),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )
    add_metric(
        "sentence_p75_duration_seconds",
        _excel_duration(_percentile(duration_values, 75)),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )
    add_metric(
        "sentence_p90_duration_seconds",
        _excel_duration(_percentile(duration_values, 90)),
        "seconds",
        "duration_stats",
        value_kind="duration",
        notes="Displayed as Excel duration with millisecond precision.",
    )

    rows.extend(
        _build_bucket_metric_rows(
            values=word_counts,
            bucket_definitions=WORD_COUNT_BUCKETS,
            metric_prefix="sentence_word_count",
            metric_group="word_buckets",
        ),
    )
    rows.extend(
        _build_bucket_metric_rows(
            values=duration_values,
            bucket_definitions=DURATION_BUCKETS,
            metric_prefix="sentence_duration",
            metric_group="duration_buckets",
        ),
    )

    add_metric(
        "sentences_with_speaker_count",
        sentence_with_speaker_count,
        "count",
        "speaker",
        value_kind="int",
    )
    add_metric(
        "sentences_with_speaker_pct",
        _safe_ratio(sentence_with_speaker_count, sentence_count),
        "pct",
        "speaker",
        value_kind="percent",
    )
    add_metric(
        "sentences_with_uncertain_source_count",
        sentences_with_uncertain_source_count,
        "count",
        "speaker",
        value_kind="int",
    )
    add_metric(
        "sentences_with_uncertain_source_pct",
        _safe_ratio(sentences_with_uncertain_source_count, sentence_count),
        "pct",
        "speaker",
        value_kind="percent",
    )
    add_metric(
        "sentences_with_unassigned_source_count",
        sentences_with_unassigned_source_count,
        "count",
        "speaker",
        value_kind="int",
    )
    add_metric(
        "sentences_with_unassigned_source_pct",
        _safe_ratio(sentences_with_unassigned_source_count, sentence_count),
        "pct",
        "speaker",
        value_kind="percent",
    )
    add_metric(
        "sentences_with_speaker_change_inside_count",
        sentences_with_speaker_change_inside_count,
        "count",
        "speaker",
        value_kind="int",
    )
    add_metric(
        "sentences_with_speaker_change_inside_pct",
        _safe_ratio(sentences_with_speaker_change_inside_count, sentence_count),
        "pct",
        "speaker",
        value_kind="percent",
    )
    add_metric(
        "sentences_with_provenance_overlap_count",
        sentences_with_provenance_overlap_count,
        "count",
        "speaker",
        value_kind="int",
    )
    add_metric(
        "sentences_with_provenance_overlap_pct",
        _safe_ratio(sentences_with_provenance_overlap_count, sentence_count),
        "pct",
        "speaker",
        value_kind="percent",
    )
    add_metric(
        "sentences_multi_utterance_count",
        multi_utterance_sentence_count,
        "count",
        "speaker",
        value_kind="int",
    )
    add_metric(
        "sentences_multi_utterance_pct",
        _safe_ratio(multi_utterance_sentence_count, sentence_count),
        "pct",
        "speaker",
        value_kind="percent",
    )

    for label, count in sorted(speaker_resolution_counts.items()):
        metric_label = _metric_label_fragment(label)
        add_metric(
            f"speaker_resolution_status_{metric_label}_count",
            count,
            "count",
            "speaker",
            value_kind="int",
        )
        add_metric(
            f"speaker_resolution_status_{metric_label}_pct",
            _safe_ratio(count, sentence_count),
            "pct",
            "speaker",
            value_kind="percent",
        )

    for label, count in sorted(speaker_stability_counts.items()):
        metric_label = _metric_label_fragment(label)
        add_metric(
            f"speaker_stability_label_{metric_label}_count",
            count,
            "count",
            "speaker",
            value_kind="int",
        )
        add_metric(
            f"speaker_stability_label_{metric_label}_pct",
            _safe_ratio(count, sentence_count),
            "pct",
            "speaker",
            value_kind="percent",
        )

    for label, count in sorted(speaker_assignment_method_counts.items()):
        metric_label = _metric_label_fragment(label)
        add_metric(
            f"speaker_assignment_method_{metric_label}_count",
            count,
            "count",
            "speaker",
            value_kind="int",
        )
        add_metric(
            f"speaker_assignment_method_{metric_label}_pct",
            _safe_ratio(count, sentence_count),
            "pct",
            "speaker",
            value_kind="percent",
        )

    for label in ("good", "fragment", "borderline", "run_on"):
        count = semantic_quality_counts.get(label, 0)
        add_metric(
            f"semantic_quality_{label}_count",
            count,
            "count",
            "semantic",
            value_kind="int",
        )
        add_metric(
            f"semantic_quality_{label}_pct",
            _safe_ratio(count, sentence_count),
            "pct",
            "semantic",
            value_kind="percent",
        )

    short_sentence_pressure = _safe_ratio(
        _count_matching_values(word_counts, WORD_COUNT_BUCKETS[1][1]),
        sentence_count,
    )
    very_short_sentence_pressure = _safe_ratio(
        _count_matching_values(word_counts, WORD_COUNT_BUCKETS[0][1]),
        sentence_count,
    )
    long_sentence_pressure = _safe_ratio(
        _count_matching_values(word_counts, WORD_COUNT_BUCKETS[5][1]),
        sentence_count,
    )
    very_long_sentence_pressure = _safe_ratio(
        _count_matching_values(word_counts, WORD_COUNT_BUCKETS[6][1]),
        sentence_count,
    )
    speaker_instability_pressure_pct = _safe_ratio(
        speaker_instability_pressure,
        sentence_count,
    )
    semantic_risk_pressure_pct = _safe_ratio(
        semantic_risk_pressure,
        sentence_count,
    )

    add_metric(
        "short_sentence_pressure_pct",
        short_sentence_pressure,
        "pct",
        "risk",
        value_kind="percent",
        notes="Share of sentences with <= 3 words.",
    )
    add_metric(
        "very_short_sentence_pressure_pct",
        very_short_sentence_pressure,
        "pct",
        "risk",
        value_kind="percent",
        notes="Share of sentences with <= 2 words.",
    )
    add_metric(
        "long_sentence_pressure_pct",
        long_sentence_pressure,
        "pct",
        "risk",
        value_kind="percent",
        notes="Share of sentences with >= 25 words.",
    )
    add_metric(
        "very_long_sentence_pressure_pct",
        very_long_sentence_pressure,
        "pct",
        "risk",
        value_kind="percent",
        notes="Share of sentences with >= 35 words.",
    )
    add_metric(
        "speaker_instability_pressure_pct",
        speaker_instability_pressure_pct,
        "pct",
        "risk",
        value_kind="percent",
        notes="Share of sentences with speaker change inside or uncertain/unassigned sources.",
    )
    add_metric(
        "semantic_risk_pressure_pct",
        semantic_risk_pressure_pct,
        "pct",
        "risk",
        value_kind="percent",
        notes="Share of sentences labelled as fragment, borderline or run-on.",
    )

    add_metric(
        "avg_utterances_per_sentence",
        avg_utterances_per_sentence,
        "count",
        "derived",
    )
    add_metric(
        "sentences_per_segment",
        _safe_ratio(sentence_count, len({row.get('segment_id') for row in sentence_rows if row.get('segment_id')})),
        "count",
        "derived",
    )
    add_metric(
        "sentence_length_profile",
        _sentence_length_profile(short_sentence_pressure or 0.0, long_sentence_pressure or 0.0),
        "label",
        "profiles",
        value_kind="text",
    )
    add_metric(
        "speaker_stability_profile",
        _speaker_stability_profile(
            speaker_instability_pressure_pct or 0.0,
            _safe_ratio(sentence_with_speaker_count, sentence_count) or 0.0,
        ),
        "label",
        "profiles",
        value_kind="text",
    )
    add_metric(
        "sentence_debug_overview",
        _sentence_debug_overview(
            semantic_risk_pressure_pct or 0.0,
            speaker_instability_pressure_pct or 0.0,
        ),
        "label",
        "profiles",
        value_kind="text",
    )

    return rows


def _build_input_detail_rows(input_sources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return one summary row per original input source."""

    rows: list[dict[str, Any]] = []
    for index, source in enumerate(input_sources, start=1):
        file_name = str(
            source.get("original_filename")
            or Path(str(source.get("original_path") or "")).name,
        )
        rows.append(
            _summary_row(
                "input_file",
                f"file_{index}_duration",
                _excel_duration(_safe_float(source.get("duration_seconds"))),
                value_kind="duration",
                notes=file_name,
            ),
        )
    return rows


def _build_review_candidate_rows(
    utterance_rows: list[dict[str, Any]],
    sentence_rows: list[dict[str, Any]],
    *,
    existing_comments: dict[str, str],
) -> list[dict[str, Any]]:
    """Return compact review rows ordered by severity and time."""

    review_rows: list[dict[str, Any]] = []
    for row in utterance_rows:
        if row["review_priority"] == "low":
            continue
        review_rows.append(
            {
                "candidate_id": f"utterance:{row['id']}",
                "kind": "utterance",
                "review_reason": _build_review_reason(
                    quality=row["speaker_status"],
                    review_flags=str(row["review_flags"] or ""),
                ),
                "suggested_action": _suggest_review_action(
                    kind="utterance",
                    review_priority=str(row["review_priority"] or ""),
                    review_flags=str(row["review_flags"] or ""),
                ),
                "human_comment": existing_comments.get(
                    f"utterance:{row['id']}",
                    existing_comments.get(str(row["id"]), ""),
                ),
                "id": row["id"],
                "start": row["start"],
                "review_priority": row["review_priority"],
                "speaker": row["speaker"],
                "segment_id": row["segment_id"],
                "quality": row["speaker_status"],
                "review_flags": row["review_flags"],
                "text": row["text"],
                "notes": row["notes"],
            },
        )
    for row in sentence_rows:
        if row["review_priority"] == "low":
            continue
        review_rows.append(
            {
                "candidate_id": f"sentence:{row['id']}",
                "kind": "sentence",
                "review_reason": _build_review_reason(
                    quality=row["semantic_quality_label"],
                    review_flags=str(row["review_flags"] or ""),
                ),
                "suggested_action": _suggest_review_action(
                    kind="sentence",
                    review_priority=str(row["review_priority"] or ""),
                    review_flags=str(row["review_flags"] or ""),
                ),
                "human_comment": existing_comments.get(
                    f"sentence:{row['id']}",
                    existing_comments.get(str(row["id"]), ""),
                ),
                "id": row["id"],
                "start": row["start"],
                "review_priority": row["review_priority"],
                "speaker": row["speaker"],
                "segment_id": row["segment_id"],
                "quality": row["semantic_quality_label"],
                "review_flags": row["review_flags"],
                "text": row["text"],
                "notes": row["notes"],
            },
        )

    priority_order = {"high": 0, "medium": 1, "low": 2}
    return sorted(
        review_rows,
        key=lambda row: (
            priority_order.get(str(row["review_priority"]), 9),
            row["start"] if row["start"] is not None else timedelta.max,
            str(row["kind"]),
            str(row["id"]),
        ),
    )


def _qa_sort_key(candidate: dict[str, Any]) -> tuple[int, float, str]:
    """Return a deterministic ordering key for QA candidate rows."""

    start_seconds = _qa_time(candidate, "start")
    if start_seconds is None:
        return (1, 10**9, str(candidate.get("qa_candidate_id") or ""))
    return (0, start_seconds, str(candidate.get("qa_candidate_id") or ""))


def _build_timing_stage_rows(
    timing_stages: list[dict[str, Any]],
    *,
    total_pipeline_duration_seconds: float | None,
    total_input_duration_seconds: float | None,
) -> list[dict[str, Any]]:
    """Return one summary row per pipeline stage timing."""

    rows: list[dict[str, Any]] = []
    for stage in timing_stages:
        duration_seconds = _safe_float(stage.get("duration_seconds"))
        notes: list[str] = []

        status = str(stage.get("status") or "").strip()
        note = str(stage.get("note") or "").strip()
        if status:
            notes.append(f"status={status}")
        if note:
            notes.append(f"note={note}")
        if stage.get("used_cache"):
            notes.append("used_cache=true")
        if stage.get("used_existing_artifact"):
            notes.append("used_existing_artifact=true")
        if stage.get("forced_recompute"):
            notes.append("forced_recompute=true")
        started_at = str(stage.get("started_at") or "").strip()
        finished_at = str(
            stage.get("finished_at") or stage.get("ended_at") or "",
        ).strip()
        if started_at:
            notes.append(f"started_at={started_at}")
        if finished_at:
            notes.append(f"finished_at={finished_at}")

        stage_factor = _safe_ratio(duration_seconds, total_input_duration_seconds)
        if stage_factor is not None:
            notes.append(f"rtf={stage_factor:.2f}x")

        metadata = _dict_value(stage.get("metadata"))
        if metadata:
            notes.append(
                "metadata="
                + ", ".join(
                    f"{key}={metadata[key]}"
                    for key in sorted(metadata)
                ),
            )

        rows.append(
            _summary_row(
                "timing_stage",
                str(stage.get("stage_name") or ""),
                _excel_duration(duration_seconds),
                value_kind="duration",
                value2=_safe_ratio(duration_seconds, total_pipeline_duration_seconds),
                value2_kind="percent",
                notes=" | ".join(notes),
            ),
        )
    return rows


def _qa_time(
    candidate: dict[str, Any],
    boundary: str,
    fallback: float | None = None,
) -> float | None:
    """Return the best available timing bound for one QA candidate."""

    direct_value = _safe_float(candidate.get(f"{boundary}_seconds"))
    if direct_value is not None:
        return direct_value

    timing = _dict_value(
        candidate.get("question_timing")
        if boundary == "start"
        else candidate.get("answer_timing"),
    )
    session_value = _safe_float(timing.get(f"session_{boundary}_seconds"))
    if session_value is not None:
        return session_value
    return _safe_float(timing.get(f"{boundary}_seconds"), fallback=fallback)


def _qa_duration(candidate: dict[str, Any]) -> float | None:
    """Return the QA candidate duration in seconds when timing is available."""

    start_seconds = _qa_time(candidate, "start")
    end_seconds = _qa_time(candidate, "end")
    if start_seconds is None or end_seconds is None:
        return None
    return max(0.0, end_seconds - start_seconds)


def _build_qa_review_flags(
    *,
    candidate: dict[str, Any],
    metadata: dict[str, Any],
    confidence_score: float | None,
) -> list[str]:
    """Return compact review cues for one QA row."""

    flags: list[str] = []
    flags.extend(_coerce_string_list(candidate.get("review_flags")))
    reason_codes = _coerce_string_list(candidate.get("reason_codes"))
    input_layer = str(metadata.get("input_layer") or "").strip()
    has_answer = bool(str(candidate.get("answer_text") or "").strip())

    if not has_answer:
        flags.append("missing_answer")
    if confidence_score is not None and confidence_score < 0.45:
        flags.append("low_confidence")
    elif confidence_score is not None and confidence_score < 0.75:
        flags.append("medium_confidence")
    if input_layer and input_layer != "sentences":
        flags.append("fallback_input_layer")
    if "speaker_conflict_penalty" in reason_codes:
        flags.append("speaker_conflict")
    if "competing_question_nearby" in reason_codes:
        flags.append("competing_question")
    if "answer_missing_utterance_grounding" in reason_codes:
        flags.append("answer_grounding_missing")
    if "question_missing_utterance_grounding" in reason_codes:
        flags.append("question_grounding_missing")
    if "segment_unknown" in reason_codes:
        flags.append("segment_unknown")
    if bool(candidate.get("answer_is_question")):
        flags.append("answer_is_question")

    return _ordered_unique(flags)


def _qa_review_priority(
    *,
    candidate: dict[str, Any],
    review_flags: Sequence[str],
    confidence_score: float | None,
) -> str:
    """Return a review-priority label for one QA row."""

    if any(
        flag in {
            "missing_answer",
            "low_confidence",
            "fallback_input_layer",
            "speaker_conflict",
            "answer_grounding_missing",
            "answer_is_question",
        }
        for flag in review_flags
    ):
        return "high"
    if any(
        flag in {
            "medium_confidence",
            "competing_question",
            "question_grounding_missing",
            "segment_unknown",
        }
        for flag in review_flags
    ):
        return "medium"
    if confidence_score is not None and confidence_score >= 0.75:
        return "low"
    return "medium"


def _build_qa_notes(
    *,
    candidate: dict[str, Any],
    metadata: dict[str, Any],
    question_debug: dict[str, Any],
    answer_debug: dict[str, Any],
    pairing_debug: dict[str, Any],
    grounding_debug: dict[str, Any],
    confidence_debug: dict[str, Any],
    context_debug: dict[str, Any],
) -> str:
    """Return a compact notes string with the most relevant QA debug details."""

    notes: list[str] = []
    search_signals = _dict_value(answer_debug.get("search_signals"))
    ranking_debug = _dict_value(answer_debug.get("ranking_debug"))
    input_layer = str(
        metadata.get("input_layer") or question_debug.get("input_layer") or "",
    ).strip()
    if input_layer:
        notes.append(f"input_layer={input_layer}")

    question_score = _safe_float(
        confidence_debug.get("question_score"),
        fallback=_safe_float(question_debug.get("question_score")),
    )
    if question_score is not None:
        notes.append(f"question_score={question_score:.2f}")

    didactic_question_score = _safe_float(
        candidate.get("didactic_question_score"),
        fallback=_safe_float(confidence_debug.get("didactic_question_score")),
    )
    if didactic_question_score is not None:
        notes.append(f"didactic_question_score={didactic_question_score:.2f}")

    answer_score = _safe_float(
        confidence_debug.get("answer_score"),
        fallback=_safe_float(answer_debug.get("answer_score")),
    )
    if answer_score is not None:
        notes.append(f"answer_score={answer_score:.2f}")

    base_confidence = _safe_float(confidence_debug.get("base_confidence"))
    if base_confidence is not None:
        notes.append(f"base_confidence={base_confidence:.2f}")

    final_confidence = _safe_float(confidence_debug.get("final_confidence"))
    if final_confidence is not None:
        notes.append(f"final_confidence={final_confidence:.2f}")

    segment_relation = str(pairing_debug.get("segment_relation") or "").strip()
    if segment_relation:
        notes.append(f"segment_relation={segment_relation}")

    search_stop_reason = str(pairing_debug.get("search_stop_reason") or "").strip()
    if search_stop_reason:
        notes.append(f"search_stop={search_stop_reason}")

    requested_search_strategy = str(
        pairing_debug.get("requested_search_strategy") or "",
    ).strip()
    if requested_search_strategy:
        notes.append(f"requested_search={requested_search_strategy}")

    effective_search_strategy = str(
        pairing_debug.get("effective_search_strategy")
        or pairing_debug.get("search_strategy")
        or "",
    ).strip()
    if effective_search_strategy:
        notes.append(f"effective_search={effective_search_strategy}")

    requested_ranking_strategy = str(
        pairing_debug.get("requested_ranking_strategy")
        or ranking_debug.get("requested_ranking_strategy")
        or "",
    ).strip()
    if requested_ranking_strategy:
        notes.append(f"requested_ranking={requested_ranking_strategy}")

    effective_ranking_strategy = str(
        pairing_debug.get("effective_ranking_strategy")
        or pairing_debug.get("ranking_strategy")
        or ranking_debug.get("effective_ranking_strategy")
        or "",
    ).strip()
    if effective_ranking_strategy:
        notes.append(f"effective_ranking={effective_ranking_strategy}")

    search_fallback_reason = str(
        pairing_debug.get("search_fallback_reason") or "",
    ).strip()
    if search_fallback_reason:
        notes.append(f"search_fallback={search_fallback_reason}")

    search_backend_error = str(
        pairing_debug.get("search_backend_error") or "",
    ).strip()
    if search_backend_error:
        notes.append(f"search_backend_error={search_backend_error}")

    ranking_fallback_reason = str(
        pairing_debug.get("ranking_fallback_reason")
        or ranking_debug.get("ranking_fallback_reason")
        or "",
    ).strip()
    if ranking_fallback_reason:
        notes.append(f"ranking_fallback={ranking_fallback_reason}")

    search_model_name = str(pairing_debug.get("semantic_model_name") or "").strip()
    if search_model_name:
        notes.append(f"search_model={search_model_name}")

    reranking_model_name = str(
        pairing_debug.get("semantic_reranking_model_name")
        or ranking_debug.get("semantic_reranking_model_name")
        or "",
    ).strip()
    if reranking_model_name:
        notes.append(f"reranking_model={reranking_model_name}")

    semantic_similarity = _safe_float(search_signals.get("semantic_similarity"))
    if semantic_similarity is not None:
        notes.append(f"semantic_similarity={semantic_similarity:.2f}")

    semantic_relevance_score = _safe_float(
        pairing_debug.get("semantic_relevance_score"),
        fallback=_safe_float(ranking_debug.get("semantic_relevance_score")),
    )
    if semantic_relevance_score is not None:
        notes.append(f"semantic_relevance={semantic_relevance_score:.2f}")

    combined_rank_score = _safe_float(
        ranking_debug.get("combined_score"),
        fallback=_safe_float(
            _dict_value(answer_debug.get("partial_scores")).get("combined_rank_score"),
        ),
    )
    if combined_rank_score is not None:
        notes.append(f"combined_rank_score={combined_rank_score:.2f}")

    candidate_channel = str(search_signals.get("candidate_channel") or "").strip()
    if candidate_channel:
        notes.append(f"candidate_channel={candidate_channel}")

    speaker_influence = str(pairing_debug.get("speaker_influence") or "").strip()
    if speaker_influence:
        notes.append(f"speaker_influence={speaker_influence}")

    if bool(pairing_debug.get("deferred_answer_search_used")):
        notes.append("deferred_answer_search=true")

    if bool(candidate.get("answer_is_question")):
        notes.append("answer_is_question=true")

    context_strategy = str(
        candidate.get("context_strategy") or context_debug.get("context_strategy") or "",
    ).strip()
    if context_strategy:
        notes.append(f"context_strategy={context_strategy}")

    context_confidence = str(
        candidate.get("context_confidence")
        or context_debug.get("context_confidence")
        or "",
    ).strip()
    if context_confidence:
        notes.append(f"context_confidence={context_confidence}")

    context_sentence_ids = _coerce_string_list(
        candidate.get("context_sentence_ids"),
        fallback=_coerce_string_list(context_debug.get("context_sentence_ids")),
    )
    if context_sentence_ids:
        notes.append("context_sentence_ids=" + ", ".join(context_sentence_ids))

    question_timing_source = str(
        grounding_debug.get("question_timing_source") or "",
    ).strip()
    if question_timing_source:
        notes.append(f"question_timing_source={question_timing_source}")

    answer_timing_source = str(
        grounding_debug.get("answer_timing_source") or "",
    ).strip()
    if answer_timing_source:
        notes.append(f"answer_timing_source={answer_timing_source}")

    question_grounded = _coerce_string_list(
        grounding_debug.get("question_grounded_utterance_ids"),
    )
    if question_grounded:
        notes.append("question_grounded=" + ", ".join(question_grounded))

    answer_grounded = _coerce_string_list(
        grounding_debug.get("answer_grounded_utterance_ids"),
    )
    if answer_grounded:
        notes.append("answer_grounded=" + ", ".join(answer_grounded))

    return " | ".join(notes)


def _summary_row(
    section: str,
    metric: str,
    value: Any,
    *,
    value_kind: str = "text",
    value2: Any = None,
    value2_kind: str = "text",
    notes: str = "",
) -> dict[str, Any]:
    """Return one structured summary row with formatting hints."""

    return {
        "section": section,
        "metric": metric,
        "value": value,
        "value2": value2,
        "value_kind": value_kind,
        "value2_kind": value2_kind,
        "notes": notes,
    }


def _build_segment_lookup(segments: list[dict[str, Any]]) -> dict[str, list[str]]:
    """Return a chunk-to-segment lookup derived from the exported segments."""

    return _build_segment_lookup_for_field(
        segments,
        lookup_field="transcript_chunk_ids",
    )


def _build_segment_lookup_for_field(
    segments: list[dict[str, Any]],
    *,
    lookup_field: str,
) -> dict[str, list[str]]:
    """Return a generic lookup from one segment reference field to segment ids."""

    lookup: dict[str, list[str]] = {}
    for segment in segments:
        segment_id = str(segment.get("segment_id") or "").strip()
        if not segment_id:
            continue
        reference_ids = segment.get(lookup_field, [])
        if not isinstance(reference_ids, list):
            continue
        for reference_id in reference_ids:
            normalized_reference_id = str(reference_id or "").strip()
            if not normalized_reference_id:
                continue
            segment_ids = lookup.setdefault(normalized_reference_id, [])
            if segment_id not in segment_ids:
                segment_ids.append(segment_id)
    return lookup


def _build_sentence_lookup_for_utterance_ids(
    sentences: list[dict[str, Any]],
) -> dict[str, list[str]]:
    """Return an utterance-to-sentence lookup from serialized sentence data."""

    lookup: dict[str, list[str]] = {}
    for sentence in sentences:
        sentence_id = str(sentence.get("sentence_id") or "").strip()
        if not sentence_id:
            continue
        source_utterance_ids = sentence.get("source_utterance_ids", [])
        if not isinstance(source_utterance_ids, list):
            continue
        for utterance_id in source_utterance_ids:
            normalized_utterance_id = str(utterance_id or "").strip()
            if not normalized_utterance_id:
                continue
            sentence_ids = lookup.setdefault(normalized_utterance_id, [])
            if sentence_id not in sentence_ids:
                sentence_ids.append(sentence_id)
    return lookup


def _build_sentence_overlap_lookup(
    utterance_sentence_lookup: dict[str, list[str]],
) -> dict[str, list[str]]:
    """Return sentence-to-overlap mappings for provenance diagnostics."""

    overlap_lookup: dict[str, set[str]] = {}
    for sentence_ids in utterance_sentence_lookup.values():
        if len(sentence_ids) <= 1:
            continue
        for sentence_id in sentence_ids:
            overlap_lookup.setdefault(sentence_id, set()).update(
                other_sentence_id
                for other_sentence_id in sentence_ids
                if other_sentence_id != sentence_id
            )
    return {
        sentence_id: sorted(overlap_sentence_ids)
        for sentence_id, overlap_sentence_ids in overlap_lookup.items()
    }


def _build_utterance_lookup(
    utterances: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Return a compact utterance lookup keyed by utterance id."""

    lookup: dict[str, dict[str, Any]] = {}
    for utterance in utterances:
        scoped_key = scoped_utterance_key(
            utterance.get("audio_source_id"),
            utterance.get("utterance_id"),
        )
        if not scoped_key:
            continue
        lookup[scoped_key] = utterance
    return lookup


def _source_utterance_diagnostics(
    *,
    source_utterance_keys: Sequence[str],
    utterance_lookup: dict[str, dict[str, Any]],
    sentence: dict[str, Any],
) -> dict[str, Any]:
    """Return source-level speaker and provenance diagnostics for one sentence."""

    resolved_utterances = [
        utterance_lookup[source_utterance_key]
        for source_utterance_key in source_utterance_keys
        if source_utterance_key in utterance_lookup
    ]
    source_speaker_ids = _ordered_unique(
        str(utterance.get("speaker_id") or "").strip()
        for utterance in resolved_utterances
        if str(utterance.get("speaker_id") or "").strip()
    )
    assigned_source_utterance_count = sum(
        1
        for utterance in resolved_utterances
        if str(utterance.get("speaker_id") or "").strip()
        and not bool(utterance.get("speaker_is_uncertain"))
    )
    uncertain_source_utterance_count = sum(
        1 for utterance in resolved_utterances if bool(utterance.get("speaker_is_uncertain"))
    )
    unassigned_source_utterance_count = sum(
        1 for utterance in resolved_utterances if not str(utterance.get("speaker_id") or "").strip()
    )
    short_fragment_source_utterance_count = 0
    if not resolved_utterances:
        metadata = _dict_value(sentence.get("metadata"))
        speaker_evidence = _dict_value(metadata.get("speaker_evidence"))
        source_speaker_ids = _coerce_string_list(
            speaker_evidence.get("assigned_speakers") or metadata.get("source_speaker_ids"),
        )
        assigned_source_utterance_count = _safe_int(
            metadata.get("assigned_source_utterance_count"),
            fallback=_safe_int(speaker_evidence.get("stable_assigned_count"), fallback=0),
        ) or 0
        uncertain_source_utterance_count = _safe_int(
            metadata.get("uncertain_source_utterance_count"),
            fallback=_safe_int(speaker_evidence.get("uncertain_count"), fallback=0),
        ) or 0
        unassigned_source_utterance_count = _safe_int(
            metadata.get("unassigned_source_utterance_count"),
            fallback=_safe_int(speaker_evidence.get("unassigned_count"), fallback=0),
        ) or 0
        short_fragment_source_utterance_count = _safe_int(
            metadata.get("short_fragment_source_utterance_count"),
            fallback=_safe_int(speaker_evidence.get("short_fragment_count"), fallback=0),
        ) or 0
    else:
        metadata = _dict_value(sentence.get("metadata"))
        speaker_evidence = _dict_value(metadata.get("speaker_evidence"))
        short_fragment_source_utterance_count = _safe_int(
            metadata.get("short_fragment_source_utterance_count"),
            fallback=_safe_int(speaker_evidence.get("short_fragment_count"), fallback=0),
        ) or 0

    return {
        "source_speaker_ids": source_speaker_ids,
        "distinct_source_speaker_count": len(source_speaker_ids),
        "assigned_source_utterance_count": assigned_source_utterance_count,
        "uncertain_source_utterance_count": uncertain_source_utterance_count,
        "unassigned_source_utterance_count": unassigned_source_utterance_count,
        "short_fragment_source_utterance_count": short_fragment_source_utterance_count,
    }


def _utterance_sort_key(utterance: Any) -> tuple[int, float, str, float, str]:
    """Return a deterministic ordering key for utterance rows."""

    if not isinstance(utterance, dict):
        return (1, 10**9, "", 10**9, "")

    session_start_seconds = _safe_float(utterance.get("session_start_seconds"))
    if session_start_seconds is not None:
        primary_order = 0
        start_seconds = session_start_seconds
    else:
        primary_order = 1
        start_seconds = _safe_float(utterance.get("start_seconds"), fallback=10**9)

    return (
        primary_order,
        start_seconds,
        str(utterance.get("audio_source_id") or ""),
        _safe_float(utterance.get("end_seconds"), fallback=10**9),
        str(utterance.get("utterance_id") or ""),
    )


def _utterance_time(
    utterance: dict[str, Any],
    boundary: str,
    fallback: float | None = None,
) -> float | None:
    """Return session-relative utterance timing when available."""

    session_key = f"session_{boundary}_seconds"
    boundary_key = f"{boundary}_seconds"
    value = utterance.get(session_key, utterance.get(boundary_key, fallback))
    return _safe_float(value, fallback=fallback)


def _utterance_duration(utterance: dict[str, Any]) -> float | None:
    """Return the utterance duration in seconds when timing is available."""

    start_seconds = _utterance_time(utterance, "start")
    end_seconds = _utterance_time(utterance, "end")
    if start_seconds is None or end_seconds is None:
        return None
    return max(0.0, end_seconds - start_seconds)


def _sentence_sort_key(sentence: Any) -> tuple[int, float, str, float, str]:
    """Return a deterministic ordering key for sentence rows."""

    if not isinstance(sentence, dict):
        return (1, 10**9, "", 10**9, "")

    session_start_seconds = _safe_float(sentence.get("session_start_seconds"))
    if session_start_seconds is not None:
        primary_order = 0
        start_seconds = session_start_seconds
    else:
        primary_order = 1
        start_seconds = _safe_float(sentence.get("start_seconds"), fallback=10**9)

    return (
        primary_order,
        start_seconds,
        str(sentence.get("audio_source_id") or ""),
        _safe_float(sentence.get("end_seconds"), fallback=10**9),
        str(sentence.get("sentence_id") or ""),
    )


def _sentence_time(
    sentence: dict[str, Any],
    boundary: str,
    fallback: float | None = None,
) -> float | None:
    """Return session-relative sentence timing when available."""

    session_key = f"session_{boundary}_seconds"
    boundary_key = f"{boundary}_seconds"
    value = sentence.get(session_key, sentence.get(boundary_key, fallback))
    return _safe_float(value, fallback=fallback)


def _sentence_duration(sentence: dict[str, Any]) -> float | None:
    """Return the sentence duration in seconds when timing is available."""

    start_seconds = _sentence_time(sentence, "start")
    end_seconds = _sentence_time(sentence, "end")
    if start_seconds is None or end_seconds is None:
        return None
    return max(0.0, end_seconds - start_seconds)


def _excel_duration(value: float | None) -> timedelta | None:
    """Return an Excel-friendly duration value with millisecond precision."""

    if value is None:
        return None
    return timedelta(seconds=round(value, 3))


def _word_count(utterance: dict[str, Any]) -> int:
    """Return a robust utterance word count from source ids or plain text."""

    source_word_ids = utterance.get("source_word_ids", [])
    if isinstance(source_word_ids, list) and source_word_ids:
        return len(source_word_ids)
    return _text_word_count(utterance.get("text"))


def _resolve_speaker_status(utterance: dict[str, Any]) -> str:
    """Return the human-facing speaker status used by the Excel sheet."""

    if bool(utterance.get("speaker_is_uncertain")):
        return "uncertain"
    return str(utterance.get("speaker_attribution_status") or "").strip()


def _resolve_segment_ids(
    utterance: dict[str, Any],
    segment_lookup: dict[str, list[str]],
    *,
    sentence_segment_lookup: dict[str, list[str]],
    utterance_sentence_lookup: dict[str, list[str]],
) -> list[str]:
    """Return the higher-level segment ids linked to one utterance."""

    utterance_id = str(utterance.get("utterance_id") or "").strip()
    if utterance_id:
        sentence_ids = utterance_sentence_lookup.get(
            scoped_utterance_key(utterance.get("audio_source_id"), utterance_id),
            [],
        )
        segment_ids_from_sentence = _ordered_unique(
            segment_id
            for sentence_id in sentence_ids
            for segment_id in sentence_segment_lookup.get(sentence_id, [])
        )
        if segment_ids_from_sentence:
            return segment_ids_from_sentence

    transcript_chunk_id = str(utterance.get("transcript_chunk_id") or "").strip()
    if not transcript_chunk_id:
        return []
    return list(segment_lookup.get(transcript_chunk_id, []))


def _resolve_unit_segment_ids(
    *,
    unit_id: str,
    lookup: dict[str, list[str]],
) -> list[str]:
    """Return linked segment ids for one generic unit identifier."""

    if not unit_id:
        return []
    return list(lookup.get(unit_id, []))


def _build_review_flags(
    *,
    speaker: str,
    speaker_status: str,
    speaker_change: bool,
    segment_is_missing: bool,
    has_sentence_mapping_conflict: bool,
    utterance: dict[str, Any],
) -> list[str]:
    """Return compact review cues for manual QA."""

    flags: list[str] = []
    if speaker_status == "uncertain":
        flags.append("uncertain_speaker")
    elif not speaker:
        flags.append("speaker_unassigned")
    if speaker_change:
        flags.append("speaker_changed")
    if segment_is_missing:
        flags.append("segment_missing")
    if has_sentence_mapping_conflict:
        flags.append("sentence_mapping_conflict")
    if _is_audio_quality_degraded(utterance):
        flags.append("audio_quality_degraded")
    return flags


def _utterance_review_priority(*, review_flags: Sequence[str]) -> str:
    """Return a review-priority label for one utterance row."""

    if any(
        flag in {"speaker_unassigned", "segment_missing", "sentence_mapping_conflict"}
        for flag in review_flags
    ):
        return "high"
    if any(
        flag in {"uncertain_speaker", "speaker_changed", "audio_quality_degraded"}
        for flag in review_flags
    ):
        return "medium"
    return "low"


def _build_notes(
    utterance: dict[str, Any],
    segment_ids: list[str],
    *,
    utterance_sentence_lookup: dict[str, list[str]],
) -> str:
    """Return a compact notes string with the most relevant debug details."""

    notes: list[str] = []
    metadata = _dict_value(utterance.get("metadata"))
    speaker_attribution = _dict_value(metadata.get("speaker_attribution"))
    reason = str(speaker_attribution.get("reason") or "").strip()
    if reason:
        notes.append(f"speaker_reason={reason}")

    audio_quality = _dict_value(speaker_attribution.get("audio_quality"))
    degraded_reasons = audio_quality.get("degraded_reasons", [])
    if isinstance(degraded_reasons, list) and degraded_reasons:
        notes.append(
            "audio_quality=" + ", ".join(str(item) for item in degraded_reasons),
        )

    build_strategy = str(metadata.get("build_strategy") or "").strip()
    if build_strategy:
        notes.append(f"build_strategy={build_strategy}")

    utterance_id = str(utterance.get("utterance_id") or "").strip()
    linked_sentence_ids = utterance_sentence_lookup.get(
        scoped_utterance_key(utterance.get("audio_source_id"), utterance_id),
        [],
    )
    if linked_sentence_ids:
        notes.append("sentence_ids=" + ", ".join(linked_sentence_ids))
    if len(linked_sentence_ids) > 1:
        notes.append("sentence_mapping_conflict=True")
    if len(segment_ids) > 1:
        notes.append("segment_candidates=" + ", ".join(segment_ids))

    return " | ".join(notes)


def _build_sentence_review_flags(
    *,
    sentence: dict[str, Any],
    segment_is_missing: bool,
    has_provenance_overlap: bool,
) -> list[str]:
    """Return compact review cues for one sentence row."""

    flags = _coerce_string_list(sentence.get("sentence_review_flags"))
    if segment_is_missing:
        flags.append("segment_missing")
    if has_provenance_overlap:
        flags.append("provenance_overlap")
    return _ordered_unique(flags)


def _sentence_review_priority(
    *,
    sentence: dict[str, Any],
    review_flags: Sequence[str],
    segment_is_missing: bool,
) -> str:
    """Return a review-priority label for one sentence row."""

    explicit_priority = str(sentence.get("review_priority") or "").strip().lower()
    if explicit_priority in {"high", "medium", "low"}:
        if segment_is_missing and explicit_priority != "high":
            return "high"
        return explicit_priority

    if segment_is_missing or any(
        flag in {
            "unassigned_source",
            "speaker_change_inside",
            "merge_risky",
            "provenance_overlap",
        }
        for flag in review_flags
    ):
        return "high"
    if any(
        flag in {
            "uncertain_source",
            "semantic_fragment",
            "semantic_run_on",
            "merge_borderline",
            "semantic_borderline",
        }
        for flag in review_flags
    ):
        return "medium"
    return "low"


def _build_sentence_notes(
    sentence: dict[str, Any],
    segment_ids: list[str],
    *,
    source_diagnostics: dict[str, Any],
    provenance_overlap_sentence_ids: Sequence[str],
) -> str:
    """Return a compact notes string for one sentence row."""

    notes: list[str] = []
    metadata = _dict_value(sentence.get("metadata"))
    source_utterance_count = _safe_int(metadata.get("source_utterance_count"))
    if source_utterance_count is not None:
        notes.append(f"source_utterance_count={source_utterance_count}")
    source_speaker_ids = _coerce_string_list(source_diagnostics.get("source_speaker_ids"))
    if source_speaker_ids:
        notes.append("source_speakers=" + ", ".join(source_speaker_ids))
    speaker_resolution_status = str(sentence.get("speaker_resolution_status") or "").strip()
    if speaker_resolution_status:
        notes.append(f"speaker_resolution_status={speaker_resolution_status}")
    speaker_confidence_label = str(sentence.get("speaker_confidence_label") or "").strip()
    if speaker_confidence_label:
        notes.append(f"speaker_confidence={speaker_confidence_label}")
    speaker_assignment_method = str(
        sentence.get("speaker_assignment_method")
        or metadata.get("speaker_assignment_method")
        or "",
    ).strip()
    if speaker_assignment_method:
        notes.append(f"speaker_assignment_method={speaker_assignment_method}")
    dominant_weight = _safe_float(
        metadata.get("dominant_speaker_weight"),
        fallback=_safe_float(_dict_value(metadata.get("speaker_evidence")).get("dominant_weight")),
    )
    if dominant_weight is not None:
        notes.append(f"dominant_weight={dominant_weight:.2f}")
    dominance_margin = _safe_float(
        metadata.get("dominance_margin"),
        fallback=_safe_float(_dict_value(metadata.get("speaker_evidence")).get("dominance_margin")),
    )
    if dominance_margin is not None:
        notes.append(f"dominance_margin={dominance_margin:.2f}")
    speaker_evidence_summary = str(sentence.get("speaker_evidence_summary") or "").strip()
    if speaker_evidence_summary:
        notes.append(f"speaker_evidence={speaker_evidence_summary}")
    if "speaker_boundary_respected" in metadata:
        notes.append(
            "speaker_boundary_respected="
            f"{bool(metadata.get('speaker_boundary_respected'))}",
        )
    if provenance_overlap_sentence_ids:
        notes.append(
            "provenance_overlap_with="
            + ", ".join(str(item) for item in provenance_overlap_sentence_ids),
        )
    if len(segment_ids) > 1:
        notes.append("segment_candidates=" + ", ".join(segment_ids))
    return " | ".join(notes)


def _is_audio_quality_degraded(utterance: dict[str, Any]) -> bool:
    """Return whether one utterance carries degraded audio-quality metadata."""

    metadata = _dict_value(utterance.get("metadata"))
    speaker_attribution = _dict_value(metadata.get("speaker_attribution"))
    audio_quality = _dict_value(speaker_attribution.get("audio_quality"))
    return bool(audio_quality.get("is_degraded"))


def _format_sentence_span(sentence: dict[str, Any]) -> str:
    """Return a readable start-end span for source utterance indexes."""

    start_index = _safe_int(sentence.get("source_utterance_start_index"))
    end_index = _safe_int(sentence.get("source_utterance_end_index"))
    if start_index is None and end_index is None:
        return ""
    if start_index is None:
        return str(end_index)
    if end_index is None or end_index == start_index:
        return str(start_index)
    return f"{start_index}-{end_index}"


def _style_header_row(worksheet: Worksheet) -> None:
    """Apply a lightweight style to the Excel header row."""

    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = TOP_ALIGNMENT


def _style_core_headers(
    worksheet: Worksheet,
    *,
    core_headers: Sequence[str],
    editable_headers: Sequence[str] = (),
) -> None:
    """Highlight human-review columns without removing technical columns."""

    column_map = _column_letter_map(
        [cell.value for cell in worksheet[1] if cell.value is not None],
    )
    for header in core_headers:
        column_letter = column_map.get(header)
        if column_letter is None:
            continue
        worksheet[f"{column_letter}1"].fill = CORE_HEADER_FILL
    for header in editable_headers:
        column_letter = column_map.get(header)
        if column_letter is None:
            continue
        worksheet[f"{column_letter}1"].fill = EDITABLE_HEADER_FILL


def _set_column_widths(
    worksheet: Worksheet,
    *,
    header_order: Sequence[str],
    widths: dict[str, float],
) -> None:
    """Apply column widths using header names instead of fixed letters."""

    column_map = _column_letter_map(header_order)
    for header, column_letter in column_map.items():
        worksheet.column_dimensions[column_letter].width = widths.get(header, 18)


def _hide_technical_columns(
    worksheet: Worksheet,
    *,
    header_order: Sequence[str],
    visible_headers: Sequence[str],
) -> None:
    """Hide technical columns while keeping them available in the workbook."""

    visible_set = set(visible_headers)
    column_map = _column_letter_map(header_order)
    for header, column_letter in column_map.items():
        if header in visible_set:
            continue
        worksheet.column_dimensions[column_letter].hidden = True
        worksheet.column_dimensions[column_letter].outline_level = 1


def _apply_row_height(
    worksheet: Worksheet,
    row_index: int,
    *,
    cell_specs: Sequence[tuple[Any, int]],
    minimum_height: float = 32,
    maximum_height: float = 120,
) -> None:
    """Set a readable row height based on the wrapped cell contents."""

    estimated_lines = 1
    for value, target_width in cell_specs:
        text = str(value or "")
        if not text:
            continue
        line_count = text.count("\n") + 1
        wrapped_lines = max(1, round(len(text) / max(target_width, 1)))
        estimated_lines = max(estimated_lines, line_count, wrapped_lines)

    worksheet.row_dimensions[row_index].height = min(
        maximum_height,
        max(minimum_height, 16 * estimated_lines),
    )


def _format_utterances_sheet(worksheet: Worksheet) -> None:
    """Apply readability-oriented formatting to the utterances worksheet."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    column_map = _column_letter_map(HEADERS)

    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

    for row_index in range(2, worksheet.max_row + 1):
        for column_letter in (
            column_map["start"],
            column_map["end"],
            column_map["duration_s"],
        ):
            worksheet[f"{column_letter}{row_index}"].number_format = "[h]:mm:ss.000"
            worksheet[f"{column_letter}{row_index}"].alignment = TOP_ALIGNMENT

        text_cell = worksheet[f"{column_map['text']}{row_index}"]
        speaker_cell = worksheet[f"{column_map['speaker']}{row_index}"]
        review_priority_cell = worksheet[f"{column_map['review_priority']}{row_index}"]
        review_flags_cell = worksheet[f"{column_map['review_flags']}{row_index}"]
        speaker_status_cell = worksheet[f"{column_map['speaker_status']}{row_index}"]
        speaker_change_cell = worksheet[f"{column_map['is_speaker_changed']}{row_index}"]
        is_uncertain_cell = worksheet[f"{column_map['is_uncertain_speaker']}{row_index}"]

        text_cell.alignment = WRAP_ALIGNMENT
        worksheet[f"{column_map['notes']}{row_index}"].alignment = WRAP_ALIGNMENT

        for column_letter in (
            column_map["id"],
            column_map["speaker"],
            column_map["speaker_status"],
            column_map["is_uncertain_speaker"],
            column_map["is_unassigned_speaker"],
            column_map["is_speaker_changed"],
            column_map["is_segment_missing"],
            column_map["is_audio_quality_degraded"],
            column_map["review_priority"],
            column_map["segment_id"],
            column_map["final_sentence_id"],
            column_map["sentence_assignment_count"],
            column_map["has_sentence_mapping_conflict"],
            column_map["word_count"],
            column_map["review_flags"],
        ):
            worksheet[f"{column_letter}{row_index}"].alignment = TOP_ALIGNMENT

        if speaker_status_cell.value == "uncertain":
            speaker_status_cell.fill = UNCERTAIN_FILL
        if review_priority_cell.value in {"high", "medium"}:
            review_priority_cell.fill = REVIEW_FILL
        if review_flags_cell.value:
            review_flags_cell.fill = REVIEW_FILL

        speaker_fill = _speaker_fill_for_value(speaker_cell.value)
        if speaker_fill is not None:
            text_cell.fill = speaker_fill
            speaker_cell.fill = speaker_fill

        text_cell.font = _build_text_font(
            underline=bool(is_uncertain_cell.value),
            review_priority=review_priority_cell.value,
        )

        if bool(speaker_change_cell.value):
            _apply_row_border(
                worksheet,
                row_index,
                (
                    column_map["start"],
                    column_map["end"],
                    column_map["duration_s"],
                    column_map["speaker"],
                    column_map["text"],
                    column_map["review_flags"],
                ),
            )


def _format_summary_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Apply readability-oriented formatting to the run summary worksheet."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_letter, width in SUMMARY_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

    for row_index, row in enumerate(rows, start=2):
        worksheet[f"A{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"B{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"C{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"D{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"E{row_index}"].alignment = WRAP_ALIGNMENT
        _apply_number_format(worksheet[f"C{row_index}"], row["value_kind"])
        _apply_number_format(worksheet[f"D{row_index}"], row["value2_kind"])


def _format_sentences_sheet(worksheet: Worksheet) -> None:
    """Apply readability-oriented formatting to the sentences worksheet."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    column_map = _column_letter_map(SENTENCE_HEADERS)
    _style_core_headers(
        worksheet,
        core_headers=SENTENCE_CORE_HEADERS,
        editable_headers=("human_comment",),
    )
    _set_column_widths(
        worksheet,
        header_order=SENTENCE_HEADERS,
        widths=SENTENCE_COLUMN_WIDTHS,
    )
    _hide_technical_columns(
        worksheet,
        header_order=SENTENCE_HEADERS,
        visible_headers=SENTENCE_CORE_HEADERS,
    )

    for row_index in range(2, worksheet.max_row + 1):
        for column_letter in (
            column_map["start"],
            column_map["end"],
            column_map["duration_s"],
        ):
            worksheet[f"{column_letter}{row_index}"].number_format = "[h]:mm:ss.000"
            worksheet[f"{column_letter}{row_index}"].alignment = TOP_ALIGNMENT

        for column_letter in (
            column_map["sentence_id"],
            column_map["time_range"],
            column_map["id"],
            column_map["speaker"],
            column_map["detected_language"],
            column_map["source_utterance_count"],
            column_map["source_utterance_span"],
            column_map["is_multi_utterance"],
            column_map["has_uncertain_source"],
            column_map["has_unassigned_source"],
            column_map["has_speaker_change_inside"],
            column_map["is_semantic_fragment"],
            column_map["is_semantic_run_on"],
            column_map["is_merge_risky"],
            column_map["semantic_quality_label"],
            column_map["length_bucket"],
            column_map["duration_bucket"],
            column_map["speaker_stability_label"],
            column_map["merge_safety_label"],
            column_map["review_priority"],
            column_map["segment_id"],
            column_map["word_count"],
            column_map["distinct_source_speaker_count"],
            column_map["assigned_source_utterance_count"],
            column_map["uncertain_source_utterance_count"],
            column_map["unassigned_source_utterance_count"],
            column_map["speaker_resolution_status"],
            column_map["speaker_confidence_label"],
            column_map["has_provenance_overlap"],
            column_map["review_flags"],
            column_map["speaker_assignment_method"],
            column_map["dominant_speaker_weight"],
            column_map["second_speaker_weight"],
            column_map["dominance_margin"],
            column_map["dominant_speaker_share"],
            column_map["short_fragment_source_utterance_count"],
        ):
            worksheet[f"{column_letter}{row_index}"].alignment = TOP_ALIGNMENT

        text_cell = worksheet[f"{column_map['text']}{row_index}"]
        speaker_cell = worksheet[f"{column_map['speaker']}{row_index}"]
        qa_summary_cell = worksheet[f"{column_map['qa_debug_summary']}{row_index}"]
        human_comment_cell = worksheet[f"{column_map['human_comment']}{row_index}"]
        review_priority_cell = worksheet[f"{column_map['review_priority']}{row_index}"]
        review_flags_cell = worksheet[f"{column_map['review_flags']}{row_index}"]
        uncertain_source_cell = worksheet[f"{column_map['has_uncertain_source']}{row_index}"]
        speaker_resolution_cell = worksheet[
            f"{column_map['speaker_resolution_status']}{row_index}"
        ]
        speaker_change_cell = worksheet[
            f"{column_map['has_speaker_change_inside']}{row_index}"
        ]

        text_cell.alignment = WRAP_ALIGNMENT
        qa_summary_cell.alignment = WRAP_ALIGNMENT
        human_comment_cell.alignment = WRAP_ALIGNMENT
        human_comment_cell.fill = EDITABLE_CELL_FILL
        worksheet[f"{column_map['source_utterance_ids']}{row_index}"].alignment = WRAP_ALIGNMENT
        worksheet[f"{column_map['source_speaker_ids']}{row_index}"].alignment = WRAP_ALIGNMENT
        worksheet[
            f"{column_map['provenance_overlap_sentence_ids']}{row_index}"
        ].alignment = WRAP_ALIGNMENT
        worksheet[f"{column_map['notes']}{row_index}"].alignment = WRAP_ALIGNMENT

        if review_priority_cell.value in {"high", "medium"}:
            review_priority_cell.fill = REVIEW_FILL
        if review_flags_cell.value:
            review_flags_cell.fill = REVIEW_FILL

        speaker_fill = _speaker_fill_for_value(speaker_cell.value)
        if speaker_fill is not None:
            text_cell.fill = speaker_fill
            speaker_cell.fill = speaker_fill

        text_cell.font = _build_text_font(
            underline=(
                bool(uncertain_source_cell.value)
                or str(speaker_resolution_cell.value or "").strip() == "uncertain"
            ),
            review_priority=review_priority_cell.value,
        )

        if bool(speaker_change_cell.value):
            _apply_row_border(
                worksheet,
                row_index,
                (
                    column_map["time_range"],
                    column_map["start"],
                    column_map["end"],
                    column_map["duration_s"],
                    column_map["speaker"],
                    column_map["text"],
                    column_map["review_flags"],
                ),
            )

        _apply_row_height(
            worksheet,
            row_index,
            cell_specs=(
                (text_cell.value, 95),
                (qa_summary_cell.value, 40),
                (human_comment_cell.value, 52),
            ),
            minimum_height=36,
        )


def _format_qa_sheet(worksheet: Worksheet) -> None:
    """Apply readability-oriented formatting to the QA worksheet."""

    worksheet.freeze_panes = "I2"
    worksheet.auto_filter.ref = worksheet.dimensions
    column_map = _column_letter_map(QA_HEADERS)
    _style_core_headers(
        worksheet,
        core_headers=QA_CORE_HEADERS,
        editable_headers=("human_comment",),
    )
    _set_column_widths(
        worksheet,
        header_order=QA_HEADERS,
        widths=QA_COLUMN_WIDTHS,
    )
    _hide_technical_columns(
        worksheet,
        header_order=QA_HEADERS,
        visible_headers=QA_CORE_HEADERS,
    )

    for row_index in range(2, worksheet.max_row + 1):
        for column_letter in (
            column_map["start"],
            column_map["end"],
            column_map["duration_s"],
        ):
            worksheet[f"{column_letter}{row_index}"].number_format = "[h]:mm:ss.000"
            worksheet[f"{column_letter}{row_index}"].alignment = TOP_ALIGNMENT

        for column_letter in (
            column_map["qa_id"],
            column_map["question_sentence_id"],
            column_map["answer_sentence_id"],
            column_map["context_sentence_id"],
            column_map["confidence"],
            column_map["context_strategy"],
            column_map["context_confidence"],
            column_map["id"],
            column_map["question_type"],
            column_map["didactic_question_score"],
            column_map["confidence_score"],
            column_map["confidence_label"],
            column_map["question_score"],
            column_map["answer_score"],
            column_map["base_confidence"],
            column_map["has_answer"],
            column_map["answer_is_question"],
            column_map["deferred_answer_search_used"],
            column_map["review_priority"],
            column_map["input_layer"],
            column_map["question_speaker_role"],
            column_map["answer_speaker_role"],
            column_map["question_segment_id"],
            column_map["answer_segment_id"],
            column_map["segment_relation"],
            column_map["requested_search_strategy"],
            column_map["effective_search_strategy"],
            column_map["requested_ranking_strategy"],
            column_map["effective_ranking_strategy"],
            column_map["search_stop_reason"],
            column_map["search_fallback_reason"],
            column_map["ranking_fallback_reason"],
            column_map["search_backend_status"],
            column_map["search_model_name"],
            column_map["search_backend"],
            column_map["semantic_similarity"],
            column_map["candidate_channel"],
            column_map["reranking_model_name"],
            column_map["reranking_backend"],
            column_map["semantic_relevance_score"],
            column_map["combined_rank_score"],
            column_map["rank_position"],
            column_map["speaker_influence"],
            column_map["question_timing_source"],
            column_map["answer_timing_source"],
            column_map["review_flags"],
        ):
            worksheet[f"{column_letter}{row_index}"].alignment = TOP_ALIGNMENT

        for column_letter in (
            column_map["review_flag_summary"],
            column_map["human_comment"],
            column_map["question_text"],
            column_map["answer_text"],
            column_map["context_text"],
            column_map["search_backend_error"],
            column_map["question_sentence_ids"],
            column_map["answer_sentence_ids"],
            column_map["context_sentence_ids"],
            column_map["question_source_utterance_ids"],
            column_map["answer_source_utterance_ids"],
            column_map["context_source_utterance_ids"],
            column_map["question_unit_ids"],
            column_map["answer_unit_ids"],
            column_map["source_segment_ids"],
            column_map["reason_codes"],
            column_map["notes"],
        ):
            worksheet[f"{column_letter}{row_index}"].alignment = WRAP_ALIGNMENT

        _apply_number_format(
            worksheet[f"{column_map['confidence_score']}{row_index}"],
            "float",
        )
        _apply_number_format(
            worksheet[f"{column_map['question_score']}{row_index}"],
            "float",
        )
        _apply_number_format(
            worksheet[f"{column_map['answer_score']}{row_index}"],
            "float",
        )
        _apply_number_format(
            worksheet[f"{column_map['base_confidence']}{row_index}"],
            "float",
        )
        _apply_number_format(
            worksheet[f"{column_map['semantic_similarity']}{row_index}"],
            "float",
        )
        _apply_number_format(
            worksheet[f"{column_map['semantic_relevance_score']}{row_index}"],
            "float",
        )
        _apply_number_format(
            worksheet[f"{column_map['combined_rank_score']}{row_index}"],
            "float",
        )
        _apply_number_format(
            worksheet[f"{column_map['rank_position']}{row_index}"],
            "int",
        )

        question_text_cell = worksheet[f"{column_map['question_text']}{row_index}"]
        answer_text_cell = worksheet[f"{column_map['answer_text']}{row_index}"]
        context_text_cell = worksheet[f"{column_map['context_text']}{row_index}"]
        review_flag_summary_cell = worksheet[
            f"{column_map['review_flag_summary']}{row_index}"
        ]
        human_comment_cell = worksheet[f"{column_map['human_comment']}{row_index}"]
        review_priority_cell = worksheet[f"{column_map['review_priority']}{row_index}"]
        review_flags_cell = worksheet[f"{column_map['review_flags']}{row_index}"]
        confidence_label_cell = worksheet[f"{column_map['confidence_label']}{row_index}"]
        input_layer_cell = worksheet[f"{column_map['input_layer']}{row_index}"]
        question_role_cell = worksheet[f"{column_map['question_speaker_role']}{row_index}"]
        answer_role_cell = worksheet[f"{column_map['answer_speaker_role']}{row_index}"]
        review_flag_summary_cell.alignment = WRAP_ALIGNMENT
        human_comment_cell.alignment = WRAP_ALIGNMENT
        human_comment_cell.fill = EDITABLE_CELL_FILL

        if review_priority_cell.value in {"high", "medium"}:
            review_priority_cell.fill = REVIEW_FILL
        if review_flags_cell.value:
            review_flags_cell.fill = REVIEW_FILL
        if confidence_label_cell.value == "low":
            confidence_label_cell.fill = UNCERTAIN_FILL
        if input_layer_cell.value and input_layer_cell.value != "sentences":
            input_layer_cell.fill = UNCERTAIN_FILL

        question_fill = _speaker_fill_for_value(question_role_cell.value)
        answer_fill = _speaker_fill_for_value(answer_role_cell.value)
        if question_fill is not None:
            question_text_cell.fill = question_fill
            question_role_cell.fill = question_fill
        if answer_fill is not None:
            answer_text_cell.fill = answer_fill
            answer_role_cell.fill = answer_fill

        question_text_cell.font = _build_text_font(
            underline=False,
            review_priority=review_priority_cell.value,
        )
        answer_text_cell.font = _build_text_font(
            underline=not bool(worksheet[f"{column_map['has_answer']}{row_index}"].value),
            review_priority=review_priority_cell.value,
        )
        _apply_row_height(
            worksheet,
            row_index,
            cell_specs=(
                (question_text_cell.value, 70),
                (answer_text_cell.value, 70),
                (context_text_cell.value, 60),
                (review_flag_summary_cell.value, 32),
                (human_comment_cell.value, 48),
            ),
            minimum_height=34,
        )


def _format_review_candidates_sheet(worksheet: Worksheet) -> None:
    """Apply readability-oriented formatting to the review-candidate worksheet."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    column_map = _column_letter_map(REVIEW_CANDIDATE_HEADERS)
    _style_core_headers(
        worksheet,
        core_headers=REVIEW_CANDIDATE_CORE_HEADERS,
        editable_headers=("human_comment",),
    )
    _set_column_widths(
        worksheet,
        header_order=REVIEW_CANDIDATE_HEADERS,
        widths=REVIEW_CANDIDATE_COLUMN_WIDTHS,
    )
    _hide_technical_columns(
        worksheet,
        header_order=REVIEW_CANDIDATE_HEADERS,
        visible_headers=REVIEW_CANDIDATE_CORE_HEADERS,
    )

    for row_index in range(2, worksheet.max_row + 1):
        worksheet[f"{column_map['start']}{row_index}"].number_format = "[h]:mm:ss.000"
        for column_letter in (
            column_map["candidate_id"],
            column_map["kind"],
            column_map["review_reason"],
            column_map["suggested_action"],
            column_map["id"],
            column_map["start"],
            column_map["review_priority"],
            column_map["speaker"],
            column_map["segment_id"],
            column_map["quality"],
            column_map["review_flags"],
        ):
            worksheet[f"{column_letter}{row_index}"].alignment = TOP_ALIGNMENT
        text_cell = worksheet[f"{column_map['text']}{row_index}"]
        speaker_cell = worksheet[f"{column_map['speaker']}{row_index}"]
        review_reason_cell = worksheet[f"{column_map['review_reason']}{row_index}"]
        human_comment_cell = worksheet[f"{column_map['human_comment']}{row_index}"]
        review_priority_cell = worksheet[f"{column_map['review_priority']}{row_index}"]
        text_cell.alignment = WRAP_ALIGNMENT
        review_reason_cell.alignment = WRAP_ALIGNMENT
        human_comment_cell.alignment = WRAP_ALIGNMENT
        human_comment_cell.fill = EDITABLE_CELL_FILL
        worksheet[f"{column_map['notes']}{row_index}"].alignment = WRAP_ALIGNMENT
        if review_priority_cell.value in {"high", "medium"}:
            review_priority_cell.fill = REVIEW_FILL

        speaker_fill = _speaker_fill_for_value(speaker_cell.value)
        if speaker_fill is not None:
            text_cell.fill = speaker_fill
            speaker_cell.fill = speaker_fill

        text_cell.font = _build_text_font(
            underline=False,
            review_priority=review_priority_cell.value,
        )
        _apply_row_height(
            worksheet,
            row_index,
            cell_specs=(
                (text_cell.value, 80),
                (review_reason_cell.value, 26),
                (human_comment_cell.value, 50),
            ),
            minimum_height=34,
        )


def _format_sentence_metrics_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
) -> None:
    """Apply readability-oriented formatting to the sentence-metrics worksheet."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_letter, width in SENTENCE_METRIC_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

    for row_index, row in enumerate(rows, start=2):
        worksheet[f"A{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"B{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"C{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"D{row_index}"].alignment = TOP_ALIGNMENT
        worksheet[f"E{row_index}"].alignment = WRAP_ALIGNMENT
        _apply_number_format(worksheet[f"B{row_index}"], row["value_kind"])


def _apply_number_format(cell, value_kind: str) -> None:
    """Apply a coherent Excel number format based on the stored value kind."""

    if value_kind == "duration":
        cell.number_format = "[h]:mm:ss.000"
        return
    if value_kind == "percent":
        cell.number_format = "0.00%"
        return
    if value_kind == "ratio":
        cell.number_format = '0.00"x"'
        return
    if value_kind == "float":
        cell.number_format = "0.00"
        return
    if value_kind == "int":
        cell.number_format = "0"


def _column_letter_map(headers: Sequence[str]) -> dict[str, str]:
    """Return a header-to-column lookup for worksheet formatting helpers."""

    return {
        header: get_column_letter(index)
        for index, header in enumerate(headers, start=1)
    }


def _speaker_fill_for_value(speaker: Any) -> PatternFill | None:
    """Return a deterministic speaker fill or ``None`` for empty speakers."""

    normalized_speaker = str(speaker or "").strip()
    if not normalized_speaker:
        return None

    palette_index = (
        int(hashlib.sha1(normalized_speaker.encode("utf-8")).hexdigest()[:8], 16)
        % len(SPEAKER_FILL_COLORS)
    )
    return PatternFill(
        fill_type="solid",
        fgColor=SPEAKER_FILL_COLORS[palette_index],
    )


def _build_text_font(
    *,
    underline: bool,
    review_priority: Any,
) -> Font:
    """Return a text font tuned for fast spreadsheet review."""

    normalized_priority = str(review_priority or "").strip().lower()
    return Font(
        bold=normalized_priority == "high",
        italic=normalized_priority == "medium",
        underline="single" if underline else None,
    )


def _apply_row_border(
    worksheet: Worksheet,
    row_index: int,
    column_letters: Sequence[str],
) -> None:
    """Apply a subtle separator to cells that deserve extra review focus."""

    for column_letter in column_letters:
        worksheet[f"{column_letter}{row_index}"].border = SPEAKER_CHANGE_BORDER


def _build_bucket_metric_rows(
    *,
    values: list[float | int],
    bucket_definitions: tuple[tuple[str, Any], ...],
    metric_prefix: str,
    metric_group: str,
) -> list[dict[str, Any]]:
    """Return bucket count and percentage metrics for one numeric dimension."""

    rows: list[dict[str, Any]] = []
    total_count = len(values)
    for suffix, predicate in bucket_definitions:
        count = _count_matching_values(values, predicate)
        rows.append(
            {
                "metric_name": f"{metric_prefix}_{suffix}_count",
                "metric_value": count,
                "metric_unit": "count",
                "metric_group": metric_group,
                "notes": "",
                "value_kind": "int",
            },
        )
        rows.append(
            {
                "metric_name": f"{metric_prefix}_{suffix}_pct",
                "metric_value": _safe_ratio(count, total_count),
                "metric_unit": "pct",
                "metric_group": metric_group,
                "notes": "",
                "value_kind": "percent",
            },
        )
    return rows


def _count_matching_values(values: list[float | int], predicate) -> int:
    """Return the count of values that satisfy one centralized predicate."""

    return sum(1 for value in values if predicate(value))


def _count_by_label(labels: list[str]) -> dict[str, int]:
    """Return occurrence counts for one list of labels."""

    counts: dict[str, int] = {}
    for label in labels:
        counts[label] = counts.get(label, 0) + 1
    return counts


def _metric_label_fragment(label: str) -> str:
    """Return a metric-safe fragment derived from one human label."""

    normalized = str(label).strip().lower().replace(" ", "_").replace("-", "_")
    normalized = "".join(
        character
        for character in normalized
        if character.isalnum() or character == "_"
    ).strip("_")
    return normalized or "unknown"


def _median(values: list[float | int]) -> float | None:
    """Return the median for a non-empty numeric list."""

    if not values:
        return None
    return float(statistics.median(values))


def _percentile(values: list[float | int], percentile: float) -> float | None:
    """Return a simple linear-interpolated percentile for a non-empty list."""

    if not values:
        return None
    if percentile <= 0:
        return float(min(values))
    if percentile >= 100:
        return float(max(values))

    ordered_values = sorted(float(value) for value in values)
    last_index = len(ordered_values) - 1
    position = (percentile / 100.0) * last_index
    lower_index = int(position)
    upper_index = min(lower_index + 1, last_index)
    if lower_index == upper_index:
        return ordered_values[lower_index]
    lower_value = ordered_values[lower_index]
    upper_value = ordered_values[upper_index]
    fraction = position - lower_index
    return lower_value + ((upper_value - lower_value) * fraction)


def _sentence_length_profile(short_pressure: float, long_pressure: float) -> str:
    """Return a compact profile label for sentence length balance."""

    if short_pressure >= 0.35 and long_pressure >= 0.20:
        return "mixed"
    if short_pressure >= 0.40:
        return "short_heavy"
    if long_pressure >= 0.25:
        return "long_heavy"
    return "balanced"


def _speaker_stability_profile(
    instability_pressure: float,
    speaker_coverage: float,
) -> str:
    """Return a compact profile label for speaker stability."""

    if instability_pressure < 0.10 and speaker_coverage >= 0.80:
        return "stable"
    if instability_pressure < 0.30:
        return "mixed"
    return "fragile"


def _sentence_debug_overview(
    semantic_risk_pressure: float,
    speaker_instability_pressure: float,
) -> str:
    """Return a global overview label for quick run comparison."""

    if semantic_risk_pressure < 0.20 and speaker_instability_pressure < 0.15:
        return "good"
    if semantic_risk_pressure < 0.40 and speaker_instability_pressure < 0.30:
        return "borderline"
    return "review_needed"


def _timedelta_seconds(value: Any) -> float | None:
    """Return the total seconds for one Excel duration cell value."""

    if isinstance(value, timedelta):
        return float(value.total_seconds())
    return _safe_float(value)


def _sum_duration(items: list[dict[str, Any]]) -> float | None:
    """Return the sum of available duration values."""

    durations = [
        duration
        for item in items
        if (duration := _safe_float(item.get("duration_seconds"))) is not None
    ]
    if not durations:
        return None
    return round(sum(durations), 3)


def _text_word_count(value: Any) -> int:
    """Return a simple whitespace-based word count."""

    text = str(value or "").strip()
    if not text:
        return 0
    return len([token for token in text.split() if token])


def _join_values(values: Any) -> str:
    """Return a readable comma-separated string from a scalar or list value."""

    if isinstance(values, list):
        normalized_values = [str(value) for value in values if str(value).strip()]
        return ", ".join(normalized_values)
    return str(values or "")


def _join_notes(base_note: str, examples: Sequence[str]) -> str:
    """Return a compact summary note optionally enriched with examples."""

    normalized_examples = [str(example).strip() for example in examples if str(example).strip()]
    if not normalized_examples:
        return base_note
    return f"{base_note} Examples: {'; '.join(normalized_examples)}"


def _safe_float(value: Any, fallback: float | None = None) -> float | None:
    """Return a float value when conversion is possible."""

    if value is None:
        return fallback
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def _safe_int(value: Any, fallback: int | None = None) -> int | None:
    """Return an integer value when conversion is possible."""

    if value is None:
        return fallback
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def _coerce_string_list(
    value: Any,
    *,
    fallback: list[str] | None = None,
) -> list[str]:
    """Return a compact list of non-empty strings from arbitrary input."""

    if not isinstance(value, list):
        return list(fallback or [])
    return [
        str(item).strip()
        for item in value
        if str(item).strip()
    ]


def _split_delimited_values(value: Sequence[str] | str | Any) -> list[str]:
    """Return ordered non-empty values from either a list or a CSV-like string."""

    if isinstance(value, list):
        return _ordered_unique(value)
    normalized_value = str(value or "").strip()
    if not normalized_value:
        return []
    return _ordered_unique(part.strip() for part in normalized_value.split(","))


def _compact_identifier(value: str) -> str:
    """Return a shorter identifier for dense human-review columns."""

    normalized_value = str(value or "").strip()
    if not normalized_value:
        return ""
    tail = normalized_value.split("_")[-1]
    if tail.isdigit():
        return tail
    return normalized_value


def _ordered_unique(values) -> list[str]:
    """Return a compact ordered list without duplicates."""

    ordered: list[str] = []
    for value in values:
        normalized_value = str(value or "").strip()
        if not normalized_value or normalized_value in ordered:
            continue
        ordered.append(normalized_value)
    return ordered


def _unique_non_empty_strings(values: Sequence[str]) -> list[str]:
    """Return ordered non-empty strings without duplicates."""

    return _ordered_unique(values)


def _find_timing_stage(
    timing_stages: Sequence[dict[str, Any]],
    stage_name: str,
) -> dict[str, Any]:
    """Return the first timing stage matching the requested stage name."""

    normalized_stage_name = str(stage_name or "").strip()
    if not normalized_stage_name:
        return {}
    for stage in timing_stages:
        if str(stage.get("stage_name") or "").strip() == normalized_stage_name:
            return stage
    return {}


def _safe_ratio(
    numerator: float | int | None,
    denominator: float | int | None,
) -> float | None:
    """Return a safe ratio or `None` when it cannot be computed."""

    if numerator is None or denominator is None:
        return None
    try:
        normalized_denominator = float(denominator)
        if normalized_denominator == 0.0:
            return None
        return float(numerator) / normalized_denominator
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def _safe_average(values: list[float]) -> float | None:
    """Return the arithmetic mean for a non-empty numeric list."""

    if not values:
        return None
    return sum(values) / len(values)


def _per_minute(
    value: float | int | None,
    duration_seconds: float | None,
) -> float | None:
    """Return the normalized per-minute rate for one metric."""

    if duration_seconds is None:
        return None
    return _safe_ratio(value, duration_seconds / 60.0)


def _per_hour(
    value: float | int | None,
    duration_seconds: float | None,
) -> float | None:
    """Return the normalized per-hour rate for one metric."""

    if duration_seconds is None:
        return None
    return _safe_ratio(value, duration_seconds / 3600.0)


def _dict_value(value: Any) -> dict[str, Any]:
    """Return a dictionary value or an empty dictionary."""

    if isinstance(value, dict):
        return value
    return {}


def _list_of_dicts(value: Any) -> list[dict[str, Any]]:
    """Return only dictionary entries from a list-like value."""

    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]
